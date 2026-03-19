import re
import json
import sys
import shutil
import logging
import time
import socket
import threading
import http.server
import socketserver
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, TypedDict

import requests
import pandas as pd
from tqdm import tqdm


# ==========================================================
# CONFIGURAÇÕES
# ==========================================================

@dataclass
class Config:
    client_id:     str = "l7be8e85ac9f86487594c3f8c54442059b"
    client_secret: str = "850917fb7a524d7f967d57ff95061501"

    url_token:    str = "https://apis.fedex.com/oauth/token"
    url_tracking: str = "https://apis.fedex.com/track/v1/trackingnumbers"

    arquivo_awbs:          str = "awbs.xlsx"
    arquivo_historico:     str = "historico_status.xlsx"
    arquivo_ultimo_status: str = "ultimo_status_gerado.xlsx"
    arquivo_log:           str = "tracking.log"

    max_workers: int   = 5
    delay_entre: float = 0.5
    timeout:     int   = 30
    porta_servidor: int = 8888


# ==========================================================
# MODELS
# ==========================================================

class ResultadoAWB(TypedDict, total=False):
    AWB: str
    AWB_ORIGINAL: str
    CATEGORIA: str
    STATUS_FEDEX: str
    ULTIMO_EVENTO: str
    MOTIVO_CATEGORIA: str
    DATA_CONSULTA: str
    SUCESSO: bool
    DATA_CHEGADA_MEMPHIS: str  # Data do primeiro scan em Memphis
    PEDIDO: str               # Número do pedido vinculado ao AWB
    PRODUTO: str              # Produto vinculado ao AWB
    PAIS_ORIGEM: str          # País de origem da remessa
    REGIAO: str               # Região geográfica de origem
    DATA_CRIACAO: str         # Data do primeiro evento registrado
    DATA_ENTREGA: str         # Data de entrega (se entregue)
    LEAD_TIME_DIAS: int       # Dias entre criação e entrega
    TIMELINE_JSON: str        # Timeline de etapas extraída dos scanEvents (JSON string)
    EVENTOS_JSON: str         # Todos os scanEvents brutos para painel de detalhe (JSON string)
    ENDERECO_ENTREGA: str     # Cidade/Estado/País de destino


# ==========================================================
# LOGGER
# ==========================================================

def setup_logger(log_file: str) -> logging.Logger:
    logger = logging.getLogger("fedex_tracker")
    if logger.handlers:
        return logger
    logger.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s [%(levelname)-8s] %(message)s", "%d/%m/%Y %H:%M:%S")
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger


# ==========================================================
# AWB — LIMPEZA E VALIDAÇÃO
# ==========================================================

def limpar_awb(awb: str) -> str:
    return re.sub(r'\s+', '', awb.strip())

def validar_awb(awb: str) -> bool:
    return bool(re.match(r'^\d{12,22}$', awb))


# ==========================================================
# CLASSIFICAÇÃO — 5 CATEGORIAS
# ==========================================================

HUBS_INTERNACIONAIS = [
    # EUA
    "MEMPHIS", "MEM", "INDIANAPOLIS", "IND", "MIAMI", "MIA",
    "NEWARK", "EWR", "CHICAGO", "ORD", "LOS ANGELES", "LAX",
    "AGUADILLA", "BQN",
    # Europa
    "FRANKFURT", "FRA", "PARIS", "CDG", "COLOGNE", "CGN",
    "LONDON", "LHR", "AMSTERDAM", "AMS", "LIEGE", "LGG",
    "ENFIELD",
    # Oriente Médio
    "DUBAI", "DXB",
    # Ásia
    "HONG KONG", "HKG", "SHANGHAI", "PVG", "TOKYO", "NRT",
    "SINGAPORE", "SIN", "MUMBAI", "BOM", "DELHI", "DEL",
    "BANGALORE", "BLR", "CHENNAI", "MAA", "GUANGZHOU", "CAN",
    "BEIJING", "PEK", "SEOUL", "ICN", "TAIPEI", "TPE",
    "BANGKOK", "BKK", "KUALA LUMPUR", "KUL",
    # América
    "TORONTO", "YYZ", "MEXICO CITY", "MEX",
]

TERMOS_BRASIL = [
    "CAMPINAS", "VCP", "VIRACOPOS", "BRAZIL", "BRASIL",
    "BARUERI", "SAO PAULO", "SAO PAULO", "RIO DE JANEIRO",
    "CURITIBA", "PORTO ALEGRE", "BELO HORIZONTE", "BRASILIA",
    "MANAUS", "RECIFE", "FORTALEZA", "SALVADOR", "BR",
]

TERMOS_ATRASO_ALFANDEGA = [
    # Termos em português (retornados pela API FedEx Brasil)
    "ATRASO NA LIBERACAO",
    "ATRASO NA LIBERAÇÃO",
    "ATRASO NA LIBERACAO - IMPORTACAO",
    "ATRASO NA LIBERAÇÃO - IMPORTAÇÃO",
    # Termos em inglês
    "CLEARANCE DELAY",
    "CUSTOMS DELAY",
    "HELD BY CUSTOMS",
    "CUSTOMS HOLD",
    "RETIDO NA ALFANDEGA",
    "RETIDO NA ALFÂNDEGA",
    "AGUARDANDO LIBERACAO",
    "AGUARDANDO LIBERAÇÃO",
    "PENDING CUSTOMS",
    "CLEARANCE IN PROGRESS",
    "IMPORT HOLD",
]

TERMOS_LIBERACAO = [
    "INTERNATIONAL SHIPMENT RELEASE",
    "RELEASED BY CUSTOMS",
    "CUSTOMS CLEARED",
    "CLEARANCE COMPLETED",
    # Termos em português retornados pela API FedEx Brasil
    "LIBERACAO DA REMESSA INTERNACIONAL",
    "LIBERAÇÃO DA REMESSA INTERNACIONAL",
    "REMESSA LIBERADA",
    "LIBERADO PELA ALFANDEGA",
    "LIBERADO PELA ALFÂNDEGA",
]

# ATENÇÃO: "IMPORT" foi removido propositalmente pois bate em "IMPORTACAO"
# causando classificação errada. A liberação real tem termos específicos acima.

TERMOS_OUT_FOR_DELIVERY = [
    "ON FEDEX VEHICLE FOR DELIVERY",
    "EM UM VEICULO DA FEDEX PARA ENTREGA",
    "EM UM VEÍCULO DA FEDEX PARA ENTREGA",
    "OUT FOR DELIVERY",
    "EM VEICULO PARA ENTREGA",
    "EM VEÍCULO PARA ENTREGA",
    "SAIU PARA ENTREGA",
    "WITH DELIVERY COURIER",
]

TERMOS_ENTREGUE = [
    "DELIVERED",
    "ENTREGUE",
    "DELIVERY MADE",
]


def _textos_eventos(eventos: list) -> list:
    textos = []
    for ev in eventos:
        textos.append(ev.get("eventDescription", "").upper())
        textos.append(ev.get("scanLocation", {}).get("city", "").upper())
        textos.append(ev.get("scanLocation", {}).get("stateOrProvinceCode", "").upper())
        textos.append(ev.get("scanLocation", {}).get("countryCode", "").upper())
    return textos


def _contem(textos: list, termos: list) -> bool:
    for texto in textos:
        for termo in termos:
            if termo in texto:
                return True
    return False


def classificar(eventos: list, status_fedex: str) -> tuple:
    """
    Classifica o AWB em uma das 5 categorias.

    REGRA FUNDAMENTAL: O STATUS ATUAL (status_fedex) e o ÚLTIMO EVENTO
    têm prioridade absoluta sobre o histórico anterior.

    Ordem de verificação:
    1. DELIVERED
    2. OUT FOR DELIVERY
    3. ATRASO NA LIBERACAO  → CUSTOMS INSPECTION  (status atual tem prioridade)
    4. LIBERACAO DA REMESSA → NATIONAL TRANSIT    (status atual tem prioridade)
    5. SAIU DA ORIGEM       → COMING TO BRAZIL    (último evento)
    6. Histórico completo como fallback
    """
    status_upper         = status_fedex.upper()
    ultimo_desc          = eventos[0].get("eventDescription", "").upper() if eventos else ""
    ultimo_local_city    = eventos[0].get("scanLocation", {}).get("city", "").upper() if eventos else ""
    ultimo_local_country = eventos[0].get("scanLocation", {}).get("countryCode", "").upper() if eventos else ""
    ultimo_textos        = [status_upper, ultimo_desc, ultimo_local_city, ultimo_local_country]

    # Histórico completo (para fallback)
    todos_textos = _textos_eventos(eventos)

    # ── 1. DELIVERED ─────────────────────────────────────────────────────────
    GATILHOS_ENTREGUE = ["DELIVERED", "ENTREGUE", "DELIVERY MADE"]
    if _contem(ultimo_textos, GATILHOS_ENTREGUE):
        return "DELIVERED", "Entrega confirmada no destino"

    # ── 2. OUT FOR DELIVERY ──────────────────────────────────────────────────
    GATILHOS_OFD = [
        "ON FEDEX VEHICLE FOR DELIVERY",
        "EM UM VEICULO DA FEDEX PARA ENTREGA",
        "EM UM VEÍCULO DA FEDEX PARA ENTREGA",
        "OUT FOR DELIVERY",
        "EM VEICULO PARA ENTREGA",
        "EM VEÍCULO PARA ENTREGA",
        "SAIU PARA ENTREGA",
        "WITH DELIVERY COURIER",
    ]
    if _contem(ultimo_textos, GATILHOS_OFD):
        return "OUT FOR DELIVERY", "Pacote saiu para entrega — chegada prevista hoje"

    # ── REGRA SÃO PAULO ──────────────────────────────────────────────────────
    # Remessa com último scan em São Paulo/SP = está na cidade de destino final
    # Obrigatoriamente em OUT FOR DELIVERY
    TERMOS_SAO_PAULO = ["SAO PAULO", "SÃO PAULO"]
    ultimo_city = eventos[0].get("scanLocation", {}).get("city", "").upper() if eventos else ""
    ultimo_state = eventos[0].get("scanLocation", {}).get("stateOrProvinceCode", "").upper() if eventos else ""
    if any(t in ultimo_city for t in TERMOS_SAO_PAULO) and ultimo_state in ("SP", ""):
        return "OUT FOR DELIVERY", "Remessa em São Paulo/SP — saindo para entrega"

    # ── 3. CUSTOMS INSPECTION (status atual = atraso) ────────────────────────
    # Verifica status_upper E último evento — cobre casos onde o status genérico
    # é "NA INSTALAÇÃO LOCAL DA FEDEX" mas o eventDescription é "ATRASO NA LIBERAÇÃO"
    if _contem(ultimo_textos, TERMOS_ATRASO_ALFANDEGA):
        return "CUSTOMS INSPECTION", "Atraso na liberação alfandegária — retido pela Receita Federal"

    # ── 4. NATIONAL TRANSIT (status atual = liberado) ────────────────────────
    # REGRA DUPLA OBRIGATÓRIA:
    # 1. Status atual deve ser um gatilho de liberação alfandegária
    # 2. O Brasil DEVE aparecer no histórico de eventos
    # Sem as duas condições, a remessa NÃO pode ser NATIONAL TRANSIT.
    GATILHOS_LIBERACAO = [
        "LIBERACAO DA REMESSA INTERNACIONAL",
        "LIBERAÇÃO DA REMESSA INTERNACIONAL",
        "INTERNATIONAL SHIPMENT RELEASE",
        "RELEASED BY CUSTOMS",
        "CUSTOMS CLEARED",
        "CLEARANCE COMPLETED",
        "REMESSA LIBERADA",
        "LIBERADO PELA ALFANDEGA",
        "LIBERADO PELA ALFÂNDEGA",
    ]
    tem_brasil_historico = _contem(todos_textos, TERMOS_BRASIL)
    if _contem([status_upper], GATILHOS_LIBERACAO) and tem_brasil_historico:
        return "NATIONAL TRANSIT", "Liberação alfandegária confirmada — em trânsito nacional"
    elif _contem([status_upper], GATILHOS_LIBERACAO) and not tem_brasil_historico:
        # Tem liberação mas sem Brasil no histórico — ainda em trânsito internacional
        return "COMING TO BRAZIL", "Liberação detectada mas Brasil não confirmado no histórico"

    # ── 5. COMING TO BRAZIL (último evento = saiu da origem) ─────────────────
    GATILHOS_ORIGEM = [
        "DEIXOU O LOCAL DE ORIGEM FEDEX",
        "LEFT FEDEX ORIGIN FACILITY",
        "DEPARTED FEDEX LOCATION",
        "SHIPMENT LEFT FEDEX ORIGIN",
    ]
    if _contem([status_upper, ultimo_desc], GATILHOS_ORIGEM):
        return "COMING TO BRAZIL", "Saiu do local de origem — em rota internacional para o Brasil"

    # ── REGRA CAMPINAS: pacote disponível para liberação sem liberação confirmada ──
    # "PACOTE DISPONÍVEL PARA LIBERAÇÃO" em Campinas/VCP = aguardando alfândega
    TERMOS_DISPONIVEL_LIBERACAO = [
        "PACOTE DISPONIVEL PARA LIBERACAO",
        "PACOTE DISPONÍVEL PARA LIBERAÇÃO",
        "PACKAGE AVAILABLE FOR CLEARANCE",
        "AVAILABLE FOR CLEARANCE",
    ]
    TERMOS_CAMPINAS = ["CAMPINAS", "VCP", "VIRACOPOS"]
    ultimo_city_raw = eventos[0].get("scanLocation", {}).get("city", "").upper() if eventos else ""
    em_campinas = any(t in ultimo_city_raw for t in TERMOS_CAMPINAS)
    disponivel_liberacao = _contem(ultimo_textos, TERMOS_DISPONIVEL_LIBERACAO)
    sem_liberacao_confirmada = not _contem(todos_textos, GATILHOS_LIBERACAO)
    if (em_campinas or disponivel_liberacao) and disponivel_liberacao and sem_liberacao_confirmada:
        return "CUSTOMS INSPECTION", "Pacote disponível para liberação em Campinas — aguardando alfândega"

    # ── 6. FALLBACK: analisa histórico completo ──────────────────────────────
    tem_brasil    = _contem(todos_textos, TERMOS_BRASIL)
    tem_liberacao = _contem(todos_textos, GATILHOS_LIBERACAO)
    tem_atraso    = _contem(todos_textos, TERMOS_ATRASO_ALFANDEGA)
    tem_hub       = _contem(todos_textos, HUBS_INTERNACIONAIS)
    tem_origem    = _contem(todos_textos, GATILHOS_ORIGEM)

    # VALIDAÇÃO CRÍTICA: o último evento está no Brasil?
    # Se o último evento é em hub internacional (Memphis, Mumbai, etc),
    # a remessa NÃO pode ser NATIONAL TRANSIT — ainda está fora do Brasil.
    ultimo_no_brasil = _contem(ultimo_textos, TERMOS_BRASIL)
    ultimo_em_hub    = _contem(ultimo_textos, HUBS_INTERNACIONAIS)

    # Brasil + atraso no histórico = CUSTOMS INSPECTION
    if tem_brasil and tem_atraso and not tem_liberacao and ultimo_no_brasil:
        return "CUSTOMS INSPECTION", "Carga no Brasil com atraso na liberação alfandegária"

    # Brasil + liberação + último evento no Brasil = NATIONAL TRANSIT
    if tem_brasil and tem_liberacao and ultimo_no_brasil:
        return "NATIONAL TRANSIT", "Liberado na alfândega — em trânsito nacional"

    # Tem liberação mas último evento fora do Brasil = ainda vindo
    if tem_liberacao and not ultimo_no_brasil:
        return "COMING TO BRAZIL", "Em trânsito internacional — ainda não chegou ao Brasil"

    # Brasil sem liberação e último evento no Brasil = CUSTOMS INSPECTION
    if tem_brasil and not tem_liberacao and ultimo_no_brasil:
        return "CUSTOMS INSPECTION", "Carga no Brasil aguardando liberação pela Receita Federal"

    # Hub internacional ou saiu da origem = COMING TO BRAZIL
    if tem_hub or tem_origem or ultimo_em_hub:
        return "COMING TO BRAZIL", "Em trânsito internacional — ainda não chegou ao Brasil"

    # Nenhum movimento expressivo = LABEL CREATED
    return "LABEL CREATED", "Etiqueta criada — aguardando coleta ou primeira movimentação"


# ==========================================================
# DIAS ÚTEIS — FERIADOS NACIONAIS BRASILEIROS
# ==========================================================

def dias_uteis_br(d1, d2) -> int:
    """
    Conta dias úteis entre duas datas (d1 inclusive, d2 exclusive),
    excluindo finais de semana e feriados nacionais fixos + móveis brasileiros.
    d1 e d2 podem ser date ou datetime.
    """
    from datetime import date as _date, timedelta as _td
    import math as _math

    if hasattr(d1, "date"): d1 = d1.date()
    if hasattr(d2, "date"): d2 = d2.date()
    if d1 >= d2:
        return 0

    def _pascoa(ano):
        # Algoritmo de Meeus/Jones/Butcher
        a = ano % 19
        b = ano // 100
        c = ano % 100
        d = b // 4
        e = b % 4
        f = (b + 8) // 25
        g = (b - f + 1) // 3
        h = (19 * a + b - d - g + 15) % 30
        i = c // 4
        k = c % 4
        l = (32 + 2 * e + 2 * i - h - k) % 7
        m = (a + 11 * h + 22 * l) // 451
        mes = (h + l - 7 * m + 114) // 31
        dia = ((h + l - 7 * m + 114) % 31) + 1
        return _date(ano, mes, dia)

    def _feriados(ano):
        pascoa = _pascoa(ano)
        td = _td
        feriados = {
            # Fixos nacionais
            _date(ano, 1,  1),   # Confraternização Universal
            _date(ano, 4, 21),   # Tiradentes
            _date(ano, 5,  1),   # Dia do Trabalho
            _date(ano, 9,  7),   # Independência
            _date(ano, 10, 12),  # N. Sra. Aparecida
            _date(ano, 11,  2),  # Finados
            _date(ano, 11, 15),  # Proclamação da República
            _date(ano, 11, 20),  # Consciência Negra (lei 14.759/2023)
            _date(ano, 12, 25),  # Natal
            # Móveis (base na Páscoa)
            pascoa - td(days=48),  # Segunda de Carnaval
            pascoa - td(days=47),  # Terça de Carnaval
            pascoa - td(days=2),   # Sexta-feira Santa
            pascoa,                # Páscoa
            pascoa + td(days=60),  # Corpus Christi
        }
        return feriados

    # Cache de feriados por ano para performance
    _cache = {}
    def _get_feriados(ano):
        if ano not in _cache:
            _cache[ano] = _feriados(ano)
        return _cache[ano]

    count = 0
    cur = d1
    while cur < d2:
        if cur.weekday() < 5 and cur not in _get_feriados(cur.year):
            count += 1
        cur += _td(days=1)
    return count

class TokenManager:
    def __init__(self, config: Config):
        self.config     = config
        self._token     = None
        self._expira_em = 0
        self.logger     = logging.getLogger("fedex_tracker")

    def get_token(self) -> str:
        if self._token and time.time() < self._expira_em - 60:
            return self._token

        self.logger.info("Obtendo token de acesso da API FedEx...")

        response = requests.post(
            self.config.url_token,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            data={
                "grant_type":    "client_credentials",
                "client_id":     self.config.client_id,
                "client_secret": self.config.client_secret,
            },
            timeout=self.config.timeout,
        )

        if response.status_code != 200:
            raise Exception(
                f"Erro ao obter token: HTTP {response.status_code} — {response.text[:200]}"
            )

        dados           = response.json()
        self._token     = dados["access_token"]
        self._expira_em = time.time() + int(dados.get("expires_in", 3600))

        self.logger.info("Token obtido! (válido por 1 hora)")
        return self._token


# ═══════════════════════════════════════════════════════════════════════════════
# MAPEAMENTO DE REGIÕES
# ═══════════════════════════════════════════════════════════════════════════════

PAISES_REGIAO = {
    # América do Norte
    "US": "América do Norte", "CA": "América do Norte", "MX": "América do Norte",
    "USA": "América do Norte",
    # Europa
    "DE": "Europa", "FR": "Europa", "GB": "Europa", "UK": "Europa",
    "IT": "Europa", "ES": "Europa", "NL": "Europa", "BE": "Europa",
    "CH": "Europa", "AT": "Europa", "SE": "Europa", "NO": "Europa",
    "DK": "Europa", "FI": "Europa", "PL": "Europa", "PT": "Europa",
    "IE": "Europa", "CZ": "Europa", "HU": "Europa", "RO": "Europa",
    # Oriente Médio
    "AE": "Oriente Médio", "IL": "Oriente Médio", "TR": "Oriente Médio",
    "SA": "Oriente Médio", "QA": "Oriente Médio", "KW": "Oriente Médio",
    "BH": "Oriente Médio", "OM": "Oriente Médio", "JO": "Oriente Médio",
    "LB": "Oriente Médio", "EG": "Oriente Médio", "IR": "Oriente Médio",
    # Ásia
    "CN": "Ásia", "IN": "Ásia", "JP": "Ásia", "KR": "Ásia",
    "SG": "Ásia", "TW": "Ásia", "TH": "Ásia", "MY": "Ásia",
    "HK": "Ásia", "PH": "Ásia", "ID": "Ásia", "VN": "Ásia",
}

def pais_para_regiao(country_code: str) -> str:
    return PAISES_REGIAO.get(str(country_code).upper(), "Outros")


def extrair_dados_relatorio(eventos: list) -> dict:
    """
    Extrai país de origem, data de criação e data de entrega dos eventos.
    Retorna dict com: pais_origem, regiao, data_criacao, data_entrega, lead_time_dias
    """
    if not eventos:
        return {"pais_origem": "", "regiao": "", "data_criacao": "", "data_entrega": "", "lead_time_dias": None}

    # Eventos em ordem cronológica (último = mais recente no índice 0)
    # Inverte para ter o mais antigo primeiro
    eventos_ord = list(reversed(eventos))

    # País de origem = país do PRIMEIRO evento
    pais_origem = ""
    data_criacao = ""
    for ev in eventos_ord:
        loc = ev.get("scanLocation", {})
        pais = loc.get("countryCode", "")
        if pais and pais.upper() != "BR":
            pais_origem = pais.upper()
            ts = ev.get("date", "") or ev.get("timestamp", "")
            if ts:
                data_criacao = str(ts)[:10]
            break

    # Data de entrega = evento com DELIVERED no último (índice 0)
    data_entrega = ""
    for ev in eventos:  # mais recente primeiro
        desc = ev.get("eventDescription", "").upper()
        if any(t in desc for t in ["DELIVERED", "ENTREGUE", "DELIVERY MADE"]):
            ts = ev.get("date", "") or ev.get("timestamp", "")
            if ts:
                data_entrega = str(ts)[:10]
            break

    # Lead time em dias úteis brasileiros
    lead_time = None
    if data_criacao and data_entrega:
        try:
            from datetime import datetime as _dt
            d1 = _dt.strptime(data_criacao[:10], "%Y-%m-%d").date()
            d2 = _dt.strptime(data_entrega[:10], "%Y-%m-%d").date()
            lead_time = dias_uteis_br(d1, d2)
        except:
            pass

    regiao = pais_para_regiao(pais_origem)

    return {
        "pais_origem":    pais_origem,
        "regiao":         regiao,
        "data_criacao":   data_criacao,
        "data_entrega":   data_entrega,
        "lead_time_dias": lead_time,
    }


# ==========================================================
# CONSULTA À API OFICIAL
# ==========================================================

def consultar_awb(awb_original: str, awb_limpo: str, token: str, config: Config, pedido: str = "", produto: str = "") -> ResultadoAWB:
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M")
    logger    = logging.getLogger("fedex_tracker")

    try:
        time.sleep(config.delay_entre)

        response = requests.post(
            config.url_tracking,
            headers={
                "Content-Type":  "application/json",
                "Authorization": f"Bearer {token}",
                "X-locale":      "pt_BR",
            },
            json={
                "includeDetailedScans": True,
                "trackingInfo": [{"trackingNumberInfo": {"trackingNumber": awb_limpo}}]
            },
            timeout=config.timeout,
        )

        if response.status_code == 401:
            raise Exception("Token inválido ou expirado")
        if response.status_code == 429:
            logger.warning(f"[{awb_limpo}] Rate limit — aguardando 30s...")
            time.sleep(30)
            raise Exception("Rate limit atingido")
        if response.status_code != 200:
            raise Exception(f"HTTP {response.status_code}: {response.text[:200]}")

        dados  = response.json()
        output = (
            dados
            .get("output", {})
            .get("completeTrackResults", [{}])[0]
            .get("trackResults", [{}])[0]
        )

        # Verifica erros da API
        erro = output.get("error", {})
        if erro:
            codigo = erro.get("code", "")
            msg    = erro.get("message", "")
            if codigo == "TRACKING.TRACKINGNUMBER.NOTFOUND":
                return ResultadoAWB(
                    AWB=awb_limpo, AWB_ORIGINAL=awb_original, PEDIDO=pedido,
                    PRODUTO=produto,
                    CATEGORIA="AWB NAO ENCONTRADO",
                    STATUS_FEDEX="NOT FOUND",
                    ULTIMO_EVENTO="Número não encontrado na base FedEx",
                    MOTIVO_CATEGORIA="AWB inválido ou não registrado",
                    DATA_CONSULTA=timestamp, SUCESSO=False,
                )
            raise Exception(f"{codigo}: {msg}")

        # Status da FedEx
        status_fedex = (
            output.get("latestStatusDetail", {})
            .get("description", "UNKNOWN")
            .upper().strip()
        )

        # Todos os eventos
        eventos = output.get("scanEvents", [])

        # Último evento formatado
        if eventos:
            ult    = eventos[0]
            descr  = ult.get("eventDescription", "").strip()
            city   = ult.get("scanLocation", {}).get("city", "").strip()
            pais   = ult.get("scanLocation", {}).get("countryCode", "").strip()
            data_e = ult.get("date", "").split("T")[0]
            hora_e = ult.get("date", "T").split("T")[1][:5] if "T" in ult.get("date", "") else ""
            ultimo_evento = f"{descr} — {city}/{pais} {data_e} {hora_e}".upper().strip(" —")
        else:
            ultimo_evento = "SEM EVENTOS REGISTRADOS"

        # Data de chegada em Memphis — primeiro scan com city=MEMPHIS (evento mais antigo)
        # Os eventos vêm do mais recente para o mais antigo, então invertemos
        data_chegada_memphis = ""
        for ev in reversed(eventos):
            ev_city = ev.get("scanLocation", {}).get("city", "").upper()
            if "MEMPHIS" in ev_city:
                data_chegada_memphis = ev.get("date", "").split("T")[0]
                break

        # Classificação nas 5 categorias
        categoria, motivo = classificar(eventos, status_fedex)

        # Timeline completa extraída dos scanEvents da FedEx
        # Mapeia eventDescription → categoria e deduplica progressões
        _CATS_ORDEM = ["LABEL CREATED","COMING TO BRAZIL","CUSTOMS INSPECTION","NATIONAL TRANSIT","OUT FOR DELIVERY","DELIVERED"]
        def _ev_para_cat(ev):
            desc = ev.get("eventDescription","").upper()
            city = ev.get("scanLocation",{}).get("city","").upper()
            country = ev.get("scanLocation",{}).get("countryCode","").upper()
            # Mesma lógica da classificação mas simplificada para mapeamento
            if any(t in desc for t in ["DELIVERED","ENTREGUE","DELIVERY MADE"]):
                return "DELIVERED"
            if any(t in desc for t in ["ON FEDEX VEHICLE","OUT FOR DELIVERY","SAIU PARA ENTREGA","EM UM VEICULO","EM VEÍCULO"]):
                return "OUT FOR DELIVERY"
            if any(t in desc for t in ["LIBERACAO DA REMESSA","LIBERAÇÃO DA REMESSA","INTERNATIONAL SHIPMENT RELEASE","RELEASED BY CUSTOMS","CUSTOMS CLEARED","CLEARANCE COMPLETED"]):
                return "NATIONAL TRANSIT"
            if any(t in desc for t in ["ATRASO NA LIBERACAO","ATRASO NA LIBERAÇÃO","CLEARANCE DELAY","CUSTOMS DELAY","HELD BY CUSTOMS","RETIDO NA ALFANDEGA","RETIDO NA ALFÂNDEGA","PACOTE DISPONIVEL","PACOTE DISPONÍVEL"]):
                return "CUSTOMS INSPECTION"
            if country == "BR" or city in ("CAMPINAS","VIRACOPOS","BARUERI"):
                return "CUSTOMS INSPECTION"
            if any(t in desc for t in ["LEFT FEDEX","DEPARTED FEDEX","SAIU DA","SHIPMENT LEFT","LEFT ORIGIN","PICKED UP","COLETADO"]):
                return "COMING TO BRAZIL"
            if any(t in desc for t in ["ARRIVED","CHEGOU","IN TRANSIT","EM TRANSITO","A CAMINHO","IN FEDEX POSSESSION"]):
                return "COMING TO BRAZIL"
            if any(t in desc for t in ["LABEL","ETIQUETA","SHIPMENT INFORMATION"]):
                return "LABEL CREATED"
            return None

        timeline_json = []
        _tl_prev = None
        for ev in reversed(eventos):  # do mais antigo para o mais recente
            _ev_cat = _ev_para_cat(ev)
            if _ev_cat and _ev_cat != _tl_prev:
                _ev_date = ev.get("date","").split("T")[0]
                timeline_json.append({"cat": _ev_cat, "data": _ev_date})
                _tl_prev = _ev_cat
        # Garante que a categoria atual está no final
        if timeline_json and timeline_json[-1]["cat"] != categoria:
            timeline_json.append({"cat": categoria, "data": datetime.now().strftime("%Y-%m-%d")})

        # Eventos brutos completos para painel de detalhe estilo FedEx
        eventos_json = []
        for ev in reversed(eventos):  # cronológico: mais antigo primeiro
            _dt_raw = ev.get("date","")
            _dt_d = _dt_raw.split("T")[0] if "T" in _dt_raw else _dt_raw[:10]
            _dt_h = _dt_raw.split("T")[1][:5] if "T" in _dt_raw else ""
            _loc = ev.get("scanLocation", {})
            _city    = _loc.get("city","").strip()
            _state   = _loc.get("stateOrProvinceCode","").strip()
            _country = _loc.get("countryCode","").strip()
            _local_str = ", ".join(filter(None, [_city, _state, _country]))
            _desc = ev.get("eventDescription","").strip()
            _subdesc = ev.get("eventSubDescription","").strip() if ev.get("eventSubDescription") else ""
            eventos_json.append({
                "data":    _dt_d,
                "hora":    _dt_h,
                "desc":    _desc,
                "subdesc": _subdesc,
                "local":   _local_str,
            })

        logger.debug(f"[{awb_limpo}] {categoria} | {status_fedex} | {ultimo_evento[:50]}")

        dados_rel = extrair_dados_relatorio(eventos)

        # Endereço de entrega (destinatário)
        dest = output.get("recipientInformation", {})
        dest_addr = dest.get("address", {})
        endereco_entrega = ", ".join(filter(None, [
            dest_addr.get("city",""),
            dest_addr.get("stateOrProvinceCode",""),
            dest_addr.get("countryCode",""),
        ]))

        return ResultadoAWB(
            AWB=awb_limpo,
            AWB_ORIGINAL=awb_original,
            CATEGORIA=categoria,
            STATUS_FEDEX=status_fedex,
            ULTIMO_EVENTO=ultimo_evento,
            MOTIVO_CATEGORIA=motivo,
            DATA_CONSULTA=timestamp,
            SUCESSO=True,
            DATA_CHEGADA_MEMPHIS=data_chegada_memphis,
            PEDIDO=pedido,
            PRODUTO=produto,
            PAIS_ORIGEM=dados_rel["pais_origem"],
            REGIAO=dados_rel["regiao"],
            DATA_CRIACAO=dados_rel["data_criacao"],
            DATA_ENTREGA=dados_rel["data_entrega"],
            LEAD_TIME_DIAS=dados_rel["lead_time_dias"],
            TIMELINE_JSON=json.dumps(timeline_json, ensure_ascii=False),
            EVENTOS_JSON=json.dumps(eventos_json, ensure_ascii=False),
            ENDERECO_ENTREGA=endereco_entrega,
        )

    except Exception as e:
        msg = str(e).split('\n')[0]
        logger.error(f"[{awb_limpo}] {msg}")
        return ResultadoAWB(
            AWB=awb_limpo,
            AWB_ORIGINAL=awb_original,
            CATEGORIA="ERRO",
            STATUS_FEDEX="ERRO",
            ULTIMO_EVENTO=msg[:200],
            MOTIVO_CATEGORIA="Falha na consulta",
            DATA_CONSULTA=timestamp,
            SUCESSO=False,
            PEDIDO=pedido,
            PRODUTO=produto,
        )


# ==========================================================
# STORAGE
# ==========================================================

def carregar_awbs(arquivo: str) -> List[tuple]:
    path = Path(arquivo)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {arquivo}")
    df = pd.read_excel(path)
    df.columns = df.columns.str.strip()
    if "AWB" not in df.columns:
        raise ValueError(f"Coluna 'AWB' não encontrada. Colunas: {list(df.columns)}")
    col_pedido  = "Pedido" if "Pedido" in df.columns else ("PEDIDO" if "PEDIDO" in df.columns else None)
    col_produto = "PRODUTO" if "PRODUTO" in df.columns else ("Produto" if "Produto" in df.columns else None)
    logger    = logging.getLogger("fedex_tracker")
    resultado = []
    for _, row in df.iterrows():
        awb_original = str(row["AWB"]) if pd.notna(row["AWB"]) else ""
        if not awb_original or awb_original == "nan":
            continue
        pedido = str(row[col_pedido]) if col_pedido and pd.notna(row[col_pedido]) else ""
        if pedido.endswith(".0"):
            pedido = pedido[:-2]
        produto = str(row[col_produto]).strip() if col_produto and pd.notna(row[col_produto]) else ""
        awb_limpo = limpar_awb(awb_original)
        if validar_awb(awb_limpo):
            resultado.append((awb_original.strip(), awb_limpo, pedido, produto))
        else:
            logger.warning(f"AWB ignorado: '{awb_original}'")
    logger.info(f"{len(resultado)} AWBs válidos")
    return resultado


def _fazer_backup(arquivo: str) -> None:
    path = Path(arquivo)
    if path.exists():
        ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = path.with_name(f"{path.stem}_bak_{ts}{path.suffix}")
        shutil.copy(path, backup)


def salvar_resultados(resultados, arquivo_historico, arquivo_ultimo) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    import re as _re

    logger = logging.getLogger("fedex_tracker")
    df = pd.DataFrame(resultados)
    df["AWB"] = df["AWB"].astype(str).str.strip()

    # Garante que PEDIDO está preenchido — faz join com awbs.xlsx se necessário
    if "PEDIDO" not in df.columns or df["PEDIDO"].isna().all() or (df["PEDIDO"] == "").all():
        try:
            awbs_ref = pd.read_excel(arquivo_historico.replace("historico_status", "awbs")
                                     if "historico_status" in arquivo_historico
                                     else str(Path(arquivo_historico).parent / "awbs.xlsx"))
            awbs_ref.columns = awbs_ref.columns.str.strip()
            col_ped = "Pedido" if "Pedido" in awbs_ref.columns else ("PEDIDO" if "PEDIDO" in awbs_ref.columns else None)
            if col_ped:
                awbs_ref["AWB_CLEAN"] = awbs_ref["AWB"].astype(str).str.replace(r"\s+", "", regex=True)
                awbs_ref[col_ped] = awbs_ref[col_ped].astype(str).str.replace(r"\.0$", "", regex=True)
                pedido_map = dict(zip(awbs_ref["AWB_CLEAN"], awbs_ref[col_ped]))
                df["PEDIDO"] = df["AWB"].astype(str).map(pedido_map).fillna("")
        except Exception as e:
            logger.warning(f"Não foi possível carregar pedidos do awbs.xlsx: {e}")
            df["PEDIDO"] = df.get("PEDIDO", "")
    else:
        df["PEDIDO"] = df["PEDIDO"].fillna("").astype(str).str.replace(r"\.0$", "", regex=True)

    ORDEM_CAT = {
        "LABEL CREATED": 1, "COMING TO BRAZIL": 2, "CUSTOMS INSPECTION": 3,
        "NATIONAL TRANSIT": 4, "OUT FOR DELIVERY": 5, "DELIVERED": 6,
        "AWB NAO ENCONTRADO": 7, "ERRO": 8,
    }

    # Paleta de cores por categoria
    CORES = {
        "LABEL CREATED":       {"bg": "D9D9D9", "font": "595959"},  # Cinza
        "COMING TO BRAZIL":    {"bg": "BDD7EE", "font": "1F4E79"},  # Azul claro
        "CUSTOMS INSPECTION":  {"bg": "FFE699", "font": "7F6000"},  # Amarelo
        "NATIONAL TRANSIT":    {"bg": "C6EFCE", "font": "276221"},  # Verde claro
        "OUT FOR DELIVERY":    {"bg": "9DC3E6", "font": "1F4E79"},  # Azul médio
        "DELIVERED":           {"bg": "375623", "font": "FFFFFF"},  # Verde escuro
        "AWB NAO ENCONTRADO":  {"bg": "F4CCCC", "font": "990000"},  # Vermelho claro
        "ERRO":                {"bg": "FF0000", "font": "FFFFFF"},  # Vermelho
    }

    # LIMIAR de atraso antigo (dias)
    DIAS_ATRASO_CUSTOMS = 5   # CUSTOMS INSPECTION: alerta após 5 dias
    DIAS_ATRASO_MEMPHIS  = 3   # COMING TO BRAZIL em Memphis: alerta após 3 dias
    DIAS_ATRASO = DIAS_ATRASO_CUSTOMS  # compatibilidade

    df["_ordem"] = df["CATEGORIA"].map(lambda x: ORDEM_CAT.get(x, 9))
    df = df.sort_values("_ordem").drop(columns=["_ordem"]).reset_index(drop=True)

    # Calcula dias úteis em atraso
    def dias_no_status(ultimo_evento_str):
        try:
            partes = str(ultimo_evento_str).split("—")
            data_str = partes[-1].strip().split(" ")
            data_part = [p for p in data_str if _re.match(r"\d{4}-\d{2}-\d{2}", p)]
            if data_part:
                from datetime import date
                d = datetime.strptime(data_part[0], "%Y-%m-%d").date()
                return dias_uteis_br(d, date.today())
        except:
            pass
        return None

    df["_dias"] = df["ULTIMO_EVENTO"].apply(dias_no_status)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Estilos reutilizáveis ──────────────────────────────────────────────
    def make_fill(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

    def make_border():
        side = Side(style="thin", color="BFBFBF")
        return Border(left=side, right=side, top=side, bottom=side)

    def header_style(cell, bg="1F4E79", fg="FFFFFF"):
        cell.font      = Font(bold=True, color=fg, name="Arial", size=10)
        cell.fill      = make_fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = make_border()

    def data_style(cell, bg="FFFFFF", fg="000000", bold=False, wrap=False):
        cell.font      = Font(color=fg, name="Arial", size=9, bold=bold)
        cell.fill      = make_fill(bg)
        cell.alignment = Alignment(vertical="center", wrap_text=wrap)
        cell.border    = make_border()

    COLUNAS = ["AWB", "PEDIDO", "CATEGORIA", "STATUS_FEDEX", "ULTIMO_EVENTO", "MOTIVO_CATEGORIA", "DATA_CONSULTA", "DATA_CHEGADA_MEMPHIS", "REGIAO", "TIMELINE_JSON", "EVENTOS_JSON", "ENDERECO_ENTREGA"]
    LARGURAS = [22, 14, 22, 28, 52, 40, 18, 14, 18, 10, 10, 20]
    HEADERS  = ["AWB", "PEDIDO", "CATEGORIA", "STATUS FEDEX", "ÚLTIMO EVENTO", "MOTIVO", "DATA CONSULTA", "CHEGADA MEMPHIS", "REGIÃO", "TIMELINE", "EVENTOS", "DEST"]

    def escrever_aba(ws, df_aba, titulo_aba):
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 32

        for col_idx, (header, largura) in enumerate(zip(HEADERS, LARGURAS), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            header_style(cell)
            ws.column_dimensions[get_column_letter(col_idx)].width = largura

        for row_idx, (_, row) in enumerate(df_aba.iterrows(), 2):
            cat    = row["CATEGORIA"]
            cores  = CORES.get(cat, {"bg": "FFFFFF", "font": "000000"})
            dias   = row.get("_dias")
            em_customs = (cat == "CUSTOMS INSPECTION" and dias and dias >= DIAS_ATRASO_CUSTOMS)
            # Alerta Memphis: usa a data de CHEGADA em Memphis, não o último scan
            data_memphis_str = str(row.get("DATA_CHEGADA_MEMPHIS", ""))
            dias_memphis = None
            if data_memphis_str and data_memphis_str != "nan" and data_memphis_str != "":
                try:
                    from datetime import date as _date
                    d_mem = datetime.strptime(data_memphis_str[:10], "%Y-%m-%d").date()
                    dias_memphis = dias_uteis_br(d_mem, _date.today())
                except:
                    pass
            em_memphis = (cat == "COMING TO BRAZIL" and dias_memphis is not None
                          and dias_memphis >= DIAS_ATRASO_MEMPHIS)
            atraso = em_customs or em_memphis

            for col_idx, col_name in enumerate(COLUNAS, 1):
                val  = row.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if pd.notna(val) else "")

                if atraso:
                    # Destaque vermelho forte para AWBs com atraso antigo
                    data_style(cell, bg="FF4C4C", fg="FFFFFF", bold=True, wrap=(col_idx in [4, 5]))
                else:
                    data_style(cell, bg=cores["bg"], fg=cores["font"], wrap=(col_idx in [4, 5]))

                ws.row_dimensions[row_idx].height = 28 if col_idx == 1 else ws.row_dimensions[row_idx].height

            # Marca visualmente atrasos — coluna CATEGORIA é índice 3 (após AWB e PEDIDO)
            if atraso:
                cell_cat = ws.cell(row=row_idx, column=3)
                if em_memphis:
                    cell_cat.value = f"⚠ MEMPHIS +{dias_memphis}d"
                else:
                    cell_cat.value = f"⚠ {cat} ({dias}d)"
                cell_cat.font  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
                cell_cat.fill  = make_fill("C00000")  # Vermelho escuro

    # ── ABA RESUMO ─────────────────────────────────────────────────────────
    ws_resumo = wb.create_sheet("📊 RESUMO")
    ws_resumo.sheet_properties.tabColor = "1F4E79"

    # Título
    ws_resumo.merge_cells("A1:F1")
    titulo = ws_resumo["A1"]
    titulo.value     = f"FEDEX TRACKING — RELATÓRIO {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    titulo.font      = Font(bold=True, color="FFFFFF", name="Arial", size=14)
    titulo.fill      = make_fill("1F4E79")
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws_resumo.row_dimensions[1].height = 36

    # Subtítulo KPIs
    total     = len(df)
    sucesso   = int(df["SUCESSO"].sum()) if "SUCESSO" in df.columns else total
    def _dias_memphis(row):
        s = str(row.get("DATA_CHEGADA_MEMPHIS", ""))
        if s and s != "nan":
            try:
                from datetime import date as _d
                return dias_uteis_br(datetime.strptime(s[:10], "%Y-%m-%d").date(), _d.today())
            except:
                pass
        return 0

    df["_dias_memphis"] = df.apply(_dias_memphis, axis=1)
    em_atraso = int(
        ((df["CATEGORIA"] == "CUSTOMS INSPECTION") & (df["_dias"] >= DIAS_ATRASO_CUSTOMS)).sum() +
        ((df["CATEGORIA"] == "COMING TO BRAZIL") & (df["_dias_memphis"] >= DIAS_ATRASO_MEMPHIS)).sum()
    )

    kpis = [
        ("TOTAL AWBs", total,     "2E75B6"),
        ("CONSULTADOS OK", sucesso, "375623"),
        ("⚠ EM ATRASO (+7d)", em_atraso, "C00000"),
    ]
    for col_i, (label, val, cor) in enumerate(kpis, 1):
        col_letter = get_column_letter(col_i * 2 - 1)
        col_letter2 = get_column_letter(col_i * 2)
        ws_resumo.merge_cells(f"{col_letter}3:{col_letter2}3")
        ws_resumo.merge_cells(f"{col_letter}4:{col_letter2}4")

        c_label = ws_resumo[f"{col_letter}3"]
        c_label.value = label
        c_label.font  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        c_label.fill  = make_fill(cor)
        c_label.alignment = Alignment(horizontal="center", vertical="center")

        c_val = ws_resumo[f"{col_letter}4"]
        c_val.value = val
        c_val.font  = Font(bold=True, color=cor, name="Arial", size=24)
        c_val.fill  = make_fill("F2F2F2")
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        ws_resumo.row_dimensions[4].height = 48

    # Tabela de contagem por categoria
    ws_resumo["A6"].value = "CATEGORIA"
    ws_resumo["B6"].value = "QUANTIDADE"
    ws_resumo["C6"].value = "%"
    for cell, bg in [(ws_resumo["A6"], "1F4E79"), (ws_resumo["B6"], "1F4E79"), (ws_resumo["C6"], "1F4E79")]:
        header_style(cell, bg=bg)

    contagem = df["CATEGORIA"].value_counts().to_dict()
    cats_ordem = [c for c in ORDEM_CAT.keys() if c in contagem]

    chart_cats, chart_vals = [], []
    row_t = 7
    for cat in cats_ordem:
        qtd   = contagem[cat]
        cores = CORES.get(cat, {"bg": "FFFFFF", "font": "000000"})

        c_cat = ws_resumo.cell(row=row_t, column=1, value=cat)
        c_qtd = ws_resumo.cell(row=row_t, column=2, value=qtd)
        c_pct = ws_resumo.cell(row=row_t, column=3, value=f"=B{row_t}/B${7+len(cats_ordem)}")

        for c, fmt in [(c_cat, None), (c_qtd, None), (c_pct, "0.0%")]:
            c.font      = Font(color=cores["font"], name="Arial", size=10, bold=True)
            c.fill      = make_fill(cores["bg"])
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = make_border()
            if fmt: c.number_format = fmt

        chart_cats.append(cat)
        chart_vals.append(qtd)
        row_t += 1

    # Total
    c_tot_label = ws_resumo.cell(row=row_t, column=1, value="TOTAL")
    c_tot_val   = ws_resumo.cell(row=row_t, column=2, value=f"=SUM(B7:B{row_t-1})")
    c_tot_pct   = ws_resumo.cell(row=row_t, column=3, value="100%")
    for c in [c_tot_label, c_tot_val, c_tot_pct]:
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill      = make_fill("1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = make_border()

    ws_resumo.column_dimensions["A"].width = 26
    ws_resumo.column_dimensions["B"].width = 14
    ws_resumo.column_dimensions["C"].width = 10

    # Gráfico de barras
    chart = BarChart()
    chart.type    = "bar"
    chart.title   = "AWBs por Categoria"
    chart.y_axis.title = "Quantidade"
    chart.x_axis.title = ""
    chart.style   = 10
    chart.width   = 18
    chart.height  = 12
    chart.legend  = None

    data_ref   = Reference(ws_resumo, min_col=2, min_row=6, max_row=row_t - 1)
    cats_ref   = Reference(ws_resumo, min_col=1, min_row=7, max_row=row_t - 1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "2E75B6"

    ws_resumo.add_chart(chart, "E3")

    # ── ABA GERAL (todos os AWBs) ──────────────────────────────────────────
    ws_geral = wb.create_sheet("📋 TODOS")
    ws_geral.sheet_properties.tabColor = "2E75B6"
    escrever_aba(ws_geral, df, "TODOS")

    # ── ABAS POR CATEGORIA ─────────────────────────────────────────────────
    ICONES = {
        "LABEL CREATED":      "🏷",
        "COMING TO BRAZIL":   "✈",
        "CUSTOMS INSPECTION": "🔍",
        "NATIONAL TRANSIT":   "🚚",
        "OUT FOR DELIVERY":   "📦",
        "DELIVERED":          "✅",
        "AWB NAO ENCONTRADO": "❓",
        "ERRO":               "❌",
    }
    TAB_CORES = {
        "LABEL CREATED":      "808080",
        "COMING TO BRAZIL":   "2E75B6",
        "CUSTOMS INSPECTION": "FFB900",
        "NATIONAL TRANSIT":   "375623",
        "OUT FOR DELIVERY":   "0070C0",
        "DELIVERED":          "375623",
        "AWB NAO ENCONTRADO": "C00000",
        "ERRO":               "FF0000",
    }

    # Cria aba para TODAS as categorias — mesmo as que não têm AWBs no momento
    TODAS_CATS = ["COMING TO BRAZIL", "CUSTOMS INSPECTION",
                  "NATIONAL TRANSIT", "OUT FOR DELIVERY", "DELIVERED"]

    for cat in TODAS_CATS:
        df_cat = df[df["CATEGORIA"] == cat].reset_index(drop=True)
        icone  = ICONES.get(cat, "")
        nome   = f"{icone} {cat}"[:31]
        ws     = wb.create_sheet(nome)
        ws.sheet_properties.tabColor = TAB_CORES.get(cat, "FFFFFF")
        escrever_aba(ws, df_cat, cat)

        # Mini-cabeçalho colorido na aba
        cores = CORES.get(cat, {"bg": "FFFFFF", "font": "000000"})
        ws.insert_rows(1)
        ws.merge_cells(f"A1:{get_column_letter(len(COLUNAS))}1")
        c_titulo = ws["A1"]
        qtd = len(df_cat)
        c_titulo.value     = f"{cat}  —  {qtd} AWB{'s' if qtd != 1 else ''}" if qtd > 0 else f"{cat}  —  Nenhuma AWB no momento"
        c_titulo.font      = Font(bold=True, color=cores["font"], name="Arial", size=12)
        c_titulo.fill      = make_fill(cores["bg"])
        c_titulo.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A3"

    # ── ABA MUDANÇAS ───────────────────────────────────────────────────────
    # Sempre cria a aba — compara consulta atual com o histórico anterior
    ws_mud = wb.create_sheet("🔄 MUDANÇAS")
    ws_mud.sheet_properties.tabColor = "7030A0"

    # Transições válidas: apenas progressões reais de uma etapa para outra
    TRANSICOES_VALIDAS = {
        ("COMING TO BRAZIL",   "CUSTOMS INSPECTION"),
        ("COMING TO BRAZIL",   "NATIONAL TRANSIT"),
        ("COMING TO BRAZIL",   "OUT FOR DELIVERY"),
        ("COMING TO BRAZIL",   "DELIVERED"),
        ("CUSTOMS INSPECTION", "NATIONAL TRANSIT"),
        ("CUSTOMS INSPECTION", "OUT FOR DELIVERY"),
        ("CUSTOMS INSPECTION", "DELIVERED"),
        ("NATIONAL TRANSIT",   "OUT FOR DELIVERY"),
        ("NATIONAL TRANSIT",   "DELIVERED"),
        ("OUT FOR DELIVERY",   "DELIVERED"),
        ("LABEL CREATED",      "COMING TO BRAZIL"),
        ("LABEL CREATED",      "CUSTOMS INSPECTION"),
        ("LABEL CREATED",      "NATIONAL TRANSIT"),
        ("LABEL CREATED",      "OUT FOR DELIVERY"),
        ("LABEL CREATED",      "DELIVERED"),
    }

    mudancas = pd.DataFrame()
    try:
        path_hist_check = Path(arquivo_historico)
        if path_hist_check.exists():
            hist_df = pd.read_excel(path_hist_check)
            if "DATA_CONSULTA" in hist_df.columns and "AWB" in hist_df.columns and "CATEGORIA" in hist_df.columns:
                hist_df["AWB"] = hist_df["AWB"].astype(str).str.strip()
                hist_df["DATA_CONSULTA"] = pd.to_datetime(hist_df["DATA_CONSULTA"], dayfirst=True, errors="coerce")

                consulta_atual_str = df["DATA_CONSULTA"].iloc[0] if len(df) > 0 else ""
                try:
                    consulta_atual_dt = pd.to_datetime(consulta_atual_str, dayfirst=True, errors="coerce")
                except:
                    consulta_atual_dt = None

                # Mudanças DO DIA: compara com o último snapshot ANTES de hoje
                if consulta_atual_dt is not None and pd.notna(consulta_atual_dt):
                    consulta_atual_date = consulta_atual_dt.date()
                    # Pega apenas registros de HOJE para comparar (mudanças do dia)
                    hist_hoje = hist_df[hist_df["DATA_CONSULTA"].dt.date == consulta_atual_date]
                    # Baseline: última consulta antes de hoje
                    hist_ant = hist_df[hist_df["DATA_CONSULTA"].dt.date < consulta_atual_date]
                    # Se já há consultas hoje, compara com a consulta imediatamente anterior do dia
                    if len(hist_hoje) > 0 and len(hist_ant) == 0:
                        hist_ant = hist_df  # fallback
                else:
                    hist_ant = hist_df

                # Fallback: se não há datas anteriores, busca snapshot ou ultimo_status_gerado
                if len(hist_ant) == 0:
                    pasta = Path(arquivo_historico).parent
                    # Tenta snapshots primeiro (mais recente antes de hoje)
                    candidatos = [pasta / "ultimo_status_anterior.xlsx"]
                    candidatos += sorted(pasta.glob("snapshot_*.xlsx"), reverse=True)
                    candidatos += [pasta / "ultimo_status_gerado.xlsx"]
                    for _cand in candidatos:
                        if not _cand.exists():
                            continue
                        try:
                            _sheet = "📋 TODOS" if "snapshot" not in _cand.name else None
                            _kw = {"sheet_name": _sheet} if _sheet else {}
                            try:
                                ult_df = pd.read_excel(_cand, sheet_name="📋 TODOS")
                            except Exception:
                                ult_df = pd.read_excel(_cand)
                            col_data = "DATA CONSULTA" if "DATA CONSULTA" in ult_df.columns else "DATA_CONSULTA"
                            ult_df = ult_df.rename(columns={col_data: "DATA_CONSULTA"})
                            ult_df["AWB"] = ult_df["AWB"].astype(str).str.strip()
                            ult_df["DATA_CONSULTA"] = pd.to_datetime(ult_df["DATA_CONSULTA"], dayfirst=True, errors="coerce")
                            _candidatos_ant = ult_df[ult_df["DATA_CONSULTA"] < consulta_atual_dt] if consulta_atual_dt is not None else ult_df
                            if len(_candidatos_ant) > 0:
                                hist_ant = _candidatos_ant
                                logger.info(f"Fallback: usando '{_cand.name}' para comparação ({len(hist_ant)} registros)")
                                break
                        except Exception as e_ult:
                            logger.warning(f"Fallback '{_cand.name}' falhou: {e_ult}")

                if len(hist_ant) > 0:
                    # Normaliza categorias antigas que possam ter ⚠ ou sufixos
                    import re as _re3
                    def _norm_cat(c):
                        c = str(c)
                        c = _re3.sub(r"^[^A-Z]*", "", c)
                        c = _re3.sub(r"\s*\(\d+d\)\s*$", "", c)
                        c = _re3.sub(r"\s*MEMPHIS.*$", "", c)
                        return c.strip()
                    hist_ant = hist_ant.copy()
                    hist_ant["CATEGORIA"] = hist_ant["CATEGORIA"].apply(_norm_cat)

                    # Último status de cada AWB antes da consulta atual
                    # Força AWB como string nos dois lados para evitar erro de tipo no merge
                    hist_ant["AWB"] = hist_ant["AWB"].astype(str).str.strip()
                    ultimo_ant = (hist_ant.sort_values("DATA_CONSULTA")
                                         .groupby("AWB", as_index=False)
                                         .last()[["AWB", "CATEGORIA"]]
                                         .rename(columns={"CATEGORIA": "CAT_ANTERIOR"}))

                    df_comp = df.copy()
                    df_comp["AWB"] = df_comp["AWB"].astype(str).str.strip()
                    df_comp = df_comp.merge(ultimo_ant, on="AWB", how="left")

                    # Normaliza CATEGORIA atual antes de comparar
                    import re as _re4
                    def _nc(c):
                        c = str(c)
                        c = _re4.sub(r"^⚠\s*", "", c)
                        c = _re4.sub(r"\s*\(\d+d\)\s*$", "", c)
                        c = _re4.sub(r"\s*MEMPHIS.*$", "", c)
                        return c.strip()
                    df_comp = df_comp.copy()
                    df_comp["CATEGORIA_NORM"] = df_comp["CATEGORIA"].apply(_nc)

                    # Filtra apenas transições válidas (progressões reais)
                    def eh_transicao_valida(row):
                        ant = row.get("CAT_ANTERIOR")
                        atu = row.get("CATEGORIA_NORM")
                        if pd.isna(ant) or not ant:
                            return False
                        return (str(ant), str(atu)) in TRANSICOES_VALIDAS

                    mudancas = df_comp[df_comp.apply(eh_transicao_valida, axis=1)].copy().reset_index(drop=True)
    except Exception as e:
        logger.warning(f"Erro ao comparar histórico para mudanças: {e}")

    # Cabeçalho título — sempre mostra, com ou sem mudanças
    n_cols = len(COLUNAS) + 1
    ws_mud.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c_t = ws_mud["A1"]
    c_t.value     = (f"AWBs QUE MUDARAM DE STATUS  —  {len(mudancas)} alteração(ões)"
                     if len(mudancas) > 0 else "AWBs QUE MUDARAM DE STATUS  —  Nenhuma mudança detectada")
    c_t.font      = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    c_t.fill      = make_fill("7030A0") if len(mudancas) > 0 else make_fill("595959")
    c_t.alignment = Alignment(horizontal="center", vertical="center")
    ws_mud.row_dimensions[1].height = 28

    # Cabeçalhos das colunas
    headers_mud  = HEADERS + ["STATUS ANTERIOR"]
    larguras_mud = LARGURAS + [22]
    for col_idx, (h, larg) in enumerate(zip(headers_mud, larguras_mud), 1):
        cell = ws_mud.cell(row=2, column=col_idx, value=h)
        header_style(cell)
        ws_mud.column_dimensions[get_column_letter(col_idx)].width = larg
    ws_mud.freeze_panes = "A3"
    ws_mud.row_dimensions[2].height = 28

    if len(mudancas) == 0:
        ws_mud.merge_cells(f"A3:{get_column_letter(n_cols)}3")
        c_vazio = ws_mud["A3"]
        c_vazio.value     = "Nenhuma mudança de status desde a última consulta."
        c_vazio.font      = Font(italic=True, color="595959", name="Arial", size=10)
        c_vazio.alignment = Alignment(horizontal="center", vertical="center")
        ws_mud.row_dimensions[3].height = 28
    else:
        for row_idx, (_, row) in enumerate(mudancas.iterrows(), 3):
            cat      = row["CATEGORIA"]
            cat_ant  = row.get("CAT_ANTERIOR", "")
            cores    = CORES.get(cat,     {"bg": "FFFFFF", "font": "000000"})
            cores_ant= CORES.get(cat_ant, {"bg": "EEEEEE", "font": "333333"})

            dias_m   = row.get("_dias")
            data_mem = str(row.get("DATA_CHEGADA_MEMPHIS", ""))
            dias_mem_m = None
            if data_mem and data_mem != "nan":
                try:
                    from datetime import date as _d2
                    dias_mem_m = dias_uteis_br(datetime.strptime(data_mem[:10], "%Y-%m-%d").date(), _d2.today())
                except: pass
            em_c = (cat == "CUSTOMS INSPECTION" and dias_m and dias_m >= DIAS_ATRASO_CUSTOMS)
            em_m = (cat == "COMING TO BRAZIL" and dias_mem_m is not None and dias_mem_m >= DIAS_ATRASO_MEMPHIS)
            at   = em_c or em_m

            for col_idx, col_name in enumerate(COLUNAS, 1):
                val  = row.get(col_name, "")
                cell = ws_mud.cell(row=row_idx, column=col_idx, value=str(val) if pd.notna(val) else "")
                if at:
                    data_style(cell, bg="C00000", fg="FFFFFF", bold=True, wrap=(col_idx in [5, 6]))
                else:
                    data_style(cell, bg=cores["bg"], fg=cores["font"], wrap=(col_idx in [5, 6]))

            # Última coluna: STATUS ANTERIOR com cor original
            cell_ant = ws_mud.cell(row=row_idx, column=len(COLUNAS)+1, value=cat_ant)
            data_style(cell_ant, bg=cores_ant["bg"], fg=cores_ant["font"], bold=True)

            # Coluna CATEGORIA (índice 3) mostra a transição com seta
            cell_cat = ws_mud.cell(row=row_idx, column=3)
            cell_cat.value = f"{cat_ant} → {cat}"
            if at:
                cell_cat.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
                cell_cat.fill = make_fill("C00000")
            else:
                cell_cat.font = Font(bold=True, color=cores["font"], name="Arial", size=9)
                cell_cat.fill = make_fill(cores["bg"])

            ws_mud.row_dimensions[row_idx].height = 28

    # Remove coluna auxiliar _dias antes de salvar histórico
    df_save = df.drop(columns=["_dias", "_dias_memphis"], errors="ignore").copy()

    # Normaliza CATEGORIA no histórico — remove alertas visuais (⚠ CUSTOMS INSPECTION (10d) → CUSTOMS INSPECTION)
    import re as _re2
    def _limpar_cat(c):
        c = str(c)
        c = _re2.sub(r"^⚠\s*", "", c)          # remove ⚠ do início
        c = _re2.sub(r"\s*\(\d+d\)\s*$", "", c) # remove (Xd) do fim
        c = _re2.sub(r"\s*MEMPHIS.*$", "", c)   # remove sufixo MEMPHIS
        return c.strip()
    df_save["CATEGORIA"] = df_save["CATEGORIA"].apply(_limpar_cat)

    # Salva cópia ANTES de sobrescrever — usada para comparar mudanças no HTML
    import shutil as _shutil
    path_ult_pre = Path(arquivo_ultimo)
    path_anterior = path_ult_pre.parent / "ultimo_status_anterior.xlsx"
    if path_ult_pre.exists():
        try:
            _shutil.copy2(str(path_ult_pre), str(path_anterior))
            logger.info(f"Anterior salvo → '{path_anterior}'")
        except Exception as _e_s:
            logger.warning(f"Backup anterior falhou: {_e_s}")

    # ── ABA RELATÓRIO MENSAL ────────────────────────────────────────────────
    try:
        from datetime import datetime as _dtm
        import calendar as _cal
        hoje = _dtm.now()
        mes_ant = (hoje.month - 2) % 12 + 1
        ano_ant = hoje.year if hoje.month > 1 else hoje.year - 1
        nome_mes = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                    "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"][mes_ant - 1]

        # Filtra entregues no mês anterior
        df_rel = df[df["CATEGORIA"].apply(lambda c: "DELIVERED" in str(c).upper())].copy()
        if "DATA_ENTREGA" in df_rel.columns:
            df_rel["_de"] = pd.to_datetime(df_rel["DATA_ENTREGA"], errors="coerce")
            df_rel = df_rel[
                (df_rel["_de"].dt.month == mes_ant) &
                (df_rel["_de"].dt.year  == ano_ant)
            ]

        ws_rel = wb.create_sheet(f"📅 MENSAL {nome_mes.upper()[:3]}")
        ws_rel.sheet_properties.tabColor = "1F4E79"

        REGIOES = ["América do Norte", "Europa", "Oriente Médio", "Ásia", "Outros"]
        COR_REG  = {"América do Norte":"BDD7EE","Europa":"C6EFCE","Oriente Médio":"FFE699","Ásia":"F4CCFF","Outros":"D9D9D9"}

        # Título
        ws_rel.merge_cells("A1:H1")
        ct = ws_rel["A1"]
        ct.value = f"RELATÓRIO MENSAL — {nome_mes.upper()} {ano_ant}"
        ct.font  = Font(bold=True, color="FFFFFF", name="Arial", size=13)
        ct.fill  = make_fill("1F4E79")
        ct.alignment = Alignment(horizontal="center", vertical="center")
        ws_rel.row_dimensions[1].height = 32

        row_idx = 3
        total_geral = len(df_rel)

        # Resumo geral
        ws_rel.merge_cells(f"A{row_idx}:H{row_idx}")
        c = ws_rel[f"A{row_idx}"]
        c.value = f"Total geral Expresso: {total_geral} envios"
        c.font  = Font(bold=True, name="Arial", size=11)
        c.alignment = Alignment(vertical="center")
        ws_rel.row_dimensions[row_idx].height = 24
        row_idx += 1

        # Linha de totais por região
        totais_reg = {}
        if "REGIAO" in df_rel.columns:
            totais_reg = df_rel["REGIAO"].value_counts().to_dict()
        partes = [f"{r}: {totais_reg.get(r,0)}" for r in REGIOES if totais_reg.get(r,0) > 0]
        ws_rel.merge_cells(f"A{row_idx}:H{row_idx}")
        c2 = ws_rel[f"A{row_idx}"]
        c2.value = "  •  " + "  •  ".join(partes) if partes else "Sem dados de região"
        c2.font  = Font(name="Arial", size=10, color="595959")
        c2.alignment = Alignment(vertical="center")
        ws_rel.row_dimensions[row_idx].height = 20
        row_idx += 2

        # Detalhes por região
        for regiao in REGIOES:
            df_r = df_rel[df_rel.get("REGIAO", pd.Series()) == regiao] if "REGIAO" in df_rel.columns else pd.DataFrame()
            if len(df_r) == 0:
                continue

            cor = COR_REG.get(regiao, "EEEEEE")

            # Cabeçalho da região
            ws_rel.merge_cells(f"A{row_idx}:H{row_idx}")
            ch = ws_rel[f"A{row_idx}"]
            ch.value = regiao.upper()
            ch.font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            ch.fill  = make_fill("2E4057")
            ch.alignment = Alignment(horizontal="center", vertical="center")
            ws_rel.row_dimensions[row_idx].height = 24
            row_idx += 1

            # KPIs da região
            total_r = len(df_r)
            lt_vals = df_r["LEAD_TIME_DIAS"].dropna().astype(float) if "LEAD_TIME_DIAS" in df_r.columns else pd.Series()
            media_lt = round(lt_vals.mean()) if len(lt_vals) > 0 else "-"
            min_lt   = int(lt_vals.min()) if len(lt_vals) > 0 else "-"
            max_lt   = int(lt_vals.max()) if len(lt_vals) > 0 else "-"
            prod_top = df_r["PRODUTO"].value_counts().index[0] if "PRODUTO" in df_r.columns and len(df_r) > 0 else "-"

            kpis = [
                ("Total envios", str(total_r)),
                ("Lead time médio", f"{media_lt} dias"),
                ("Lead time mín.", f"{min_lt} dias"),
                ("Lead time máx.", f"{max_lt} dias"),
                ("Produto top", str(prod_top)),
            ]

            for col_i, (label, valor) in enumerate(kpis, 1):
                cl = ws_rel.cell(row=row_idx, column=col_i*2-1, value=label)
                cl.font = Font(bold=True, name="Arial", size=9, color="595959")
                cl.alignment = Alignment(horizontal="center", vertical="center")
                cv = ws_rel.cell(row=row_idx, column=col_i*2, value=valor)
                cv.font = Font(bold=True, name="Arial", size=11)
                cv.fill = make_fill(cor)
                cv.alignment = Alignment(horizontal="center", vertical="center")

            ws_rel.row_dimensions[row_idx].height = 28
            row_idx += 2

            # Tabela de AWBs da região
            hdrs = ["AWB", "PEDIDO", "PRODUTO", "PAÍS", "DATA ENVIO", "DATA ENTREGA", "LEAD TIME", "STATUS"]
            for ci, h in enumerate(hdrs, 1):
                c = ws_rel.cell(row=row_idx, column=ci, value=h)
                header_style(c)
                ws_rel.column_dimensions[get_column_letter(ci)].width = [18,10,22,8,14,14,12,18][ci-1]
            ws_rel.row_dimensions[row_idx].height = 22
            row_idx += 1

            for _, rr in df_r.iterrows():
                lt_val = rr.get("LEAD_TIME_DIAS","")
                lt_str = f"{int(lt_val)}d" if lt_val and str(lt_val) not in ("","nan","None") else "-"
                vals = [
                    str(rr.get("AWB","")),
                    str(rr.get("PEDIDO","")),
                    str(rr.get("PRODUTO","")),
                    str(rr.get("PAIS_ORIGEM","")),
                    str(rr.get("DATA_CRIACAO",""))[:10],
                    str(rr.get("DATA_ENTREGA",""))[:10],
                    lt_str,
                    str(rr.get("STATUS_FEDEX",""))[:30],
                ]
                for ci, v in enumerate(vals, 1):
                    c = ws_rel.cell(row=row_idx, column=ci, value=v)
                    data_style(c, bg=cor, fg="000000", wrap=False)
                ws_rel.row_dimensions[row_idx].height = 20
                row_idx += 1

            row_idx += 1  # espaço entre regiões

    except Exception as _e_rel:
        logger.warning(f"Aba mensal: {_e_rel}")

    wb.save(arquivo_ultimo)
    logger.info(f"Relatório salvo → '{arquivo_ultimo}'")

    _fazer_backup(arquivo_historico)
    path_hist = Path(arquivo_historico)
    if path_hist.exists():
        hist  = pd.read_excel(path_hist)
        final = pd.concat([hist, df_save], ignore_index=True)
    else:
        final = df_save

    final = final.drop_duplicates(subset=["AWB", "DATA_CONSULTA"], keep="last")

    # Mantém apenas os últimos 3 meses no histórico
    try:
        if "DATA_CONSULTA" in final.columns:
            final["_dc_dt"] = pd.to_datetime(final["DATA_CONSULTA"], dayfirst=True, errors="coerce")
            cutoff = pd.Timestamp.now() - pd.DateOffset(months=3)
            final = final[final["_dc_dt"].isna() | (final["_dc_dt"] >= cutoff)]
            final = final.drop(columns=["_dc_dt"])
    except Exception as _eh:
        logger.warning(f"Não foi possível truncar histórico: {_eh}")

    final.to_excel(arquivo_historico, index=False)
    logger.info(f"Histórico → '{arquivo_historico}' ({len(final)} registros)")


def gerar_resumo(resultados) -> str:
    total    = len(resultados)
    sucessos = sum(1 for r in resultados if r["SUCESSO"])

    ordem = ["LABEL CREATED", "COMING TO BRAZIL", "CUSTOMS INSPECTION",
             "NATIONAL TRANSIT", "OUT FOR DELIVERY", "DELIVERED",
             "AWB NAO ENCONTRADO", "ERRO"]

    contagem = {}
    for r in resultados:
        cat = r["CATEGORIA"]
        contagem[cat] = contagem.get(cat, 0) + 1

    linhas = [
        "", "=" * 60,
        f"  RESUMO DO RASTREIO — {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        "=" * 60,
        f"  Total consultado : {total}",
        f"  Com sucesso      : {sucessos}",
        f"  Com erro         : {total - sucessos}",
        "", "  Por categoria (ordem de progressão):",
    ]
    for cat in ordem:
        if cat in contagem:
            linhas.append(f"    {cat:<30} {contagem[cat]:>4}x")
    linhas.append("=" * 60)
    return "\n".join(linhas)


# ==========================================================
# MAIN
# ==========================================================

def gerar_html_relatorio(arquivo_excel: str, arquivo_html: str, arquivo_historico: str = "") -> None:
    import json, re as _re_html
    logger = logging.getLogger("fedex_tracker")

    def _norm(c):
        c = str(c)
        c = _re_html.sub(r"^[^A-Z]*", "", c)
        c = _re_html.sub(r"\s*\(\d+d\)\s*$", "", c)
        c = _re_html.sub(r"\s*MEMPHIS.*$", "", c)
        return c.strip()

    try:
        sheets = pd.read_excel(arquivo_excel, sheet_name=None)
    except Exception as e:
        logger.warning(f"HTML: não foi possível abrir {arquivo_excel}: {e}")
        return

    df = sheets.get("📋 TODOS", pd.DataFrame())
    if df.empty:
        return

    df.columns = df.columns.str.strip()
    df = df.rename(columns={"STATUS FEDEX":"STATUS_FEDEX","ÚLTIMO EVENTO":"ULTIMO_EVENTO","DATA CONSULTA":"DATA_CONSULTA","CHEGADA MEMPHIS":"DATA_CHEGADA_MEMPHIS","REGIÃO":"REGIAO","TIMELINE":"TIMELINE_JSON","EVENTOS":"EVENTOS_JSON","DEST":"ENDERECO_ENTREGA"})
    df["CATEGORIA_NORM"] = df["CATEGORIA"].apply(_norm)
    df = df.fillna("")

    # Recalcula dias úteis no status a partir do ULTIMO_EVENTO (campo salvo no Excel)
    def _recalc_dias(ultimo_evento_str):
        try:
            partes = str(ultimo_evento_str).split("—")
            data_str = partes[-1].strip().split(" ")
            data_part = [p for p in data_str if re.match(r"\d{4}-\d{2}-\d{2}", p)]
            if data_part:
                from datetime import date
                d = datetime.strptime(data_part[0], "%Y-%m-%d").date()
                return dias_uteis_br(d, date.today())
        except Exception:
            pass
        return None

    def _recalc_dias_memphis(data_memphis_str):
        try:
            s = str(data_memphis_str)
            if s and s not in ("nan",""):
                from datetime import date
                d = datetime.strptime(s[:10], "%Y-%m-%d").date()
                return dias_uteis_br(d, date.today())
        except Exception:
            pass
        return None

    df["_dias"]         = df["ULTIMO_EVENTO"].apply(_recalc_dias)
    df["_dias_memphis"] = df.get("DATA_CHEGADA_MEMPHIS", pd.Series(dtype=str)).apply(_recalc_dias_memphis)

    df_mud = sheets.get("🔄 MUDANÇAS", pd.DataFrame())

    # Dados mensais — tenta ler a aba mais recente do tipo 📅 MENSAL
    df_mensal = pd.DataFrame()
    nome_mes_rel = ""
    for sheet_name in sheets:
        if "MENSAL" in sheet_name.upper():
            df_mensal = sheets[sheet_name]
            nome_mes_rel = sheet_name
            break
    mudancas_rows = []
    if not df_mud.empty:
        for _, r in df_mud.iterrows():
            vals = [str(v) if pd.notna(v) else "" for v in r.values]
            if vals[0] and vals[0] not in ("nan","None") and _re_html.match(r"\d{9,}", vals[0]):
                mudancas_rows.append(vals[:3])

    data_consulta = df["DATA_CONSULTA"].iloc[0] if len(df) > 0 else ""
    total = len(df)
    ORDEM = ["COMING TO BRAZIL","CUSTOMS INSPECTION","NATIONAL TRANSIT","OUT FOR DELIVERY","DELIVERED","LABEL CREATED"]
    ICONS = {"COMING TO BRAZIL":"✈","CUSTOMS INSPECTION":"🔍","NATIONAL TRANSIT":"🚚","OUT FOR DELIVERY":"📦","DELIVERED":"✅","LABEL CREATED":"🏷"}
    cat_classes = {"COMING TO BRAZIL":"coming","CUSTOMS INSPECTION":"customs","NATIONAL TRANSIT":"national","OUT FOR DELIVERY":"ofd","DELIVERED":"delivered","LABEL CREATED":"label"}
    cats = df["CATEGORIA_NORM"].value_counts().to_dict()

    # Carrega histórico para ETA e linha do tempo por AWB
    _hist_rows = {}
    _eta_map   = {}
    try:
        _path_h = Path(arquivo_historico if arquivo_historico else "historico_status.xlsx")
        if _path_h.exists():
            _hdf = pd.read_excel(_path_h)
            _hdf.columns = _hdf.columns.str.strip()
            if "AWB" in _hdf.columns and "CATEGORIA" in _hdf.columns and "DATA_CONSULTA" in _hdf.columns:
                _hdf["AWB"] = _hdf["AWB"].astype(str).str.strip()
                _hdf["_dc"] = pd.to_datetime(_hdf["DATA_CONSULTA"], dayfirst=True, errors="coerce")
                _hdf = _hdf.sort_values("_dc")
                for _awb, _grp in _hdf.groupby("AWB"):
                    _timeline = []
                    _prev = None
                    for _, _row in _grp.iterrows():
                        import re as _re2
                        _cat = _re2.sub(r"^⚠\s*","", str(_row.get("CATEGORIA",""))).strip()
                        _cat = _re2.sub(r"\s*\(\d+d\)\s*$","",_cat).strip()
                        _cat = _re2.sub(r"\s*MEMPHIS.*$","",_cat).strip()
                        _dt  = str(_row["_dc"])[:10] if pd.notna(_row["_dc"]) else ""
                        if _cat != _prev and _cat and _cat not in ("nan","ERRO"):
                            _timeline.append({"cat":_cat,"data":_dt})
                            _prev = _cat
                    _hist_rows[_awb] = _timeline

                # ETA: média de dias úteis por categoria → destino final (DELIVERED) por região
                _CATS_ORDEM = ["LABEL CREATED","COMING TO BRAZIL","CUSTOMS INSPECTION","NATIONAL TRANSIT","OUT FOR DELIVERY","DELIVERED"]
                _delivered = _hdf[_hdf["CATEGORIA"].astype(str).str.upper().str.contains("DELIVERED",na=False)].copy()
                if "LEAD_TIME_DIAS" in _delivered.columns and "REGIAO" in _delivered.columns:
                    for _reg, _grp2 in _delivered.groupby("REGIAO"):
                        _lt = pd.to_numeric(_grp2["LEAD_TIME_DIAS"], errors="coerce").dropna()
                        if len(_lt) >= 2:
                            _eta_map[str(_reg)] = round(float(_lt.mean()))
    except Exception as _eh2:
        pass

    rows_json = []
    for _, r in df.iterrows():
        _awb = str(r["AWB"])
        _dias = r.get("_dias")
        _dias_val = int(_dias) if _dias is not None and not pd.isna(_dias) else None
        _regiao = str(r.get("REGIAO","")) if "REGIAO" in df.columns else ""
        _eta_dias = _eta_map.get(_regiao)
        _eta_str = ""
        if _eta_dias and _dias_val is not None and r["CATEGORIA_NORM"] not in ("DELIVERED","ERRO","AWB NAO ENCONTRADO"):
            _restante = max(0, _eta_dias - _dias_val)
            _eta_str = f"~{_restante}d úteis"

        # Timeline: usa TIMELINE_JSON da consulta atual (direto da API FedEx) como fonte primária
        # Fallback para histórico se não disponível
        _timeline = []
        _tl_raw = str(r.get("TIMELINE_JSON","")) if "TIMELINE_JSON" in df.columns else ""
        if _tl_raw and _tl_raw not in ("nan",""):
            try:
                _timeline = json.loads(_tl_raw)
            except Exception:
                _timeline = _hist_rows.get(_awb, [])
        else:
            _timeline = _hist_rows.get(_awb, [])

        # Eventos brutos para painel de detalhe
        _ev_raw = str(r.get("EVENTOS_JSON","")) if "EVENTOS_JSON" in df.columns else ""
        _eventos = []
        if _ev_raw and _ev_raw not in ("nan",""):
            try:
                _eventos = json.loads(_ev_raw)
            except Exception:
                pass

        rows_json.append({
            "awb":       _awb,
            "pedido":    str(r["PEDIDO"]),
            "produto":   str(r.get("PRODUTO","")) if "PRODUTO" in df.columns else "",
            "cat":       r["CATEGORIA_NORM"],
            "status":    str(r.get("STATUS_FEDEX","")),
            "evento":    str(r.get("ULTIMO_EVENTO","")),
            "data":      str(r.get("DATA_CONSULTA","")),
            "alerta":    "⚠" in str(r["CATEGORIA"]),
            "dias":      _dias_val,
            "eta":       _eta_str,
            "timeline":  _timeline,
            "eventos":   _eventos,
            "endereco":  str(r.get("ENDERECO_ENTREGA","")) if "ENDERECO_ENTREGA" in df.columns else "",
        })

    kpi_html = ""
    for cat in ORDEM:
        n = cats.get(cat, 0)
        cls = cat_classes.get(cat,"label")
        icon = ICONS.get(cat,"•")
        short = cat.replace("COMING TO BRAZIL","COMING").replace("CUSTOMS INSPECTION","CUSTOMS").replace("NATIONAL TRANSIT","NATIONAL").replace("OUT FOR DELIVERY","OUT FOR DEL.").replace("LABEL CREATED","LABEL")
        kpi_html += f'  <div class="kpi {cls}" onclick="filterCat(\'{cat}\')" data-cat="{cat}"><div class="kpi-icon">{icon}</div><div class="kpi-num">{n}</div><div class="kpi-label">{short}</div></div>\n'

    filter_tags = ""
    for cat in ORDEM:
        icon = ICONS.get(cat,"")
        short = cat.replace("COMING TO BRAZIL","COMING").replace("CUSTOMS INSPECTION","CUSTOMS").replace("NATIONAL TRANSIT","NATIONAL").replace("OUT FOR DELIVERY","OUT FOR DEL.").replace("LABEL CREATED","LABEL")
        filter_tags += f'  <button class="filter-tag" onclick="filterCat(\'{cat}\')" data-ftag="{cat}">{icon} {short}</button>\n'

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>FedEx Tracker — {data_consulta}</title>
<meta http-equiv="refresh" content="3600">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
:root{{
  --bg:#0d1117;--surface:#161b2a;--surface2:#1c2238;--surface3:#212840;
  --border:#21293f;--border2:#2d3755;
  --text:#dde1f0;--muted:#4b5577;--muted2:#6b7599;
  --accent:#4f7dff;--accent2:#3d6be8;
  --coming:#1a3a6c;--coming-fg:#93c5fd;--coming-bar:#4f7dff;
  --customs:#3d2e00;--customs-fg:#fcd34d;--customs-bar:#f59e0b;
  --national:#0f3320;--national-fg:#6ee7b7;--national-bar:#10b981;
  --ofd:#1e2f5c;--ofd-fg:#a5b4fc;--ofd-bar:#818cf8;
  --delivered:#052e16;--delivered-fg:#4ade80;--delivered-bar:#22c55e;
  --label:#161b2a;--label-fg:#64748b;--label-bar:#475569;
  --alert:#3d0f0f;--alert-fg:#fca5a5;
  font-size:14px
}}
body{{background:var(--bg);color:var(--text);font-family:'DM Sans',sans-serif;min-height:100vh}}

.header{{background:var(--surface);border-bottom:1px solid var(--border);padding:16px 32px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100}}
.logo{{font-family:'DM Mono',monospace;font-size:12px;color:var(--accent);letter-spacing:.1em;display:flex;align-items:center;gap:10px}}
.logo-dot{{width:7px;height:7px;background:var(--accent);border-radius:50%;animation:pulse 2.5s ease-in-out infinite}}
@keyframes pulse{{0%,100%{{opacity:1;transform:scale(1)}}50%{{opacity:.3;transform:scale(.7)}}}}
.header-right{{display:flex;align-items:center;gap:20px}}
.header-date{{font-family:'DM Mono',monospace;font-size:11px;color:var(--muted2)}}
.header-countdown{{font-family:'DM Mono',monospace;font-size:11px;color:var(--muted);background:var(--surface2);padding:4px 10px;border-radius:20px;border:1px solid var(--border)}}

.kpi-strip{{display:grid;grid-template-columns:repeat(6,1fr);gap:10px;padding:20px 32px 0}}
.kpi{{background:var(--surface);border:1px solid var(--border);border-left:3px solid transparent;border-radius:0 8px 8px 0;padding:14px 16px;cursor:pointer;transition:background .15s,border-color .15s,transform .15s;position:relative}}
.kpi:hover{{background:var(--surface2);transform:translateY(-1px)}}
.kpi.active{{background:var(--surface2);border-color:var(--accent)!important}}
.kpi.coming{{border-left-color:var(--coming-bar)}}.kpi.customs{{border-left-color:var(--customs-bar)}}.kpi.national{{border-left-color:var(--national-bar)}}.kpi.ofd{{border-left-color:var(--ofd-bar)}}.kpi.delivered{{border-left-color:var(--delivered-bar)}}.kpi.label{{border-left-color:var(--label-bar)}}
.kpi-icon{{font-size:16px;margin-bottom:6px}}.kpi-num{{font-family:'DM Mono',monospace;font-size:26px;font-weight:500;line-height:1}}.kpi-label{{font-size:9px;color:var(--muted2);text-transform:uppercase;letter-spacing:.12em;margin-top:5px}}

.mid-strip{{display:grid;grid-template-columns:1fr 1fr;gap:10px;padding:10px 32px 0}}
.panel{{background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:18px}}
.panel-title{{font-size:10px;text-transform:uppercase;letter-spacing:.14em;color:var(--muted2);margin-bottom:14px;display:flex;align-items:center;gap:8px}}
.panel-title span{{color:var(--text);font-size:12px;letter-spacing:0;text-transform:none}}
.chart-bars{{display:flex;flex-direction:column;gap:8px}}
.bar-row{{display:grid;grid-template-columns:110px 1fr 34px;align-items:center;gap:8px}}
.bar-name{{font-size:10px;color:var(--muted2);text-align:right}}.bar-track{{background:var(--surface2);border-radius:3px;height:14px;overflow:hidden}}.bar-fill{{height:100%;border-radius:3px;transition:width .9s cubic-bezier(.4,0,.2,1)}}.bar-pct{{font-family:'DM Mono',monospace;font-size:10px;color:var(--muted2);text-align:right}}
.mud-list{{display:flex;flex-direction:column;gap:6px;max-height:210px;overflow-y:auto}}.mud-list::-webkit-scrollbar{{width:3px}}.mud-list::-webkit-scrollbar-thumb{{background:var(--border2);border-radius:3px}}
.mud-row{{background:var(--surface2);border-left:2px solid var(--border2);border-radius:0 6px 6px 0;padding:8px 12px;display:flex;align-items:center;gap:10px;font-size:12px;transition:border-color .15s}}
.mud-row:hover{{border-left-color:var(--accent)}}
.mud-awb{{font-family:'DM Mono',monospace;font-size:10px;color:var(--muted2);min-width:105px}}.mud-pedido{{font-family:'DM Mono',monospace;font-size:10px;color:var(--accent);min-width:48px}}
.mud-trans{{display:flex;align-items:center;gap:5px;flex:1}}.mud-cat{{padding:2px 7px;border-radius:3px;font-size:9px;font-weight:600;text-transform:uppercase;letter-spacing:.06em}}.mud-arrow{{color:var(--muted);font-size:10px}}.mud-empty{{color:var(--muted);font-size:12px;text-align:center;padding:20px}}

.search-bar{{padding:14px 32px 0;display:flex;gap:8px;align-items:center;flex-wrap:wrap}}
.search-input{{flex:1;min-width:200px;background:var(--surface);border:1px solid var(--border);border-radius:7px;padding:9px 14px;color:var(--text);font-family:'DM Sans',sans-serif;font-size:12px;outline:none;transition:border-color .15s}}.search-input:focus{{border-color:var(--accent)}}.search-input::placeholder{{color:var(--muted)}}
.filter-tag{{padding:5px 12px;border-radius:20px;border:1px solid var(--border);background:transparent;color:var(--muted2);font-size:10px;cursor:pointer;transition:.15s;font-family:'DM Sans',sans-serif;text-transform:uppercase;letter-spacing:.08em}}
.filter-tag.active,.filter-tag:hover{{background:var(--accent);border-color:var(--accent);color:#fff}}

.table-wrap{{padding:12px 32px 32px;overflow-x:auto}}
table{{width:100%;border-collapse:separate;border-spacing:0 3px}}
thead th{{font-size:9px;text-transform:uppercase;letter-spacing:.14em;color:var(--muted);padding:6px 14px;text-align:left}}
tbody tr{{background:var(--surface);border-left:3px solid transparent;transition:background .12s,border-color .12s}}
tbody tr:hover{{background:var(--surface2)}}
tbody tr.alerta{{background:var(--alert)!important;border-left-color:#ef4444!important}}
tbody tr.alerta:hover{{background:#4a1010!important}}
td{{padding:10px 14px;vertical-align:middle;font-size:11px}}
td:first-child{{border-radius:0}}.td-last{{border-radius:0 6px 6px 0}}
.cat-badge{{display:inline-flex;align-items:center;gap:4px;padding:3px 9px;border-radius:4px;font-size:9px;font-weight:600;text-transform:uppercase;letter-spacing:.07em;white-space:nowrap}}
.pedido-badge{{font-family:'DM Mono',monospace;font-size:10px;color:var(--accent)}}
.evento-cell{{max-width:280px;color:var(--muted2);font-size:10px;line-height:1.5}}

.tab-bar{{display:flex;border-bottom:1px solid var(--border);padding:0 32px;margin-top:14px;background:var(--surface)}}
.tab-btn{{padding:11px 18px;font-size:10px;color:var(--muted2);cursor:pointer;border-bottom:2px solid transparent;margin-bottom:-1px;transition:.15s;text-transform:uppercase;letter-spacing:.1em;background:none;border-top:none;border-left:none;border-right:none;font-family:'DM Sans',sans-serif}}
.tab-btn.active{{color:var(--text);border-bottom-color:var(--accent)}}
.tab-pane{{display:none}}.tab-pane.active{{display:block}}

.mensal-wrap{{padding:20px 32px}}
.mensal-regiao{{background:var(--surface);border:1px solid var(--border);border-radius:8px;margin-bottom:16px;overflow:visible}}
.mensal-reg-header{{padding:12px 18px;font-size:11px;font-weight:600;letter-spacing:.08em;text-transform:uppercase;color:#fff;border-left:3px solid transparent;border-radius:0}}
.mensal-kpis{{display:grid;grid-template-columns:repeat(5,1fr);border-top:1px solid var(--border)}}
.mensal-kpi{{background:var(--surface2);padding:14px;text-align:center;border-right:1px solid var(--border)}}
.mensal-kpi:last-child{{border-right:none}}
.mensal-kpi-label{{font-size:9px;color:var(--muted2);text-transform:uppercase;letter-spacing:.1em;margin-bottom:5px}}
.mensal-kpi-val{{font-family:'DM Mono',monospace;font-size:17px;font-weight:500}}
.mensal-kpi-val.destaque{{color:var(--accent);font-size:13px}}
.mensal-sem-dados{{color:var(--muted);font-size:13px;text-align:center;padding:40px}}
.mensal-kpi-clickable{{cursor:pointer;position:relative}}
.mensal-kpi-clickable:hover{{background:var(--surface3)!important}}
.mensal-pop{{display:none;position:fixed;z-index:9999;background:var(--surface2);border:1px solid var(--accent);border-radius:8px;padding:12px 14px;min-width:230px;box-shadow:0 8px 24px #0009;font-size:12px;pointer-events:none}}
.mensal-pop.open{{display:block}}
.mensal-pop-awb{{font-family:'DM Mono',monospace;font-size:13px;font-weight:600;color:var(--text);margin-bottom:8px;padding-bottom:6px;border-bottom:1px solid var(--border2)}}
.mensal-pop-row{{display:flex;justify-content:space-between;gap:12px;padding:3px 0;color:var(--muted2)}}
.mensal-pop-row span:last-child{{color:var(--text);text-align:right}}
.sparkline-wrap{{padding:14px 18px 10px;border-top:1px solid var(--border)}}
.sparkline-title{{font-size:9px;color:var(--muted2);text-transform:uppercase;letter-spacing:.1em;margin-bottom:8px}}

thead th.sortable{{cursor:pointer;user-select:none}}
thead th.sortable:hover{{color:var(--text)}}
thead th.sort-asc::after{{content:" ↑"}}
thead th.sort-desc::after{{content:" ↓"}}

.dias-cell{{font-family:'DM Mono',monospace;font-size:10px;text-align:center}}
.heat-0{{color:var(--muted)}}
.heat-1{{color:#6ee7b7}}
.heat-2{{color:#fcd34d}}
.heat-3{{color:#fb923c}}
.heat-4{{color:#f87171;font-weight:600}}

.eta-cell{{font-family:'DM Mono',monospace;font-size:9px;color:var(--muted2);text-align:center}}

.tl-popup{{display:none;position:fixed;z-index:9999;background:var(--surface2);border:1px solid var(--border2);border-radius:8px;padding:14px 16px;min-width:260px;box-shadow:0 8px 24px #000a;pointer-events:none}}
.tl-popup.open{{display:block}}
.tl-awb{{font-family:'DM Mono',monospace;font-size:12px;font-weight:600;color:var(--text);margin-bottom:4px}}
.tl-produto{{font-size:10px;color:var(--accent);margin-bottom:10px}}
.tl-steps{{display:flex;flex-direction:column;gap:0}}

.detail-panel{{display:none;margin:0 32px 24px;background:var(--surface);border:1px solid var(--border);border-radius:10px;overflow:hidden}}
.detail-panel.open{{display:block}}
.detail-header{{padding:16px 20px;border-bottom:1px solid var(--border);display:flex;align-items:flex-start;justify-content:space-between;gap:16px;flex-wrap:wrap}}
.detail-awb{{font-family:'DM Mono',monospace;font-size:16px;font-weight:600;color:var(--text)}}
.detail-sub{{font-size:11px;color:var(--muted2);margin-top:3px}}
.detail-close{{background:none;border:none;color:var(--muted2);cursor:pointer;font-size:18px;padding:0;line-height:1;transition:.15s}}
.detail-close:hover{{color:var(--text)}}
.detail-meta{{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:0;border-bottom:1px solid var(--border)}}
.detail-meta-item{{padding:12px 20px;border-right:1px solid var(--border)}}
.detail-meta-item:last-child{{border-right:none}}
.detail-meta-label{{font-size:9px;text-transform:uppercase;letter-spacing:.1em;color:var(--muted2);margin-bottom:4px}}
.detail-meta-val{{font-size:12px;color:var(--text);font-weight:500}}
.detail-events{{padding:0}}
.detail-day-group{{border-bottom:1px solid var(--border)}}
.detail-day-group:last-child{{border-bottom:none}}
.detail-day-header{{padding:10px 20px;font-size:10px;font-weight:600;color:var(--muted2);text-transform:uppercase;letter-spacing:.1em;background:var(--surface2)}}
.detail-event-row{{display:grid;grid-template-columns:80px 1fr auto;align-items:start;gap:12px;padding:10px 20px;border-top:1px solid var(--border);transition:background .1s}}
.detail-event-row:first-child{{border-top:none}}
.detail-event-row:hover{{background:var(--surface2)}}
.detail-event-time{{font-family:'DM Mono',monospace;font-size:10px;color:var(--muted2);padding-top:1px}}
.detail-event-desc{{font-size:11px;color:var(--text);line-height:1.5}}
.detail-event-subdesc{{font-size:10px;color:var(--muted2);margin-top:2px}}
.detail-event-local{{font-size:10px;color:var(--muted2);text-align:right;white-space:nowrap;padding-top:1px}}
.tl-step{{display:flex;align-items:flex-start;gap:10px;padding:4px 0}}
.tl-dot{{width:8px;height:8px;border-radius:50%;flex-shrink:0;margin-top:3px}}
.tl-line-v{{width:1px;height:14px;background:var(--border2);margin-left:3.5px}}
.tl-cat{{font-size:10px;color:var(--text);flex:1}}
.tl-date{{font-family:'DM Mono',monospace;font-size:9px;color:var(--muted2)}}
.tl-step.current .tl-cat{{color:var(--accent);font-weight:600}}
.tl-step.current .tl-dot{{box-shadow:0 0 0 2px var(--accent)}}
</style></head><body>
<div class="header">
  <div class="logo"><div class="logo-dot"></div>FEDEX TRACKER</div>
  <div class="header-right">
    <div class="header-date">Atualizado em {data_consulta}</div>
    <div class="header-countdown" id="countdown">↻ próxima em —</div>
  </div>
</div>
<div class="tab-bar">
  <button class="tab-btn active" onclick="switchTab('painel')">Painel</button>
  <button class="tab-btn" onclick="switchTab('mensal')">Relatório Mensal</button>
</div>
<div id="tab-painel" class="tab-pane active">
<div class="kpi-strip" id="kpi-strip">
{kpi_html}</div>
<div class="mid-strip">
  <div class="panel">
    <div class="panel-title">Distribuição <span id="total-badge"></span></div>
    <div class="chart-bars" id="chart-bars"></div>
  </div>
  <div class="panel">
    <div class="panel-title">🔄 Mudanças detectadas <span id="mud-count"></span></div>
    <div class="mud-list" id="mud-list"></div>
  </div>
</div>
<div class="search-bar">
  <input class="search-input" type="text" id="search" placeholder="Buscar AWB ou pedido — busca exata abre painel de detalhes..." oninput="onSearch()">
  <button class="filter-tag active" onclick="filterCat('ALL')" data-ftag="ALL">Todos</button>
{filter_tags}</div>
<div id="detail-panel" class="detail-panel"></div>
<div class="table-wrap">
<table id="main-table"><thead><tr>
  <th class="sortable" data-col="awb">AWB</th>
  <th class="sortable" data-col="pedido">PEDIDO</th>
  <th class="sortable" data-col="cat">CATEGORIA</th>
  <th class="sortable" data-col="status">STATUS FEDEX</th>
  <th>ÚLTIMO EVENTO</th>
  <th class="sortable" data-col="dias">DIAS NO STATUS</th>
  <th class="sortable" data-col="eta">ETA</th>
  <th class="sortable" data-col="data">DATA</th>
</tr></thead>
<tbody id="tbody"></tbody></table>
<div style="padding:12px 0;font-size:11px;color:var(--muted);font-family:'DM Mono',monospace" id="row-count"></div>
</div>
<div id="tl-popup" class="tl-popup"></div>
</div>
<div id="tab-mensal" class="tab-pane">
  <div class="mensal-wrap" id="mensal-content"></div>
</div>
<script>
const ROWS=__ROWS_JSON__;
const MUDANCAS=__MUD_JSON__;
const CAT_COLORS={{"COMING TO BRAZIL":{{bg:"var(--coming)",fg:"var(--coming-fg)"}},"CUSTOMS INSPECTION":{{bg:"var(--customs)",fg:"var(--customs-fg)"}},"NATIONAL TRANSIT":{{bg:"var(--national)",fg:"var(--national-fg)"}},"OUT FOR DELIVERY":{{bg:"var(--ofd)",fg:"var(--ofd-fg)"}},"DELIVERED":{{bg:"var(--delivered)",fg:"var(--delivered-fg)"}},"LABEL CREATED":{{bg:"var(--label)",fg:"var(--label-fg)"}}}};
const ORDEM=["COMING TO BRAZIL","CUSTOMS INSPECTION","NATIONAL TRANSIT","OUT FOR DELIVERY","DELIVERED","LABEL CREATED"];
const ICONS={{"COMING TO BRAZIL":"✈","CUSTOMS INSPECTION":"🔍","NATIONAL TRANSIT":"🚚","OUT FOR DELIVERY":"📦","DELIVERED":"✅","LABEL CREATED":"🏷"}};
const CAT_BAR_COLOR={{"COMING TO BRAZIL":"#4f7dff","CUSTOMS INSPECTION":"#f59e0b","NATIONAL TRANSIT":"#10b981","OUT FOR DELIVERY":"#818cf8","DELIVERED":"#22c55e","LABEL CREATED":"#475569"}};
let sortCol=null, sortDir=1;

function heatClass(dias,cat){{
  if(dias===null||dias===undefined||cat==="DELIVERED") return 'heat-0';
  if(dias<=2) return 'heat-1';
  if(dias<=5) return 'heat-2';
  if(dias<=10) return 'heat-3';
  return 'heat-4';
}}

function sortBy(col){{
  if(sortCol===col) sortDir*=-1; else{{sortCol=col;sortDir=1;}}
  document.querySelectorAll('thead th').forEach(th=>th.classList.remove('sort-asc','sort-desc'));
  document.querySelectorAll('thead th[data-col="'+col+'"]').forEach(th=>th.classList.add(sortDir===1?'sort-asc':'sort-desc'));
  renderTable();
}}
document.querySelectorAll('thead th.sortable').forEach(th=>th.addEventListener('click',()=>sortBy(th.dataset.col)));

function onSearch(){{
  const q=document.getElementById("search").value.trim();
  const panel=document.getElementById("detail-panel");
  // Busca exata por AWB (12+ dígitos) ou pedido (5+ dígitos) → abre painel de detalhe
  const exact=ROWS.find(r=>r.awb===q||r.pedido===q||(r.awb.includes(q)&&q.length>=10)||(r.pedido===q&&q.length>=4));
  if(exact&&q.length>=4){{
    renderDetail(exact);
    panel.classList.add("open");
    document.getElementById("tbody").innerHTML="";
    document.getElementById("row-count").textContent="";
  }} else {{
    panel.classList.remove("open");
    renderTable();
  }}
}}

function _fmtData(d){{
  if(!d) return"";
  try{{
    const [y,m,dy]=d.split("-");
    const meses=["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"];
    const dias=["Dom","Seg","Ter","Qua","Qui","Sex","Sáb"];
    const dt=new Date(+y,+m-1,+dy);
    return dias[dt.getDay()]+", "+meses[+m-1]+" "+dy+"/"+y.slice(2);
  }}catch(e){{return d;}}
}}

function renderDetail(r){{
  const panel=document.getElementById("detail-panel");
  const bar=CAT_BAR_COLOR[r.cat]||"var(--accent)";
  const c=CAT_COLORS[r.cat]||{{bg:"#333",fg:"#aaa"}};

  // Agrupa eventos por data
  const byDate={{}};
  (r.eventos||[]).forEach(ev=>{{
    const d=ev.data||"";
    if(!byDate[d]) byDate[d]=[];
    byDate[d].push(ev);
  }});
  const dates=Object.keys(byDate).sort((a,b)=>b.localeCompare(a)); // mais recente primeiro

  let evHtml="";
  dates.forEach(d=>{{
    evHtml+=`<div class="detail-day-group">
      <div class="detail-day-header">${{_fmtData(d)}}</div>`;
    byDate[d].slice().reverse().forEach(ev=>{{
      evHtml+=`<div class="detail-event-row">
        <div class="detail-event-time">${{ev.hora||""}}</div>
        <div>
          <div class="detail-event-desc">${{ev.desc}}</div>
          ${{ev.subdesc?`<div class="detail-event-subdesc">${{ev.subdesc}}</div>`:""}}
        </div>
        <div class="detail-event-local">${{ev.local||""}}</div>
      </div>`;
    }});
    evHtml+="</div>";
  }});

  panel.innerHTML=`
    <div class="detail-header">
      <div>
        <div class="detail-awb">${{r.awb}}</div>
        <div class="detail-sub">${{r.produto||""}}${{r.pedido?" · Pedido "+r.pedido:""}}</div>
      </div>
      <div style="display:flex;align-items:center;gap:12px">
        <span class="cat-badge" style="background:${{c.bg}};color:${{c.fg}};font-size:11px;padding:5px 12px">${{ICONS[r.cat]||""}} ${{r.cat}}</span>
        <button class="detail-close" onclick="closeDetail()">✕</button>
      </div>
    </div>
    <div class="detail-meta">
      <div class="detail-meta-item">
        <div class="detail-meta-label">Status atual</div>
        <div class="detail-meta-val">${{r.status}}</div>
      </div>
      <div class="detail-meta-item">
        <div class="detail-meta-label">Destino</div>
        <div class="detail-meta-val">${{r.endereco||"—"}}</div>
      </div>
      <div class="detail-meta-item">
        <div class="detail-meta-label">Dias no status</div>
        <div class="detail-meta-val" style="color:${{r.dias>10?"#f87171":r.dias>5?"#fb923c":r.dias>2?"#fcd34d":"#6ee7b7"}}">${{r.dias!=null?r.dias+"d úteis":"—"}}</div>
      </div>
      <div class="detail-meta-item">
        <div class="detail-meta-label">ETA estimado</div>
        <div class="detail-meta-val" style="color:var(--accent)">${{r.eta||"—"}}</div>
      </div>
      <div class="detail-meta-item">
        <div class="detail-meta-label">Última atualização</div>
        <div class="detail-meta-val">${{r.evento.split("—").pop().trim().split(" ").slice(0,2).join(" ")||"—"}}</div>
      </div>
    </div>
    <div class="detail-events">${{evHtml||'<div style="padding:20px;color:var(--muted2);font-size:12px;text-align:center">Eventos não disponíveis — aguarde o próximo ciclo de consulta</div>'}}</div>
  `;
}}

function closeDetail(){{
  document.getElementById("detail-panel").classList.remove("open");
  document.getElementById("search").value="";
  renderTable();
}}

function renderTable(){{
  const q=document.getElementById("search").value.trim().toLowerCase();
  let rows=ROWS.filter(r=>{{
    const cm=currentCat==="ALL"||r.cat===currentCat;
    const qm=!q||r.awb.includes(q)||r.pedido.toLowerCase().includes(q)||(r.produto||"").toLowerCase().includes(q);
    return cm&&qm;
  }});
  if(sortCol){{
    rows=[...rows].sort((a,b)=>{{
      let va=a[sortCol]??'', vb=b[sortCol]??'';
      if(sortCol==='dias'){{va=va??9999;vb=vb??9999;return(va-vb)*sortDir;}}
      return String(va).localeCompare(String(vb))*sortDir;
    }});
  }}
  document.getElementById("tbody").innerHTML=rows.map((r,i)=>{{
    const hc=heatClass(r.dias,r.cat);
    const bar=CAT_BAR_COLOR[r.cat]||'transparent';
    const diasTxt=r.dias!=null?r.dias+'d':'—';
    const etaTxt=r.eta||'—';
    return`<tr class="${{r.alerta?'alerta':''}}" style="border-left:3px solid ${{bar}}"
      data-awb="${{r.awb}}" onmouseenter="showTimeline(event,this)" onmouseleave="hideTimeline()">
      <td style="font-family:'DM Mono',monospace;font-size:10px">${{r.awb}}</td>
      <td><span class="pedido-badge">${{r.pedido}}</span></td>
      <td>${{catBadge(r.cat)}}</td>
      <td style="font-size:10px;color:var(--muted2)">${{r.status}}</td>
      <td class="evento-cell">${{r.evento}}</td>
      <td class="dias-cell ${{hc}}">${{diasTxt}}</td>
      <td class="eta-cell">${{etaTxt}}</td>
      <td class="td-last" style="font-family:'DM Mono',monospace;font-size:10px;color:var(--muted)">${{r.data}}</td>
    </tr>`;
  }}).join("");
  document.getElementById("row-count").textContent=`${{rows.length}} de ${{ROWS.length}} remessas`;
}}

let _tlT=null;
function showTimeline(e,el){{
  clearTimeout(_tlT);
  _tlT=setTimeout(()=>{{
    const awb=el.dataset.awb;
    const r=ROWS.find(x=>x.awb===awb);
    const pop=document.getElementById('tl-popup');
    if(!r||!r.timeline||!r.timeline.length){{pop.classList.remove('open');return;}}
    let html=`<div class="tl-awb">${{r.awb}}</div>`;
    if(r.produto) html+=`<div class="tl-produto">${{r.produto}}</div>`;
    html+='<div class="tl-steps">';
    r.timeline.forEach((step,i)=>{{
      const isCur=i===r.timeline.length-1;
      const color=CAT_BAR_COLOR[step.cat]||'#475569';
      html+=`<div class="tl-step${{isCur?' current':''}}">
        <div style="display:flex;flex-direction:column;align-items:center">
          <div class="tl-dot" style="background:${{color}}"></div>
          ${{i<r.timeline.length-1?'<div class="tl-line-v"></div>':''}}
        </div>
        <div style="flex:1;padding-bottom:2px">
          <div class="tl-cat">${{ICONS[step.cat]||''}} ${{step.cat}}</div>
          <div class="tl-date">${{step.data||''}}</div>
        </div></div>`;
    }});
    html+='</div>';
    pop.innerHTML=html;
    const rect=el.getBoundingClientRect();
    const pw=270;
    let left=rect.right+8;
    if(left+pw>window.innerWidth-8) left=rect.left-pw-8;
    let top=rect.top;
    pop.style.left=Math.max(8,left)+'px';
    pop.style.top=top+'px';
    pop.style.maxHeight=(window.innerHeight-top-16)+'px';
    pop.style.overflowY='auto';
    pop.classList.add('open');
  }},100);
}}
function hideTimeline(){{
  clearTimeout(_tlT);
  _tlT=setTimeout(()=>document.getElementById('tl-popup')?.classList.remove('open'),60);
}}
function renderChart(){{
  const total=ROWS.length;const cats={{}};
  ROWS.forEach(r=>cats[r.cat]=(cats[r.cat]||0)+1);
  const max=Math.max(...Object.values(cats),1);
  document.getElementById("total-badge").textContent=total+" remessas";
  document.getElementById("chart-bars").innerHTML=ORDEM.map(cat=>{{
    const n=cats[cat]||0;const pct=Math.round(n/total*100);const w=Math.round(n/max*100);
    const c=CAT_COLORS[cat]||{{fg:"#aaa"}};
    const short=cat.replace("COMING TO BRAZIL","COMING").replace("CUSTOMS INSPECTION","CUSTOMS").replace("NATIONAL TRANSIT","NATIONAL").replace("OUT FOR DELIVERY","OUT FOR DEL.").replace("LABEL CREATED","LABEL");
    return`<div class="bar-row"><div class="bar-name">${{short}}</div><div class="bar-track"><div class="bar-fill" style="width:${{w}}%;background:${{c.fg}};opacity:.7"></div></div><div class="bar-pct">${{pct}}%</div></div>`;
  }}).join("");
}}
function catBadge(cat){{const c=CAT_COLORS[cat]||{{bg:"#333",fg:"#aaa"}};return`<span class="cat-badge" style="background:${{c.bg}};color:${{c.fg}}">${{ICONS[cat]||""}} ${{cat}}</span>`;}}
let currentCat="ALL";

function renderMudancas(){{
  const el=document.getElementById("mud-list");
  document.getElementById("mud-count").textContent=MUDANCAS.length?MUDANCAS.length+" hoje":"";
  if(!MUDANCAS.length){{el.innerHTML='<div class="mud-empty">Nenhuma mudança detectada hoje</div>';return;}}
  el.innerHTML=MUDANCAS.map(r=>{{
    const trans=(r[2]||"").split("→").map(s=>s.trim());
    const ant=trans[0]||"";const atu=trans[1]||"";
    const ca=CAT_COLORS[ant]||{{bg:"#333",fg:"#aaa"}};const cb=CAT_COLORS[atu]||{{bg:"#333",fg:"#aaa"}};
    return`<div class="mud-row"><div class="mud-awb">${{r[0]}}</div><div class="mud-pedido">${{r[1]||""}}</div><div class="mud-trans"><span class="mud-cat" style="background:${{ca.bg}};color:${{ca.fg}}">${{ant}}</span><span class="mud-arrow">→</span><span class="mud-cat" style="background:${{cb.bg}};color:${{cb.fg}}">${{atu}}</span></div></div>`;
  }}).join("");
}}
function filterCat(cat){{
  currentCat=cat;
  document.querySelectorAll("[data-ftag]").forEach(b=>b.classList.toggle("active",b.dataset.ftag===cat));
  renderTable();
}}
renderChart();renderMudancas();renderTable();

// Countdown regressivo para próxima atualização
(function(){{
  const meta = document.querySelector('meta[http-equiv="refresh"]');
  if (!meta) return;
  let secs = parseInt(meta.content) || 3600;
  const el = document.getElementById('countdown');
  function fmt(s){{ const m=Math.floor(s/60),r=s%60; return '↻ próxima em '+m+'min '+(r<10?'0':'')+r+'s'; }}
  if(el) el.textContent = fmt(secs);
  const iv = setInterval(function(){{
    secs--;
    if(secs<=0){{ clearInterval(iv); if(el) el.textContent='↻ atualizando...'; return; }}
    if(el) el.textContent = fmt(secs);
  }}, 1000);
}})();

function switchTab(tab){{
  document.querySelectorAll('.tab-pane').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active'));
  document.getElementById('tab-'+tab).classList.add('active');
  event.target.classList.add('active');
}}

const MENSAL=__MENSAL_JSON__;
const CAT_COR_REG={{"América do Norte":"#BDD7EE","Europa":"#C6EFCE","Oriente Médio":"#FFE699","Ásia":"#F4CCFF","Outros":"#D9D9D9"}};
let mensalIdx = 0;

function _kpiPopover(label, awbObj) {{
  if (!awbObj) return `<div class="mensal-kpi"><div class="mensal-kpi-label">${{label}}</div><div class="mensal-kpi-val">—</div></div>`;
  const dias = awbObj.lt !== null ? awbObj.lt + 'd' : '—';
  const tip = `${{awbObj.awb}}${{awbObj.pedido ? ' · ' + awbObj.pedido : ''}}${{awbObj.produto ? '\\n' + awbObj.produto : ''}}`;
  return `<div class="mensal-kpi mensal-kpi-clickable" onclick="togglePop(this)" title="${{tip}}">
    <div class="mensal-kpi-label">${{label}} <span style="font-size:9px;opacity:.6">ⓘ</span></div>
    <div class="mensal-kpi-val">${{dias}}</div>
    <div class="mensal-pop">
      <div class="mensal-pop-awb">${{awbObj.awb}}</div>
      ${{awbObj.pedido ? `<div class="mensal-pop-row"><span>Pedido</span><span>${{awbObj.pedido}}</span></div>` : ''}}
      ${{awbObj.produto ? `<div class="mensal-pop-row"><span>Produto</span><span style="color:var(--accent)">${{awbObj.produto}}</span></div>` : ''}}
      <div class="mensal-pop-row"><span>Lead time</span><span style="font-weight:600">${{dias}} úteis</span></div>
    </div>
  </div>`;
}}

function togglePop(el) {{
  const pop = el.querySelector('.mensal-pop');
  if (!pop) return;
  const isOpen = pop.classList.contains('open');
  document.querySelectorAll('.mensal-pop.open').forEach(p => p.classList.remove('open'));
  if (!isOpen) {{
    pop.classList.add('open');
    const rect = el.getBoundingClientRect();
    const popW = 230;
    let left = rect.left + rect.width / 2 - popW / 2;
    left = Math.max(8, Math.min(left, window.innerWidth - popW - 8));
    pop.style.top = (rect.bottom + 8) + 'px';
    pop.style.left = left + 'px';
    pop.style.width = popW + 'px';
  }}
}}

document.addEventListener('click', function(e) {{
  if (!e.target.closest('.mensal-kpi-clickable')) {{
    document.querySelectorAll('.mensal-pop.open').forEach(p => p.classList.remove('open'));
  }}
}});

function renderMensal(){{
  const el=document.getElementById('mensal-content');
  if(!MENSAL||!MENSAL.meses||MENSAL.meses.length===0){{
    el.innerHTML='<div class="mensal-sem-dados">Nenhum dado de entrega disponível ainda.<br>As remessas entregues aparecerão aqui automaticamente.</div>';
    return;
  }}
  const meses = MENSAL.meses;
  const m = meses[mensalIdx];

  // Navegação
  const navH = `<div style="display:flex;align-items:center;gap:12px;margin-bottom:20px;flex-wrap:wrap">
    <button onclick="mudaMes(-1)" style="background:var(--panel-bg);border:1px solid var(--border);color:var(--fg);padding:6px 14px;border-radius:6px;cursor:pointer;font-size:13px" ${{mensalIdx===meses.length-1?'disabled':''}}>&larr; Anterior</button>
    <div style="display:flex;gap:6px;flex-wrap:wrap">
      ${{meses.map((mm,i)=>`<button onclick="irMes(${{i}})" style="background:${{i===mensalIdx?'var(--accent)':'var(--panel-bg)'}};border:1px solid ${{i===mensalIdx?'var(--accent)':'var(--border)'}};color:${{i===mensalIdx?'#fff':'var(--fg)'}};padding:4px 10px;border-radius:6px;cursor:pointer;font-size:12px">${{mm.periodo}}</button>`).join('')}}
    </div>
    <button onclick="mudaMes(1)" style="background:var(--panel-bg);border:1px solid var(--border);color:var(--fg);padding:6px 14px;border-radius:6px;cursor:pointer;font-size:13px" ${{mensalIdx===0?'disabled':''}}>Próximo &rarr;</button>
  </div>`;

  let h = navH;
  h += `<div style="margin-bottom:20px">
    <span style="font-size:13px;color:var(--muted)">Período: </span>
    <span style="font-size:15px;font-weight:600">${{m.periodo}}</span>
    <span style="margin-left:20px;font-size:13px;color:var(--muted)">Total entregue: </span>
    <span style="font-size:15px;font-weight:600;color:var(--accent)">${{m.total}} envios</span>
  </div>`;

  if(!m.regioes||m.regioes.length===0){{
    h+='<div class="mensal-sem-dados">Nenhuma remessa entregue neste mês.</div>';
    el.innerHTML=h; return;
  }}

  // Sparkline — volume por mês
  if(meses.length>1){{
    const maxVol=Math.max(...meses.map(mm=>mm.total),1);
    const barW=Math.max(14,Math.floor(180/meses.length));
    let spark='<div class="sparkline-wrap"><div class="sparkline-title">Volume histórico de entregas</div>';
    spark+='<div style="display:flex;align-items:flex-end;gap:3px;height:40px">';
    [...meses].reverse().forEach((mm,i)=>{{
      const h2=Math.max(4,Math.round(mm.total/maxVol*36));
      const isActive=meses.length-1-i===mensalIdx;
      spark+=`<div title="${{mm.periodo}}: ${{mm.total}} envios" onclick="irMes(${{meses.length-1-i}})"
        style="width:${{barW}}px;height:${{h2}}px;background:${{isActive?'var(--accent)':'var(--border2)'}};border-radius:2px 2px 0 0;cursor:pointer;transition:.15s;flex-shrink:0"
        onmouseover="this.style.background='var(--accent)'" onmouseout="this.style.background='${{isActive?'var(--accent)':'var(--border2)'}}'"></div>`;
    }});
    spark+='</div></div>';
    h+=spark;
  }}

  const REG_COLORS={{"América do Norte":"#4f7dff","Europa":"#10b981","Oriente Médio":"#f59e0b","Ásia":"#a78bfa","Outros":"#475569"}};
  const REG_BG={{"América do Norte":"#0d1a36","Europa":"#062010","Oriente Médio":"#1c1200","Ásia":"#150d2e","Outros":"#111827"}};

  m.regioes.forEach(r=>{{
    const cor=REG_COLORS[r.nome]||'#4f7dff';
    const bg=REG_BG[r.nome]||'#111827';
    h+=`<div class="mensal-regiao">
      <div class="mensal-reg-header" style="background:${{bg}};border-left-color:${{cor}}">${{r.nome}} — ${{r.total}} envio${{r.total>1?'s':''}}</div>
      <div class="mensal-kpis">
        <div class="mensal-kpi"><div class="mensal-kpi-label">Total</div><div class="mensal-kpi-val">${{r.total}}</div></div>
        <div class="mensal-kpi"><div class="mensal-kpi-label">Lead time médio</div><div class="mensal-kpi-val">${{r.media_lt !== null && r.media_lt !== "-" ? r.media_lt+"d" : "—"}}</div></div>
        ${{_kpiPopover('Mínimo', r.awb_min)}}
        ${{_kpiPopover('Máximo', r.awb_max)}}
        <div class="mensal-kpi"><div class="mensal-kpi-label">Produto top</div><div class="mensal-kpi-val destaque">${{r.produto_top}}</div></div>
      </div>
      ${{r.awbs && r.awbs.length>0 ? `<details style="margin-top:8px"><summary style="cursor:pointer;font-size:12px;color:var(--muted);user-select:none">Ver remessas (${{r.awbs.length}})</summary><div style="margin-top:8px;display:flex;flex-wrap:wrap;gap:6px">${{r.awbs.map(a=>`<span style="font-family:'DM Mono',monospace;font-size:11px;background:var(--bg);padding:2px 8px;border-radius:4px;border:1px solid var(--border)" title="${{a.produto}}">${{a.awb}}${{a.pedido?' · '+a.pedido:''}}${{a.lt!=null?' ('+a.lt+'d)':''}}</span>`).join('')}}</div></details>` : ''}}
    </div>`;
  }});
  el.innerHTML=h;
}}
function mudaMes(dir){{
  const total=MENSAL.meses.length;
  mensalIdx=Math.max(0,Math.min(total-1,mensalIdx-dir));
  renderMensal();
}}
function irMes(i){{
  mensalIdx=i;
  renderMensal();
}}
renderMensal();
</script></body></html>"""

    # ── Monta JSON mensal a partir do histórico completo ──────────────────────
    # Lê historico_status.xlsx para ter todas as entregas já registradas,
    # mesmo as que saíram do snapshot atual (awbs removidos do awbs.xlsx).
    from datetime import datetime as _dtm2
    REGIOES_DISPLAY = ["América do Norte", "Europa", "Oriente Médio", "Ásia", "Outros"]

    def _build_mensal_json(arquivo_hist: str, df_atual: pd.DataFrame) -> dict:
        """
        Constrói o JSON com histórico mensal de entregas.
        Fonte primária: historico_status.xlsx (acumula todos os ciclos).
        Fallback: df_atual (snapshot corrente).
        """
        frames = []

        # Tenta carregar histórico
        if arquivo_hist:
            try:
                df_h = pd.read_excel(arquivo_hist)
                df_h.columns = df_h.columns.str.strip()
                frames.append(df_h)
            except Exception as _eh:
                logger.warning(f"Mensal: não foi possível ler histórico: {_eh}")

        # Inclui snapshot atual também (cobre AWBs entregues nesse ciclo)
        frames.append(df_atual)

        if not frames:
            return {"meses": []}

        df_all = pd.concat(frames, ignore_index=True)
        df_all.columns = df_all.columns.str.strip()

        # Normaliza nome da coluna de categoria
        if "CATEGORIA" not in df_all.columns and "CATEGORIA_NORM" in df_all.columns:
            df_all["CATEGORIA"] = df_all["CATEGORIA_NORM"]

        # Filtra apenas entregues
        mask_delivered = df_all.get("CATEGORIA", pd.Series(dtype=str)).astype(str).str.upper().str.contains("DELIVERED", na=False)
        df_del = df_all[mask_delivered].copy()

        if df_del.empty or "DATA_ENTREGA" not in df_del.columns:
            return {"meses": []}

        df_del["_de"] = pd.to_datetime(df_del["DATA_ENTREGA"], errors="coerce")
        df_del = df_del[df_del["_de"].notna()]
        if df_del.empty:
            return {"meses": []}

        # Remove duplicatas: mesmo AWB entregue só conta uma vez (última ocorrência)
        if "AWB" in df_del.columns:
            df_del = df_del.sort_values("_de").drop_duplicates(subset=["AWB"], keep="last")

        df_del["_mes"] = df_del["_de"].dt.month
        df_del["_ano"] = df_del["_de"].dt.year
        df_del["_chave"] = df_del["_ano"].astype(str) + "-" + df_del["_mes"].astype(str).str.zfill(2)

        nomes_mes = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
                     "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

        # Ordena meses do mais recente para o mais antigo
        chaves = sorted(df_del["_chave"].unique(), reverse=True)
        meses_out = []

        for chave in chaves:
            df_m = df_del[df_del["_chave"] == chave]
            ano_m = int(chave.split("-")[0])
            mes_m = int(chave.split("-")[1])
            periodo = f"{nomes_mes[mes_m - 1]} {ano_m}"

            regioes_out = []
            for reg in REGIOES_DISPLAY:
                if "REGIAO" not in df_m.columns:
                    df_r = pd.DataFrame()
                else:
                    df_r = df_m[df_m["REGIAO"].astype(str).str.strip() == reg]
                if df_r.empty:
                    continue

                lt = pd.to_numeric(df_r["LEAD_TIME_DIAS"], errors="coerce") if "LEAD_TIME_DIAS" in df_r.columns else pd.Series(dtype=float)
                lt_valid = lt.dropna()
                prod_top = str(df_r["PRODUTO"].value_counts().index[0]) if "PRODUTO" in df_r.columns and not df_r["PRODUTO"].isna().all() else "-"

                # Lista de AWBs com lead time individual para identificar min/max
                awbs_list = []
                for _, row in df_r.iterrows():
                    lt_val = lt.get(row.name)
                    awbs_list.append({
                        "awb":     str(row.get("AWB", "")),
                        "pedido":  str(row.get("PEDIDO", "")).replace(".0","") if pd.notna(row.get("PEDIDO","")) else "",
                        "produto": str(row.get("PRODUTO", "")) if pd.notna(row.get("PRODUTO","")) else "",
                        "lt":      int(lt_val) if pd.notna(lt_val) else None,
                    })

                # Identifica AWBs de min e max lead time
                awbs_com_lt = [a for a in awbs_list if a["lt"] is not None]
                awb_min = min(awbs_com_lt, key=lambda a: a["lt"]) if awbs_com_lt else None
                awb_max = max(awbs_com_lt, key=lambda a: a["lt"]) if awbs_com_lt else None

                regioes_out.append({
                    "nome":       reg,
                    "total":      len(df_r),
                    "media_lt":   round(float(lt_valid.mean())) if len(lt_valid) > 0 else "-",
                    "min_lt":     int(lt_valid.min())           if len(lt_valid) > 0 else "-",
                    "max_lt":     int(lt_valid.max())           if len(lt_valid) > 0 else "-",
                    "awb_min":    awb_min,
                    "awb_max":    awb_max,
                    "produto_top": prod_top,
                    "awbs":       awbs_list,
                })

            meses_out.append({
                "periodo": periodo,
                "total":   len(df_m),
                "regioes": regioes_out,
            })

        return {"meses": meses_out}

    mensal_json = _build_mensal_json(arquivo_historico, df)

    html = html.replace("__ROWS_JSON__", json.dumps(rows_json, ensure_ascii=False))
    html = html.replace("__MUD_JSON__", json.dumps(mudancas_rows, ensure_ascii=False))
    html = html.replace("__MENSAL_JSON__", json.dumps(mensal_json, ensure_ascii=False))

    with open(arquivo_html, "w", encoding="utf-8") as f:
        f.write(html)
    logger.info(f"HTML salvo → '{arquivo_html}'")

def executar():
    config = Config()
    logger = setup_logger(config.arquivo_log)
    logger.info("=" * 60)
    logger.info("  FedEx Tracker — API Oficial")
    logger.info("=" * 60)

    try:
        lista_awbs = carregar_awbs(config.arquivo_awbs)
    except (FileNotFoundError, ValueError) as e:
        logger.error(str(e))
        sys.exit(1)

    if not lista_awbs:
        logger.error("Nenhum AWB válido encontrado.")
        sys.exit(1)

    try:
        token_mgr = TokenManager(config)
        token     = token_mgr.get_token()
    except Exception as e:
        logger.error(f"Falha na autenticação: {e}")
        sys.exit(1)

    logger.info(f"Iniciando rastreio de {len(lista_awbs)} AWBs...")

    resultados = []
    with ThreadPoolExecutor(max_workers=config.max_workers) as executor:
        futures = {
            executor.submit(consultar_awb, orig, limpo, token, config, ped, prod): limpo
            for orig, limpo, ped, prod in lista_awbs
        }
        with tqdm(total=len(lista_awbs), desc="Rastreando", unit="AWB", ncols=72) as pbar:
            for future in as_completed(futures):
                r = future.result()
                resultados.append(r)
                pbar.set_postfix_str(f"{r['AWB']} → {r['CATEGORIA'][:18]}")
                pbar.update(1)

    try:
        salvar_resultados(resultados, config.arquivo_historico, config.arquivo_ultimo_status)
    except Exception as e:
        logger.error(f"Erro ao salvar: {e}")

    # Gera relatório HTML junto com o Excel
    try:
        arquivo_html = config.arquivo_ultimo_status.replace(".xlsx", ".html")
        gerar_html_relatorio(config.arquivo_ultimo_status, arquivo_html, config.arquivo_historico)
    except Exception as e:
        logger.error(f"Erro ao gerar HTML: {e}")

    logger.info(gerar_resumo(resultados))


def get_ip_local() -> str:
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return "localhost"


def iniciar_servidor(pasta: str, porta: int) -> None:
    """Sobe servidor HTTP na rede local — roda em thread separada."""
    class Handler(http.server.SimpleHTTPRequestHandler):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, directory=pasta, **kwargs)

        def do_GET(self):
            if self.path == "/":
                self.send_response(302)
                self.send_header("Location", "/ultimo_status_gerado.html")
                self.end_headers()
            else:
                super().do_GET()

        def log_message(self, format, *args):
            pass  # silencia logs do servidor no console

    def _run():
        with socketserver.TCPServer(("", porta), Handler) as httpd:
            httpd.allow_reuse_address = True
            httpd.serve_forever()

    t = threading.Thread(target=_run, daemon=True)
    t.start()


if __name__ == "__main__":
    config = Config()
    pasta  = str(Path(config.arquivo_ultimo_status).parent.resolve())
    ip     = get_ip_local()

    # Sobe o servidor antes da primeira consulta
    iniciar_servidor(pasta, config.porta_servidor)

    def _print_banner():
        print()
        print("=" * 55)
        print("  FEDEX TRACKER — SERVIDOR ATIVO")
        print("=" * 55)
        print(f"  Seu computador : http://localhost:{config.porta_servidor}")
        print(f"  Rede local     : http://{ip}:{config.porta_servidor}")
        print("=" * 55)
        print("  Ctrl+C para parar.")
        print()

    _print_banner()

    while True:
        try:
            executar()
        except Exception as e:
            logging.getLogger("fedex_tracker").error(f"Erro no ciclo: {e}")
        _print_banner()
        print(f"  Proxima consulta em 1 hora... ({datetime.now().strftime('%H:%M')})")
        time.sleep(3600)
