import re
import json
import sys
import os
import hashlib
import shutil
import logging
import time
import socket
import threading
import http.server
import socketserver
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import List, TypedDict

import requests
import pandas as pd
from tqdm import tqdm

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass


# ==========================================================
# SEGURANÇA — UTILITÁRIOS
# ==========================================================

def _mascarar_awb(awb: str) -> str:
    s = str(awb)
    if len(s) >= 8:
        return s[:4] + "*" * (len(s) - 8) + s[-4:]
    return "****"

def _mascarar_pedido(pedido: str) -> str:
    s = str(pedido)
    if len(s) >= 3:
        return s[:3] + "*" * (len(s) - 3)
    return "***"


# ==========================================================
# CONFIGURAÇÕES
# ==========================================================

def _env(key: str, default: str = "") -> str:
    return os.environ.get(key, default)

def _carregar_env():
    try:
        from dotenv import load_dotenv
        pasta_script = Path(__file__).parent.resolve()
        env_script = pasta_script / ".env"
        if env_script.exists():
            load_dotenv(env_script)
            return
        env_cwd = Path.cwd() / ".env"
        if env_cwd.exists():
            load_dotenv(env_cwd)
    except ImportError:
        pass

_carregar_env()

@dataclass
class Config:
    client_id:     str = field(default_factory=lambda: _env("FEDEX_CLIENT_ID",     ""))
    client_secret: str = field(default_factory=lambda: _env("FEDEX_CLIENT_SECRET", ""))

    url_token:    str = "https://apis.fedex.com/oauth/token"
    url_tracking: str = "https://apis.fedex.com/track/v1/trackingnumbers"

    arquivo_awbs:          str = "awbs.xlsx"
    arquivo_historico:     str = "historico_status.xlsx"
    arquivo_ultimo_status: str = "ultimo_status_gerado.xlsx"
    arquivo_log:           str = "tracking.log"

    max_workers: int   = 5
    delay_entre: float = 0.5
    timeout:     int   = 30
    porta_servidor: int = 8888

    def validar(self) -> None:
        if not self.client_id or not self.client_secret:
            pasta = Path(__file__).parent.resolve()
            raise ValueError(
                "\n" + "=" * 60 +
                "\n  ERRO: Credenciais FedEx não configuradas!" +
                "\n" +
                "\n  OPÇÃO 1 — Crie o arquivo .env nesta pasta:" +
                f"\n  {pasta}\\.env" +
                "\n" +
                "\n  Com o conteúdo:" +
                "\n    FEDEX_CLIENT_ID=seu_client_id" +
                "\n    FEDEX_CLIENT_SECRET=seu_client_secret" +
                "\n" +
                "\n  OPÇÃO 2 — Edite diretamente no código (linha ~70):" +
                "\n    client_id = 'seu_client_id'" +
                "\n    client_secret = 'seu_client_secret'" +
                "\n" +
                "\n  Obtenha suas credenciais em: developer.fedex.com" +
                "\n" + "=" * 60
            )


# ==========================================================
# MODELS
# ==========================================================

class ResultadoAWB(TypedDict, total=False):
    AWB: str
    AWB_ORIGINAL: str
    CATEGORIA: str
    STATUS_FEDEX: str
    ULTIMO_EVENTO: str
    MOTIVO_CATEGORIA: str
    DATA_CONSULTA: str
    SUCESSO: bool
    DATA_CHEGADA_MEMPHIS: str
    PEDIDO: str
    PRODUTO: str
    PAIS_ORIGEM: str
    REGIAO: str
    DATA_CRIACAO: str
    DATA_ENTREGA: str
    LEAD_TIME_DIAS: int
    TIMELINE_JSON: str
    EVENTOS_JSON: str
    ENDERECO_ENTREGA: str


# ==========================================================
# LOGGER
# ==========================================================

def setup_logger(log_file: str) -> logging.Logger:
    logger = logging.getLogger("fedex_tracker")
    if logger.handlers:
        return logger
    logger.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s [%(levelname)-8s] %(message)s", "%d/%m/%Y %H:%M:%S")

    from logging.handlers import RotatingFileHandler
    fh = RotatingFileHandler(
        log_file, encoding="utf-8",
        maxBytes=5 * 1024 * 1024,
        backupCount=7,
    )
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger


# ==========================================================
# AWB — LIMPEZA E VALIDAÇÃO
# ==========================================================

def limpar_awb(awb: str) -> str:
    return re.sub(r'\s+', '', awb.strip())

def validar_awb(awb: str) -> bool:
    return bool(re.match(r'^\d{12,22}$', awb))


# ==========================================================
# CLASSIFICAÇÃO — 5 CATEGORIAS
# ==========================================================

HUBS_INTERNACIONAIS = [
    "MEMPHIS", "MEM", "INDIANAPOLIS", "IND", "MIAMI", "MIA",
    "NEWARK", "EWR", "CHICAGO", "ORD", "LOS ANGELES", "LAX",
    "AGUADILLA", "BQN",
    "FRANKFURT", "FRA", "PARIS", "CDG", "COLOGNE", "CGN",
    "LONDON", "LHR", "AMSTERDAM", "AMS", "LIEGE", "LGG",
    "ENFIELD",
    "DUBAI", "DXB",
    "HONG KONG", "HKG", "SHANGHAI", "PVG", "TOKYO", "NRT",
    "SINGAPORE", "SIN", "MUMBAI", "BOM", "DELHI", "DEL",
    "BANGALORE", "BLR", "CHENNAI", "MAA", "GUANGZHOU", "CAN",
    "BEIJING", "PEK", "SEOUL", "ICN", "TAIPEI", "TPE",
    "BANGKOK", "BKK", "KUALA LUMPUR", "KUL",
    "TORONTO", "YYZ", "MEXICO CITY", "MEX",
]

TERMOS_BRASIL = [
    "CAMPINAS", "VCP", "VIRACOPOS", "BRAZIL", "BRASIL",
    "BARUERI", "SAO PAULO", "SAO PAULO", "RIO DE JANEIRO",
    "CURITIBA", "PORTO ALEGRE", "BELO HORIZONTE", "BRASILIA",
    "MANAUS", "RECIFE", "FORTALEZA", "SALVADOR", "BR",
]

TERMOS_ATRASO_ALFANDEGA = [
    "ATRASO NA LIBERACAO",
    "ATRASO NA LIBERAÇÃO",
    "ATRASO NA LIBERACAO - IMPORTACAO",
    "ATRASO NA LIBERAÇÃO - IMPORTAÇÃO",
    "CLEARANCE DELAY",
    "CUSTOMS DELAY",
    "HELD BY CUSTOMS",
    "CUSTOMS HOLD",
    "RETIDO NA ALFANDEGA",
    "RETIDO NA ALFÂNDEGA",
    "AGUARDANDO LIBERACAO",
    "AGUARDANDO LIBERAÇÃO",
    "PENDING CUSTOMS",
    "CLEARANCE IN PROGRESS",
    "IMPORT HOLD",
]

TERMOS_LIBERACAO = [
    "INTERNATIONAL SHIPMENT RELEASE",
    "RELEASED BY CUSTOMS",
    "CUSTOMS CLEARED",
    "CLEARANCE COMPLETED",
    "LIBERACAO DA REMESSA INTERNACIONAL",
    "LIBERAÇÃO DA REMESSA INTERNACIONAL",
    "REMESSA LIBERADA",
    "LIBERADO PELA ALFANDEGA",
    "LIBERADO PELA ALFÂNDEGA",
]

TERMOS_OUT_FOR_DELIVERY = [
    "ON FEDEX VEHICLE FOR DELIVERY",
    "EM UM VEICULO DA FEDEX PARA ENTREGA",
    "EM UM VEÍCULO DA FEDEX PARA ENTREGA",
    "OUT FOR DELIVERY",
    "EM VEICULO PARA ENTREGA",
    "EM VEÍCULO PARA ENTREGA",
    "SAIU PARA ENTREGA",
    "WITH DELIVERY COURIER",
]

TERMOS_ENTREGUE = [
    "DELIVERED",
    "ENTREGUE",
    "DELIVERY MADE",
]


def _textos_eventos(eventos: list) -> list:
    textos = []
    for ev in eventos:
        textos.append(ev.get("eventDescription", "").upper())
        textos.append(ev.get("scanLocation", {}).get("city", "").upper())
        textos.append(ev.get("scanLocation", {}).get("stateOrProvinceCode", "").upper())
        textos.append(ev.get("scanLocation", {}).get("countryCode", "").upper())
    return textos


def _contem(textos: list, termos: list) -> bool:
    for texto in textos:
        for termo in termos:
            if termo in texto:
                return True
    return False


def classificar(eventos: list, status_fedex: str) -> tuple:
    status_upper         = status_fedex.upper()
    ultimo_desc          = eventos[0].get("eventDescription", "").upper() if eventos else ""
    ultimo_local_city    = eventos[0].get("scanLocation", {}).get("city", "").upper() if eventos else ""
    ultimo_local_country = eventos[0].get("scanLocation", {}).get("countryCode", "").upper() if eventos else ""
    ultimo_textos        = [status_upper, ultimo_desc, ultimo_local_city, ultimo_local_country]
    todos_textos = _textos_eventos(eventos)

    GATILHOS_ENTREGUE = ["DELIVERED", "ENTREGUE", "DELIVERY MADE"]
    if _contem(ultimo_textos, GATILHOS_ENTREGUE):
        return "DELIVERED", "Entrega confirmada no destino"

    GATILHOS_OFD = [
        "ON FEDEX VEHICLE FOR DELIVERY",
        "EM UM VEICULO DA FEDEX PARA ENTREGA",
        "EM UM VEÍCULO DA FEDEX PARA ENTREGA",
        "OUT FOR DELIVERY",
        "EM VEICULO PARA ENTREGA",
        "EM VEÍCULO PARA ENTREGA",
        "SAIU PARA ENTREGA",
        "WITH DELIVERY COURIER",
    ]
    if _contem(ultimo_textos, GATILHOS_OFD):
        return "OUT FOR DELIVERY", "Pacote saiu para entrega — chegada prevista hoje"

    TERMOS_SAO_PAULO = ["SAO PAULO", "SÃO PAULO"]
    ultimo_city = eventos[0].get("scanLocation", {}).get("city", "").upper() if eventos else ""
    ultimo_state = eventos[0].get("scanLocation", {}).get("stateOrProvinceCode", "").upper() if eventos else ""
    if any(t in ultimo_city for t in TERMOS_SAO_PAULO) and ultimo_state in ("SP", ""):
        return "OUT FOR DELIVERY", "Remessa em São Paulo/SP — saindo para entrega"

    if _contem(ultimo_textos, TERMOS_ATRASO_ALFANDEGA):
        return "CUSTOMS INSPECTION", "Atraso na liberação alfandegária — retido pela Receita Federal"

    GATILHOS_LIBERACAO = [
        "LIBERACAO DA REMESSA INTERNACIONAL",
        "LIBERAÇÃO DA REMESSA INTERNACIONAL",
        "INTERNATIONAL SHIPMENT RELEASE",
        "RELEASED BY CUSTOMS",
        "CUSTOMS CLEARED",
        "CLEARANCE COMPLETED",
        "REMESSA LIBERADA",
        "LIBERADO PELA ALFANDEGA",
        "LIBERADO PELA ALFÂNDEGA",
    ]
    tem_brasil_historico = _contem(todos_textos, TERMOS_BRASIL)
    if _contem([status_upper], GATILHOS_LIBERACAO) and tem_brasil_historico:
        return "NATIONAL TRANSIT", "Liberação alfandegária confirmada — em trânsito nacional"
    elif _contem([status_upper], GATILHOS_LIBERACAO) and not tem_brasil_historico:
        return "COMING TO BRAZIL", "Liberação detectada mas Brasil não confirmado no histórico"

    GATILHOS_ORIGEM = [
        "DEIXOU O LOCAL DE ORIGEM FEDEX",
        "LEFT FEDEX ORIGIN FACILITY",
        "DEPARTED FEDEX LOCATION",
        "SHIPMENT LEFT FEDEX ORIGIN",
    ]
    if _contem([status_upper, ultimo_desc], GATILHOS_ORIGEM):
        return "COMING TO BRAZIL", "Saiu do local de origem — em rota internacional para o Brasil"

    TERMOS_DISPONIVEL_LIBERACAO = [
        "PACOTE DISPONIVEL PARA LIBERACAO",
        "PACOTE DISPONÍVEL PARA LIBERAÇÃO",
        "PACKAGE AVAILABLE FOR CLEARANCE",
        "AVAILABLE FOR CLEARANCE",
    ]
    TERMOS_CAMPINAS = ["CAMPINAS", "VCP", "VIRACOPOS"]
    ultimo_city_raw = eventos[0].get("scanLocation", {}).get("city", "").upper() if eventos else ""
    em_campinas = any(t in ultimo_city_raw for t in TERMOS_CAMPINAS)
    disponivel_liberacao = _contem(ultimo_textos, TERMOS_DISPONIVEL_LIBERACAO)
    sem_liberacao_confirmada = not _contem(todos_textos, GATILHOS_LIBERACAO)
    if (em_campinas or disponivel_liberacao) and disponivel_liberacao and sem_liberacao_confirmada:
        return "CUSTOMS INSPECTION", "Pacote disponível para liberação em Campinas — aguardando alfândega"

    tem_brasil    = _contem(todos_textos, TERMOS_BRASIL)
    tem_liberacao = _contem(todos_textos, GATILHOS_LIBERACAO)
    tem_atraso    = _contem(todos_textos, TERMOS_ATRASO_ALFANDEGA)
    tem_hub       = _contem(todos_textos, HUBS_INTERNACIONAIS)
    tem_origem    = _contem(todos_textos, GATILHOS_ORIGEM)

    ultimo_no_brasil = _contem(ultimo_textos, TERMOS_BRASIL)
    ultimo_em_hub    = _contem(ultimo_textos, HUBS_INTERNACIONAIS)

    if tem_brasil and tem_atraso and not tem_liberacao and ultimo_no_brasil:
        return "CUSTOMS INSPECTION", "Carga no Brasil com atraso na liberação alfandegária"

    if tem_brasil and tem_liberacao and ultimo_no_brasil:
        return "NATIONAL TRANSIT", "Liberado na alfândega — em trânsito nacional"

    if tem_liberacao and not ultimo_no_brasil:
        return "COMING TO BRAZIL", "Em trânsito internacional — ainda não chegou ao Brasil"

    if tem_brasil and not tem_liberacao and ultimo_no_brasil:
        return "CUSTOMS INSPECTION", "Carga no Brasil aguardando liberação pela Receita Federal"

    if tem_hub or tem_origem or ultimo_em_hub:
        return "COMING TO BRAZIL", "Em trânsito internacional — ainda não chegou ao Brasil"

    return "LABEL CREATED", "Etiqueta criada — aguardando coleta ou primeira movimentação"


# ==========================================================
# DIAS ÚTEIS — FERIADOS NACIONAIS BRASILEIROS
# ==========================================================

def dias_uteis_br(d1, d2) -> int:
    from datetime import date as _date, timedelta as _td

    if hasattr(d1, "date"): d1 = d1.date()
    if hasattr(d2, "date"): d2 = d2.date()
    if d1 >= d2:
        return 0

    def _pascoa(ano):
        a = ano % 19; b = ano // 100; c = ano % 100; d = b // 4; e = b % 4
        f = (b + 8) // 25; g = (b - f + 1) // 3
        h = (19 * a + b - d - g + 15) % 30; i = c // 4; k = c % 4
        l = (32 + 2 * e + 2 * i - h - k) % 7
        m = (a + 11 * h + 22 * l) // 451
        mes = (h + l - 7 * m + 114) // 31
        dia = ((h + l - 7 * m + 114) % 31) + 1
        return _date(ano, mes, dia)

    def _feriados(ano):
        pascoa = _pascoa(ano)
        td = _td
        return {
            _date(ano, 1,  1), _date(ano, 4, 21), _date(ano, 5,  1),
            _date(ano, 9,  7), _date(ano, 10, 12), _date(ano, 11,  2),
            _date(ano, 11, 15), _date(ano, 11, 20), _date(ano, 12, 25),
            pascoa - td(days=48), pascoa - td(days=47),
            pascoa - td(days=2), pascoa, pascoa + td(days=60),
        }

    _cache = {}
    def _get_feriados(ano):
        if ano not in _cache:
            _cache[ano] = _feriados(ano)
        return _cache[ano]

    count = 0
    cur = d1
    while cur < d2:
        if cur.weekday() < 5 and cur not in _get_feriados(cur.year):
            count += 1
        cur += _td(days=1)
    return count


# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO DE INTELIGÊNCIA — 5 FEATURES
# ═══════════════════════════════════════════════════════════════════════════════

PERIODOS_SOBRECARGA = [
    (1, 20, 2, 10,  1.4, "Pós-Carnaval — pico de volume represado"),
    (3, 25, 4, 10,  1.3, "Semana Santa — operação reduzida Receita Federal"),
    (6, 25, 7, 15,  1.3, "Meio de ano — fechamento fiscal"),
    (10, 25, 11, 5, 1.3, "Pré-feriados novembro"),
    (12, 10, 12, 31, 1.5, "Dezembro — maior sobrecarga do ano"),
]

def fator_sobrecarga(data_chegada_campinas) -> tuple:
    from datetime import date as _d
    if not data_chegada_campinas:
        return 1.0, ""
    try:
        if hasattr(data_chegada_campinas, "date"):
            d = data_chegada_campinas.date()
        else:
            d = datetime.strptime(str(data_chegada_campinas)[:10], "%Y-%m-%d").date()
        for (mi, di, mf, df, fator, desc) in PERIODOS_SOBRECARGA:
            inicio = _d(d.year, mi, di)
            fim = _d(d.year if mf >= mi else d.year + 1, mf, df)
            if inicio <= d <= fim:
                return fator, desc
        return 1.0, ""
    except Exception:
        return 1.0, ""


def calcular_fingerprint_produtos(arquivo_historico: str) -> dict:
    import statistics
    result = {}
    try:
        path = Path(arquivo_historico)
        if not path.exists():
            return result
        df = pd.read_excel(path)
        df.columns = df.columns.str.strip()
        if "PRODUTO" not in df.columns or "CATEGORIA" not in df.columns:
            return result
        df_del = df[df["CATEGORIA"].astype(str).str.upper().str.contains("DELIVERED", na=False)].copy()
        if df_del.empty or "LEAD_TIME_DIAS" not in df_del.columns:
            return result
        df_del["LEAD_TIME_DIAS"] = pd.to_numeric(df_del["LEAD_TIME_DIAS"], errors="coerce")
        df_del = df_del[df_del["LEAD_TIME_DIAS"].notna() & (df_del["LEAD_TIME_DIAS"] > 0)]
        df_del["PRODUTO_NORM"] = df_del["PRODUTO"].astype(str).str.upper().str.strip()
        for produto, grp in df_del.groupby("PRODUTO_NORM"):
            lts = grp["LEAD_TIME_DIAS"].tolist()
            if len(lts) < 2:
                continue
            media = round(statistics.mean(lts), 1)
            std   = round(statistics.stdev(lts), 1) if len(lts) > 2 else round(media * 0.2, 1)
            result[produto] = {
                "media": media, "std": std,
                "min": int(min(lts)), "max": int(max(lts)),
                "n": len(lts), "limiar_alerta": round(media + std, 1),
            }
    except Exception:
        pass
    return result


def avaliar_produto(produto: str, dias_atual: int, fingerprint: dict) -> dict:
    if not produto or not fingerprint:
        return {}
    prod_norm = str(produto).upper().strip()
    fp = fingerprint.get(prod_norm)
    if not fp or fp["n"] < 2:
        return {}
    media = fp["media"]; std = fp["std"]; limiar = fp["limiar_alerta"]
    if dias_atual <= media - std:
        status = "rapido"; msg = f"Abaixo da média histórica ({media}d) — dentro do esperado"
    elif dias_atual <= media:
        status = "normal"; msg = f"Dentro da média histórica ({media}d para {fp['n']} entregas)"
    elif dias_atual <= limiar:
        status = "atencao"; msg = f"Acima da média ({media}d) — monitorar"
    else:
        pct = round((dias_atual - media) / media * 100)
        status = "critico"; msg = f"{pct}% acima da média histórica ({media}d) — atraso atípico"
    pct_historico = None
    try:
        path_h = Path("historico_status.xlsx")
        if path_h.exists():
            df_h = pd.read_excel(path_h)
            df_h["PRODUTO_NORM"] = df_h.get("PRODUTO", pd.Series()).astype(str).str.upper().str.strip()
            df_del = df_h[(df_h["PRODUTO_NORM"] == prod_norm) & (df_h["CATEGORIA"].astype(str).str.upper().str.contains("DELIVERED", na=False))].copy()
            if len(df_del) >= 2 and "LEAD_TIME_DIAS" in df_del.columns:
                lts = pd.to_numeric(df_del["LEAD_TIME_DIAS"], errors="coerce").dropna().tolist()
                pct_historico = round(sum(1 for x in lts if x <= dias_atual) / len(lts) * 100)
    except Exception:
        pass
    return {"status": status, "mensagem": msg, "media_produto": media, "std_produto": std,
            "n_historico": fp["n"], "pct_historico": pct_historico, "limiar_alerta": limiar}


def calcular_risco_atraso(produto: str, regiao: str, dias_em_campinas: int, arquivo_historico: str) -> dict:
    result = {"probabilidade": None, "casos_similares": 0, "alerta": False}
    try:
        path = Path(arquivo_historico)
        if not path.exists() or dias_em_campinas is None:
            return result
        df = pd.read_excel(path)
        df.columns = df.columns.str.strip()
        if not all(c in df.columns for c in ["CATEGORIA", "DATA_CHEGADA_MEMPHIS"]):
            return result
        mask = df["CATEGORIA"].astype(str).str.upper().str.contains("CUSTOMS|DELIVERED|NATIONAL", na=False)
        df_sim = df[mask].copy()
        if "PRODUTO" in df.columns and produto:
            prod_norm = str(produto).upper().strip()
            df_prod = df_sim[df_sim["PRODUTO"].astype(str).str.upper().str.strip() == prod_norm]
            if len(df_prod) >= 5:
                df_sim = df_prod
        if len(df_sim) < 3:
            return result
        if "LEAD_TIME_DIAS" in df_sim.columns:
            lts = pd.to_numeric(df_sim["LEAD_TIME_DIAS"], errors="coerce").dropna()
            if len(lts) >= 3:
                casos_que_demoraram = sum(1 for x in lts if x > dias_em_campinas * 1.5)
                prob = round(casos_que_demoraram / len(lts) * 100)
                media_res = round(float(lts.mean()), 1)
                result = {
                    "probabilidade": prob, "casos_similares": len(lts),
                    "media_resolucao": media_res,
                    "alerta": prob >= 40 and dias_em_campinas >= 3,
                    "mensagem": f"Com base em {len(lts)} casos similares, {prob}% demoraram mais que o atual. Lead time médio histórico: {media_res}d úteis."
                }
    except Exception:
        pass
    return result


def prever_liberacao(produto: str, regiao: str, dias_em_customs: int, arquivo_historico: str, fator_cal: float = 1.0) -> dict:
    result = {"dias_restantes": None, "confianca": None, "base": 0}
    try:
        path = Path(arquivo_historico)
        if not path.exists():
            return result
        df = pd.read_excel(path)
        df.columns = df.columns.str.strip()
        mask_del = df["CATEGORIA"].astype(str).str.upper().str.contains("DELIVERED", na=False)
        df_del = df[mask_del].copy()
        if "LEAD_TIME_DIAS" not in df_del.columns or len(df_del) < 5:
            return result
        df_del["LEAD_TIME_DIAS"] = pd.to_numeric(df_del["LEAD_TIME_DIAS"], errors="coerce")
        df_del = df_del[df_del["LEAD_TIME_DIAS"].notna()]
        if "PRODUTO" in df_del.columns and produto:
            prod_norm = str(produto).upper().strip()
            df_prod = df_del[df_del["PRODUTO"].astype(str).str.upper().str.strip() == prod_norm]
            if len(df_prod) >= 5:
                df_del = df_prod
        if "REGIAO" in df_del.columns and regiao and len(df_del) > 10:
            df_reg = df_del[df_del["REGIAO"].astype(str).str.strip() == regiao]
            if len(df_reg) >= 5:
                df_del = df_reg
        lts = df_del["LEAD_TIME_DIAS"].tolist()
        if len(lts) < 3:
            return result
        import statistics
        media_total = statistics.mean(lts)
        dias_restantes = max(0, round((media_total - dias_em_customs) * fator_cal))
        if len(lts) > 2:
            std = statistics.stdev(lts)
            cv = std / media_total if media_total > 0 else 1
            confianca = max(30, min(90, round((1 - cv) * 60 + min(len(lts), 30))))
        else:
            confianca = 40
        ajuste_desc = f"(+{round((fator_cal-1)*100)}% período de sobrecarga)" if fator_cal > 1.05 else ""
        result = {"dias_restantes": dias_restantes, "media_historica": round(media_total, 1),
                  "confianca": confianca, "base": len(lts), "ajuste_calendario": ajuste_desc}
    except Exception:
        pass
    return result


def gerar_relatorio_executivo(arquivo_historico: str, arquivo_html_saida: str) -> bool:
    logger = logging.getLogger("fedex_tracker")
    try:
        from datetime import date as _d, timedelta as _td
        hoje = _d.today()
        if hoje.weekday() != 0:
            return False
        path = Path(arquivo_historico)
        if not path.exists():
            return False
        df = pd.read_excel(path)
        df.columns = df.columns.str.strip()
        df["_dc"] = pd.to_datetime(df.get("DATA_CONSULTA", pd.Series()), dayfirst=True, errors="coerce")
        semana_inicio = hoje - _td(days=7)
        df_sem = df[df["_dc"].dt.date >= semana_inicio].copy() if "_dc" in df.columns else df.copy()
        total_sem    = len(df_sem["AWB"].unique()) if "AWB" in df_sem.columns else 0
        entregas_sem = len(df_sem[df_sem["CATEGORIA"].astype(str).str.upper().str.contains("DELIVERED", na=False)]["AWB"].unique()) if "CATEGORIA" in df_sem.columns else 0
        atrasos_sem  = len(df_sem[df_sem["CATEGORIA"].astype(str).str.contains("⚠", na=False)]["AWB"].unique()) if "CATEGORIA" in df_sem.columns else 0
        lt_media = "—"
        if "LEAD_TIME_DIAS" in df_sem.columns:
            lts = pd.to_numeric(df_sem["LEAD_TIME_DIAS"], errors="coerce").dropna()
            if len(lts) > 0:
                lt_media = f"{round(float(lts.mean()), 1)}d úteis"
        top_atraso = []
        if "PRODUTO" in df_sem.columns and "CATEGORIA" in df_sem.columns:
            df_at = df_sem[df_sem["CATEGORIA"].astype(str).str.contains("⚠|CUSTOMS", na=False)]
            if not df_at.empty:
                top_atraso = df_at["PRODUTO"].value_counts().head(5).reset_index().values.tolist()
        dist_cat = {}
        if "CATEGORIA" in df_sem.columns:
            for cat in df_sem["CATEGORIA"].astype(str):
                import re as _re
                cat_norm = _re.sub(r"^⚠\s*","",cat).strip()
                cat_norm = _re.sub(r"\s*\(\d+d\)\s*$","",cat_norm).strip()
                cat_norm = _re.sub(r"\s*MEMPHIS.*$","",cat_norm).strip()
                dist_cat[cat_norm] = dist_cat.get(cat_norm, 0) + 1
        dist_reg = {}
        if "REGIAO" in df_sem.columns:
            for reg in df_sem["REGIAO"].dropna().astype(str):
                if reg and reg != "nan":
                    dist_reg[reg] = dist_reg.get(reg, 0) + 1
        CAT_COLORS_REL = {"COMING TO BRAZIL":"#4f7dff","CUSTOMS INSPECTION":"#f59e0b","NATIONAL TRANSIT":"#10b981","OUT FOR DELIVERY":"#818cf8","DELIVERED":"#22c55e","LABEL CREATED":"#475569"}
        def barras_html(dados, color_map, default_color="#4f7dff"):
            if not dados: return "<p style='color:#888;font-size:12px'>Sem dados</p>"
            total = sum(dados.values()) or 1
            h = ""
            for k, v in sorted(dados.items(), key=lambda x: -x[1]):
                pct = round(v / total * 100); cor = color_map.get(k, default_color)
                h += f'<div style="margin-bottom:10px"><div style="display:flex;justify-content:space-between;font-size:11px;margin-bottom:3px"><span>{k}</span><span style="color:#666">{v} ({pct}%)</span></div><div style="background:#eee;border-radius:3px;height:10px"><div style="background:{cor};width:{pct}%;height:100%;border-radius:3px"></div></div></div>'
            return h
        REG_COLORS_REL = {"América do Norte":"#4f7dff","Europa":"#10b981","Oriente Médio":"#f59e0b","Ásia":"#a78bfa","Outros":"#475569"}
        top_atraso_html = "".join(f"<tr><td style='padding:6px 8px'>{r[0]}</td><td style='padding:6px 8px;text-align:center;color:#f59e0b'>{r[1]}</td></tr>" for r in top_atraso) or "<tr><td colspan='2' style='padding:12px;color:#888;text-align:center'>Nenhum produto em atraso</td></tr>"
        html = f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
<title>Relatório Executivo — Semana {semana_inicio.strftime('%d/%m')} a {hoje.strftime('%d/%m/%Y')}</title>
<style>@media print{{body{{margin:0}}.no-print{{display:none}}.page{{page-break-after:always}}}}*{{box-sizing:border-box;margin:0;padding:0;font-family:Arial,sans-serif}}body{{background:#f5f5f5;color:#222}}.page{{background:#fff;max-width:900px;margin:20px auto;padding:40px;border-radius:8px;box-shadow:0 2px 12px #0001}}.header{{border-bottom:3px solid #4f7dff;padding-bottom:16px;margin-bottom:24px;display:flex;justify-content:space-between;align-items:flex-end}}.header-title{{font-size:22px;font-weight:700;color:#1a1a2e}}.header-sub{{font-size:12px;color:#888;margin-top:4px}}.header-date{{font-size:11px;color:#888;text-align:right}}.kpis{{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:28px}}.kpi{{background:#f8f9fc;border-radius:8px;padding:16px;border-left:4px solid #4f7dff}}.kpi-label{{font-size:10px;text-transform:uppercase;letter-spacing:.1em;color:#888;margin-bottom:6px}}.kpi-val{{font-size:26px;font-weight:700;color:#1a1a2e}}.kpi-sub{{font-size:11px;color:#888;margin-top:2px}}.section{{margin-bottom:28px}}.section-title{{font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:#4f7dff;border-bottom:1px solid #eee;padding-bottom:6px;margin-bottom:14px}}.grid-2{{display:grid;grid-template-columns:1fr 1fr;gap:24px}}table{{width:100%;border-collapse:collapse;font-size:12px}}th{{background:#f0f2ff;padding:8px;text-align:left;font-size:10px;text-transform:uppercase;letter-spacing:.08em;color:#4f7dff}}td{{padding:6px 8px;border-bottom:1px solid #f0f0f0}}.footer{{text-align:center;font-size:10px;color:#bbb;margin-top:24px;padding-top:16px;border-top:1px solid #eee}}.print-btn{{display:block;margin:16px auto;padding:10px 24px;background:#4f7dff;color:#fff;border:none;border-radius:6px;font-size:13px;cursor:pointer;font-family:Arial}}.print-btn:hover{{background:#3d6be8}}</style></head><body>
<div class="no-print" style="text-align:center;padding:12px;background:#4f7dff"><button class="print-btn" onclick="window.print()">🖨️ Imprimir / Salvar como PDF</button></div>
<div class="page">
  <div class="header"><div><div class="header-title">📦 Relatório Executivo — FedEx Tracker</div><div class="header-sub">Semana de {semana_inicio.strftime('%d/%m/%Y')} a {hoje.strftime('%d/%m/%Y')}</div></div><div class="header-date">Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}</div></div>
  <div class="section"><div class="section-title">Resumo da semana</div><div class="kpis"><div class="kpi"><div class="kpi-label">Total monitorado</div><div class="kpi-val">{total_sem}</div><div class="kpi-sub">AWBs na semana</div></div><div class="kpi" style="border-left-color:#22c55e"><div class="kpi-label">Entregas</div><div class="kpi-val" style="color:#16a34a">{entregas_sem}</div><div class="kpi-sub">concluídas</div></div><div class="kpi" style="border-left-color:#f59e0b"><div class="kpi-label">Em atraso</div><div class="kpi-val" style="color:#d97706">{atrasos_sem}</div><div class="kpi-sub">alertas ativos</div></div><div class="kpi" style="border-left-color:#818cf8"><div class="kpi-label">Lead time médio</div><div class="kpi-val" style="font-size:18px">{lt_media}</div><div class="kpi-sub">entregas concluídas</div></div></div></div>
  <div class="section grid-2"><div><div class="section-title">Distribuição por status</div>{barras_html(dist_cat,CAT_COLORS_REL)}</div><div><div class="section-title">Distribuição por região de origem</div>{barras_html(dist_reg,REG_COLORS_REL)}</div></div>
  <div class="section"><div class="section-title">Produtos com mais ocorrências em atraso</div><table><thead><tr><th>Produto</th><th>Ocorrências</th></tr></thead><tbody>{top_atraso_html}</tbody></table></div>
  <div class="footer">FedEx AWB Tracker — Relatório gerado automaticamente toda segunda-feira</div>
</div></body></html>"""
        with open(arquivo_html_saida, "w", encoding="utf-8") as f:
            f.write(html)
        logger.info(f"Relatório executivo semanal gerado → '{arquivo_html_saida}'")
        return True
    except Exception as e:
        logger.warning(f"Erro ao gerar relatório executivo: {e}")
        return False


# ==========================================================
# AUTENTICAÇÃO — TOKEN OAUTH
# ==========================================================

class TokenManager:
    def __init__(self, config: Config):
        self.config = config; self._token = None; self._expira_em = 0
        self.logger = logging.getLogger("fedex_tracker")

    def get_token(self) -> str:
        if self._token and time.time() < self._expira_em - 60:
            return self._token
        self.logger.info("Obtendo token de acesso da API FedEx...")
        response = requests.post(
            self.config.url_token,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            data={"grant_type": "client_credentials", "client_id": self.config.client_id, "client_secret": self.config.client_secret},
            timeout=self.config.timeout,
        )
        if response.status_code != 200:
            raise Exception(f"Erro ao obter token: HTTP {response.status_code} — {response.text[:200]}")
        dados = response.json()
        self._token = dados["access_token"]
        self._expira_em = time.time() + int(dados.get("expires_in", 3600))
        self.logger.info("Token obtido! (válido por 1 hora)")
        return self._token


# ═══════════════════════════════════════════════════════════════════════════════
# MAPEAMENTO DE REGIÕES
# ═══════════════════════════════════════════════════════════════════════════════

PAISES_REGIAO = {
    "US": "América do Norte", "CA": "América do Norte", "MX": "América do Norte", "USA": "América do Norte",
    "DE": "Europa", "FR": "Europa", "GB": "Europa", "UK": "Europa", "IT": "Europa", "ES": "Europa",
    "NL": "Europa", "BE": "Europa", "CH": "Europa", "AT": "Europa", "SE": "Europa", "NO": "Europa",
    "DK": "Europa", "FI": "Europa", "PL": "Europa", "PT": "Europa", "IE": "Europa", "CZ": "Europa",
    "HU": "Europa", "RO": "Europa",
    "AE": "Oriente Médio", "IL": "Oriente Médio", "TR": "Oriente Médio", "SA": "Oriente Médio",
    "QA": "Oriente Médio", "KW": "Oriente Médio", "BH": "Oriente Médio", "OM": "Oriente Médio",
    "JO": "Oriente Médio", "LB": "Oriente Médio", "EG": "Oriente Médio", "IR": "Oriente Médio",
    "CN": "Ásia", "IN": "Ásia", "JP": "Ásia", "KR": "Ásia", "SG": "Ásia", "TW": "Ásia",
    "TH": "Ásia", "MY": "Ásia", "HK": "Ásia", "PH": "Ásia", "ID": "Ásia", "VN": "Ásia",
}

def pais_para_regiao(country_code: str) -> str:
    return PAISES_REGIAO.get(str(country_code).upper(), "Outros")


def extrair_dados_relatorio(eventos: list) -> dict:
    if not eventos:
        return {"pais_origem": "", "regiao": "", "data_criacao": "", "data_entrega": "", "lead_time_dias": None}
    eventos_ord = list(reversed(eventos))
    pais_origem = ""; data_criacao = ""
    for ev in eventos_ord:
        loc = ev.get("scanLocation", {}); pais = loc.get("countryCode", "")
        if pais and pais.upper() != "BR":
            pais_origem = pais.upper()
            ts = ev.get("date", "") or ev.get("timestamp", "")
            if ts: data_criacao = str(ts)[:10]
            break
    data_entrega = ""
    for ev in eventos:
        desc = ev.get("eventDescription", "").upper()
        if any(t in desc for t in ["DELIVERED", "ENTREGUE", "DELIVERY MADE"]):
            ts = ev.get("date", "") or ev.get("timestamp", "")
            if ts: data_entrega = str(ts)[:10]
            break
    lead_time = None
    if data_criacao and data_entrega:
        try:
            from datetime import datetime as _dt
            d1 = _dt.strptime(data_criacao[:10], "%Y-%m-%d").date()
            d2 = _dt.strptime(data_entrega[:10], "%Y-%m-%d").date()
            lead_time = dias_uteis_br(d1, d2)
        except: pass
    return {"pais_origem": pais_origem, "regiao": pais_para_regiao(pais_origem),
            "data_criacao": data_criacao, "data_entrega": data_entrega, "lead_time_dias": lead_time}


# ==========================================================
# CONSULTA À API OFICIAL
# ==========================================================

def consultar_awb(awb_original: str, awb_limpo: str, token: str, config: Config, pedido: str = "", produto: str = "") -> ResultadoAWB:
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M")
    logger    = logging.getLogger("fedex_tracker")
    try:
        time.sleep(config.delay_entre)
        response = requests.post(
            config.url_tracking,
            headers={"Content-Type": "application/json", "Authorization": f"Bearer {token}", "X-locale": "pt_BR"},
            json={"includeDetailedScans": True, "trackingInfo": [{"trackingNumberInfo": {"trackingNumber": awb_limpo}}]},
            timeout=config.timeout,
        )
        if response.status_code == 401: raise Exception("Token inválido ou expirado")
        if response.status_code == 429:
            logger.warning(f"[{_mascarar_awb(awb_limpo)}] Rate limit — aguardando 30s...")
            time.sleep(30); raise Exception("Rate limit atingido")
        if response.status_code != 200:
            raise Exception(f"HTTP {response.status_code}: {response.text[:200]}")
        dados  = response.json()
        output = dados.get("output", {}).get("completeTrackResults", [{}])[0].get("trackResults", [{}])[0]
        erro = output.get("error", {})
        if erro:
            codigo = erro.get("code", ""); msg = erro.get("message", "")
            if codigo == "TRACKING.TRACKINGNUMBER.NOTFOUND":
                return ResultadoAWB(AWB=awb_limpo, AWB_ORIGINAL=awb_original, PEDIDO=pedido, PRODUTO=produto,
                    CATEGORIA="AWB NAO ENCONTRADO", STATUS_FEDEX="NOT FOUND",
                    ULTIMO_EVENTO="Número não encontrado na base FedEx",
                    MOTIVO_CATEGORIA="AWB inválido ou não registrado", DATA_CONSULTA=timestamp, SUCESSO=False)
            raise Exception(f"{codigo}: {msg}")
        status_fedex = output.get("latestStatusDetail", {}).get("description", "UNKNOWN").upper().strip()
        eventos = output.get("scanEvents", [])
        if eventos:
            ult = eventos[0]
            descr = ult.get("eventDescription", "").strip()
            city  = ult.get("scanLocation", {}).get("city", "").strip()
            pais  = ult.get("scanLocation", {}).get("countryCode", "").strip()
            data_e = ult.get("date", "").split("T")[0]
            hora_e = ult.get("date", "T").split("T")[1][:5] if "T" in ult.get("date", "") else ""
            ultimo_evento = f"{descr} — {city}/{pais} {data_e} {hora_e}".upper().strip(" —")
        else:
            ultimo_evento = "SEM EVENTOS REGISTRADOS"
        data_chegada_memphis = ""
        for ev in reversed(eventos):
            if "MEMPHIS" in ev.get("scanLocation", {}).get("city", "").upper():
                data_chegada_memphis = ev.get("date", "").split("T")[0]; break
        categoria, motivo = classificar(eventos, status_fedex)

        def _ev_para_cat(ev):
            desc = ev.get("eventDescription","").upper()
            city = ev.get("scanLocation",{}).get("city","").upper()
            country = ev.get("scanLocation",{}).get("countryCode","").upper()
            if any(t in desc for t in ["DELIVERED","ENTREGUE","DELIVERY MADE"]): return "DELIVERED"
            if any(t in desc for t in ["ON FEDEX VEHICLE","OUT FOR DELIVERY","SAIU PARA ENTREGA","EM UM VEICULO","EM VEÍCULO"]): return "OUT FOR DELIVERY"
            if any(t in desc for t in ["LIBERACAO DA REMESSA","LIBERAÇÃO DA REMESSA","INTERNATIONAL SHIPMENT RELEASE","RELEASED BY CUSTOMS","CUSTOMS CLEARED","CLEARANCE COMPLETED"]): return "NATIONAL TRANSIT"
            if any(t in desc for t in ["ATRASO NA LIBERACAO","ATRASO NA LIBERAÇÃO","CLEARANCE DELAY","CUSTOMS DELAY","HELD BY CUSTOMS","RETIDO NA ALFANDEGA","RETIDO NA ALFÂNDEGA","PACOTE DISPONIVEL","PACOTE DISPONÍVEL"]): return "CUSTOMS INSPECTION"
            if country == "BR" or city in ("CAMPINAS","VIRACOPOS","BARUERI"): return "CUSTOMS INSPECTION"
            if any(t in desc for t in ["LEFT FEDEX","DEPARTED FEDEX","SAIU DA","SHIPMENT LEFT","LEFT ORIGIN","PICKED UP","COLETADO"]): return "COMING TO BRAZIL"
            if any(t in desc for t in ["ARRIVED","CHEGOU","IN TRANSIT","EM TRANSITO","A CAMINHO","IN FEDEX POSSESSION"]): return "COMING TO BRAZIL"
            if any(t in desc for t in ["LABEL","ETIQUETA","SHIPMENT INFORMATION"]): return "LABEL CREATED"
            return None

        timeline_json = []; _tl_prev = None
        for ev in reversed(eventos):
            _ev_cat = _ev_para_cat(ev)
            if _ev_cat and _ev_cat != _tl_prev:
                timeline_json.append({"cat": _ev_cat, "data": ev.get("date","").split("T")[0]})
                _tl_prev = _ev_cat
        if timeline_json and timeline_json[-1]["cat"] != categoria:
            timeline_json.append({"cat": categoria, "data": datetime.now().strftime("%Y-%m-%d")})

        eventos_json = []
        for ev in reversed(eventos):
            _dt_raw = ev.get("date","")
            _loc = ev.get("scanLocation", {})
            _local_str = ", ".join(filter(None, [_loc.get("city","").strip(), _loc.get("stateOrProvinceCode","").strip(), _loc.get("countryCode","").strip()]))
            _subdesc = ev.get("eventSubDescription","").strip() if ev.get("eventSubDescription") else ""
            eventos_json.append({"data": _dt_raw.split("T")[0] if "T" in _dt_raw else _dt_raw[:10],
                "hora": _dt_raw.split("T")[1][:5] if "T" in _dt_raw else "",
                "desc": ev.get("eventDescription","").strip(), "subdesc": _subdesc, "local": _local_str})

        logger.debug(f"[{_mascarar_awb(awb_limpo)}] {categoria} | {status_fedex} | {ultimo_evento[:50]}")
        dados_rel = extrair_dados_relatorio(eventos)
        dest = output.get("recipientInformation", {}); dest_addr = dest.get("address", {})
        endereco_entrega = ", ".join(filter(None, [dest_addr.get("city",""), dest_addr.get("stateOrProvinceCode",""), dest_addr.get("countryCode","")]))

        return ResultadoAWB(
            AWB=awb_limpo, AWB_ORIGINAL=awb_original, CATEGORIA=categoria,
            STATUS_FEDEX=status_fedex, ULTIMO_EVENTO=ultimo_evento, MOTIVO_CATEGORIA=motivo,
            DATA_CONSULTA=timestamp, SUCESSO=True, DATA_CHEGADA_MEMPHIS=data_chegada_memphis,
            PEDIDO=pedido, PRODUTO=produto,
            PAIS_ORIGEM=dados_rel["pais_origem"], REGIAO=dados_rel["regiao"],
            DATA_CRIACAO=dados_rel["data_criacao"], DATA_ENTREGA=dados_rel["data_entrega"],
            LEAD_TIME_DIAS=dados_rel["lead_time_dias"],
            TIMELINE_JSON=json.dumps(timeline_json, ensure_ascii=False),
            EVENTOS_JSON=json.dumps(eventos_json, ensure_ascii=False),
            ENDERECO_ENTREGA=endereco_entrega,
        )
    except Exception as e:
        msg = str(e).split('\n')[0]
        logger.error(f"[{_mascarar_awb(awb_limpo)}] {msg}")
        return ResultadoAWB(AWB=awb_limpo, AWB_ORIGINAL=awb_original, CATEGORIA="ERRO",
            STATUS_FEDEX="ERRO", ULTIMO_EVENTO=msg[:200], MOTIVO_CATEGORIA="Falha na consulta",
            DATA_CONSULTA=timestamp, SUCESSO=False, PEDIDO=pedido, PRODUTO=produto)


# ==========================================================
# STORAGE
# ==========================================================

def carregar_awbs(arquivo: str) -> List[tuple]:
    path = Path(arquivo)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {arquivo}")
    df = pd.read_excel(path)
    df.columns = df.columns.str.strip()
    if "AWB" not in df.columns:
        raise ValueError(f"Coluna 'AWB' não encontrada. Colunas: {list(df.columns)}")
    col_pedido  = "Pedido" if "Pedido" in df.columns else ("PEDIDO" if "PEDIDO" in df.columns else None)
    col_produto = "PRODUTO" if "PRODUTO" in df.columns else ("Produto" if "Produto" in df.columns else None)
    logger = logging.getLogger("fedex_tracker")
    resultado = []
    for _, row in df.iterrows():
        awb_original = str(row["AWB"]) if pd.notna(row["AWB"]) else ""
        if not awb_original or awb_original == "nan": continue
        pedido = str(row[col_pedido]) if col_pedido and pd.notna(row[col_pedido]) else ""
        if pedido.endswith(".0"): pedido = pedido[:-2]
        produto = str(row[col_produto]).strip() if col_produto and pd.notna(row[col_produto]) else ""
        awb_limpo = limpar_awb(awb_original)
        if validar_awb(awb_limpo):
            resultado.append((awb_original.strip(), awb_limpo, pedido, produto))
        else:
            logger.warning(f"AWB ignorado: '{_mascarar_awb(awb_original)}'  (formato inválido)")
    logger.info(f"{len(resultado)} AWBs válidos")
    return resultado


def _fazer_backup(arquivo: str) -> None:
    path = Path(arquivo)
    if path.exists():
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        shutil.copy(path, path.with_name(f"{path.stem}_bak_{ts}{path.suffix}"))


def salvar_resultados(resultados, arquivo_historico, arquivo_ultimo) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    import re as _re

    logger = logging.getLogger("fedex_tracker")
    df = pd.DataFrame(resultados)
    df["AWB"] = df["AWB"].astype(str).str.strip()

    if "PEDIDO" not in df.columns or df["PEDIDO"].isna().all() or (df["PEDIDO"] == "").all():
        try:
            awbs_ref = pd.read_excel(str(Path(arquivo_historico).parent / "awbs.xlsx"))
            awbs_ref.columns = awbs_ref.columns.str.strip()
            col_ped = "Pedido" if "Pedido" in awbs_ref.columns else ("PEDIDO" if "PEDIDO" in awbs_ref.columns else None)
            if col_ped:
                awbs_ref["AWB_CLEAN"] = awbs_ref["AWB"].astype(str).str.replace(r"\s+", "", regex=True)
                awbs_ref[col_ped] = awbs_ref[col_ped].astype(str).str.replace(r"\.0$", "", regex=True)
                pedido_map = dict(zip(awbs_ref["AWB_CLEAN"], awbs_ref[col_ped]))
                df["PEDIDO"] = df["AWB"].astype(str).map(pedido_map).fillna("")
        except Exception as e:
            logger.warning(f"Não foi possível carregar pedidos do awbs.xlsx: {e}")
            df["PEDIDO"] = df.get("PEDIDO", "")
    else:
        df["PEDIDO"] = df["PEDIDO"].fillna("").astype(str).str.replace(r"\.0$", "", regex=True)

    ORDEM_CAT = {"LABEL CREATED":1,"COMING TO BRAZIL":2,"CUSTOMS INSPECTION":3,"NATIONAL TRANSIT":4,"OUT FOR DELIVERY":5,"DELIVERED":6,"AWB NAO ENCONTRADO":7,"ERRO":8}
    CORES = {
        "LABEL CREATED":       {"bg":"D9D9D9","font":"595959"},
        "COMING TO BRAZIL":    {"bg":"BDD7EE","font":"1F4E79"},
        "CUSTOMS INSPECTION":  {"bg":"FFE699","font":"7F6000"},
        "NATIONAL TRANSIT":    {"bg":"C6EFCE","font":"276221"},
        "OUT FOR DELIVERY":    {"bg":"9DC3E6","font":"1F4E79"},
        "DELIVERED":           {"bg":"375623","font":"FFFFFF"},
        "AWB NAO ENCONTRADO":  {"bg":"F4CCCC","font":"990000"},
        "ERRO":                {"bg":"FF0000","font":"FFFFFF"},
    }
    DIAS_ATRASO_CUSTOMS = 5; DIAS_ATRASO_MEMPHIS = 3

    df["_ordem"] = df["CATEGORIA"].map(lambda x: ORDEM_CAT.get(x, 9))
    df = df.sort_values("_ordem").drop(columns=["_ordem"]).reset_index(drop=True)

    def dias_no_status(ultimo_evento_str):
        try:
            partes = str(ultimo_evento_str).split("—")
            data_str = partes[-1].strip().split(" ")
            data_part = [p for p in data_str if _re.match(r"\d{4}-\d{2}-\d{2}", p)]
            if data_part:
                from datetime import date
                d = datetime.strptime(data_part[0], "%Y-%m-%d").date()
                return dias_uteis_br(d, date.today())
        except: pass
        return None

    df["_dias"] = df["ULTIMO_EVENTO"].apply(dias_no_status)

    wb = Workbook(); wb.remove(wb.active)

    def make_fill(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)
    def make_border():
        side = Side(style="thin", color="BFBFBF")
        return Border(left=side, right=side, top=side, bottom=side)
    def header_style(cell, bg="1F4E79", fg="FFFFFF"):
        cell.font = Font(bold=True, color=fg, name="Arial", size=10)
        cell.fill = make_fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border()
    def data_style(cell, bg="FFFFFF", fg="000000", bold=False, wrap=False):
        cell.font = Font(color=fg, name="Arial", size=9, bold=bold)
        cell.fill = make_fill(bg)
        cell.alignment = Alignment(vertical="center", wrap_text=wrap)
        cell.border = make_border()

    COLUNAS = ["AWB","PEDIDO","CATEGORIA","STATUS_FEDEX","ULTIMO_EVENTO","MOTIVO_CATEGORIA","DATA_CONSULTA","DATA_CHEGADA_MEMPHIS","REGIAO","TIMELINE_JSON","EVENTOS_JSON","ENDERECO_ENTREGA"]
    LARGURAS = [22,14,22,28,52,40,18,14,18,10,10,20]
    HEADERS  = ["AWB","PEDIDO","CATEGORIA","STATUS FEDEX","ÚLTIMO EVENTO","MOTIVO","DATA CONSULTA","CHEGADA MEMPHIS","REGIÃO","TIMELINE","EVENTOS","DEST"]

    def escrever_aba(ws, df_aba, titulo_aba):
        ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 32
        for col_idx, (header, largura) in enumerate(zip(HEADERS, LARGURAS), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            header_style(cell)
            ws.column_dimensions[get_column_letter(col_idx)].width = largura
        for row_idx, (_, row) in enumerate(df_aba.iterrows(), 2):
            cat = row["CATEGORIA"]; cores = CORES.get(cat, {"bg":"FFFFFF","font":"000000"})
            dias = row.get("_dias")
            em_customs = (cat == "CUSTOMS INSPECTION" and dias and dias >= DIAS_ATRASO_CUSTOMS)
            data_memphis_str = str(row.get("DATA_CHEGADA_MEMPHIS",""))
            dias_memphis = None
            if data_memphis_str and data_memphis_str not in ("nan",""):
                try:
                    from datetime import date as _date
                    d_mem = datetime.strptime(data_memphis_str[:10], "%Y-%m-%d").date()
                    dias_memphis = dias_uteis_br(d_mem, _date.today())
                except: pass
            em_memphis = (cat == "COMING TO BRAZIL" and dias_memphis is not None and dias_memphis >= DIAS_ATRASO_MEMPHIS)
            atraso = em_customs or em_memphis
            for col_idx, col_name in enumerate(COLUNAS, 1):
                val  = row.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if pd.notna(val) else "")
                if atraso: data_style(cell, bg="FF4C4C", fg="FFFFFF", bold=True, wrap=(col_idx in [4,5]))
                else: data_style(cell, bg=cores["bg"], fg=cores["font"], wrap=(col_idx in [4,5]))
                ws.row_dimensions[row_idx].height = 28 if col_idx == 1 else ws.row_dimensions[row_idx].height
            if atraso:
                cell_cat = ws.cell(row=row_idx, column=3)
                cell_cat.value = f"⚠ MEMPHIS +{dias_memphis}d" if em_memphis else f"⚠ {cat} ({dias}d)"
                cell_cat.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
                cell_cat.fill = make_fill("C00000")

    # ── ABA RESUMO ──────────────────────────────────────────────────────────
    ws_resumo = wb.create_sheet("📊 RESUMO")
    ws_resumo.sheet_properties.tabColor = "1F4E79"
    ws_resumo.merge_cells("A1:F1")
    titulo = ws_resumo["A1"]
    titulo.value = f"FEDEX TRACKING — RELATÓRIO {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    titulo.font = Font(bold=True, color="FFFFFF", name="Arial", size=14)
    titulo.fill = make_fill("1F4E79")
    titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws_resumo.row_dimensions[1].height = 36

    total = len(df); sucesso = int(df["SUCESSO"].sum()) if "SUCESSO" in df.columns else total
    def _dias_memphis(row):
        s = str(row.get("DATA_CHEGADA_MEMPHIS",""))
        if s and s != "nan":
            try:
                from datetime import date as _d
                return dias_uteis_br(datetime.strptime(s[:10],"%Y-%m-%d").date(), _d.today())
            except: pass
        return 0
    df["_dias_memphis"] = df.apply(_dias_memphis, axis=1)
    em_atraso = int(((df["CATEGORIA"]=="CUSTOMS INSPECTION")&(df["_dias"]>=DIAS_ATRASO_CUSTOMS)).sum()+((df["CATEGORIA"]=="COMING TO BRAZIL")&(df["_dias_memphis"]>=DIAS_ATRASO_MEMPHIS)).sum())

    for col_i, (label, val, cor) in enumerate([("TOTAL AWBs",total,"2E75B6"),("CONSULTADOS OK",sucesso,"375623"),("⚠ EM ATRASO (+7d)",em_atraso,"C00000")], 1):
        cl = get_column_letter(col_i*2-1); cl2 = get_column_letter(col_i*2)
        ws_resumo.merge_cells(f"{cl}3:{cl2}3"); ws_resumo.merge_cells(f"{cl}4:{cl2}4")
        c_label = ws_resumo[f"{cl}3"]; c_label.value = label
        c_label.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        c_label.fill = make_fill(cor); c_label.alignment = Alignment(horizontal="center", vertical="center")
        c_val = ws_resumo[f"{cl}4"]; c_val.value = val
        c_val.font = Font(bold=True, color=cor, name="Arial", size=24)
        c_val.fill = make_fill("F2F2F2"); c_val.alignment = Alignment(horizontal="center", vertical="center")
        ws_resumo.row_dimensions[4].height = 48

    ws_resumo["A6"].value="CATEGORIA"; ws_resumo["B6"].value="QUANTIDADE"; ws_resumo["C6"].value="%"
    for cell, bg in [(ws_resumo["A6"],"1F4E79"),(ws_resumo["B6"],"1F4E79"),(ws_resumo["C6"],"1F4E79")]:
        header_style(cell, bg=bg)
    contagem = df["CATEGORIA"].value_counts().to_dict()
    cats_ordem = [c for c in ORDEM_CAT.keys() if c in contagem]
    row_t = 7
    for cat in cats_ordem:
        qtd = contagem[cat]; cores = CORES.get(cat, {"bg":"FFFFFF","font":"000000"})
        c_cat = ws_resumo.cell(row=row_t, column=1, value=cat)
        c_qtd = ws_resumo.cell(row=row_t, column=2, value=qtd)
        c_pct = ws_resumo.cell(row=row_t, column=3, value=f"=B{row_t}/B${7+len(cats_ordem)}")
        for c, fmt in [(c_cat,None),(c_qtd,None),(c_pct,"0.0%")]:
            c.font = Font(color=cores["font"], name="Arial", size=10, bold=True)
            c.fill = make_fill(cores["bg"]); c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = make_border()
            if fmt: c.number_format = fmt
        row_t += 1
    for c in [ws_resumo.cell(row=row_t,column=1,value="TOTAL"),ws_resumo.cell(row=row_t,column=2,value=f"=SUM(B7:B{row_t-1})"),ws_resumo.cell(row=row_t,column=3,value="100%")]:
        c.font=Font(bold=True,color="FFFFFF",name="Arial",size=10); c.fill=make_fill("1F4E79"); c.alignment=Alignment(horizontal="center",vertical="center"); c.border=make_border()
    ws_resumo.column_dimensions["A"].width=26; ws_resumo.column_dimensions["B"].width=14; ws_resumo.column_dimensions["C"].width=10
    chart = BarChart(); chart.type="bar"; chart.title="AWBs por Categoria"; chart.style=10; chart.width=18; chart.height=12; chart.legend=None
    chart.add_data(Reference(ws_resumo,min_col=2,min_row=6,max_row=row_t-1),titles_from_data=True)
    chart.set_categories(Reference(ws_resumo,min_col=1,min_row=7,max_row=row_t-1))
    chart.series[0].graphicalProperties.solidFill="2E75B6"; ws_resumo.add_chart(chart,"E3")

    ws_geral = wb.create_sheet("📋 TODOS"); ws_geral.sheet_properties.tabColor="2E75B6"
    escrever_aba(ws_geral, df, "TODOS")

    ICONES = {"LABEL CREATED":"🏷","COMING TO BRAZIL":"✈","CUSTOMS INSPECTION":"🔍","NATIONAL TRANSIT":"🚚","OUT FOR DELIVERY":"📦","DELIVERED":"✅","AWB NAO ENCONTRADO":"❓","ERRO":"❌"}
    TAB_CORES = {"LABEL CREATED":"808080","COMING TO BRAZIL":"2E75B6","CUSTOMS INSPECTION":"FFB900","NATIONAL TRANSIT":"375623","OUT FOR DELIVERY":"0070C0","DELIVERED":"375623","AWB NAO ENCONTRADO":"C00000","ERRO":"FF0000"}

    for cat in ["COMING TO BRAZIL","CUSTOMS INSPECTION","NATIONAL TRANSIT","OUT FOR DELIVERY","DELIVERED"]:
        df_cat = df[df["CATEGORIA"]==cat].reset_index(drop=True)
        nome = f"{ICONES.get(cat,'')} {cat}"[:31]
        ws = wb.create_sheet(nome); ws.sheet_properties.tabColor = TAB_CORES.get(cat,"FFFFFF")
        escrever_aba(ws, df_cat, cat)
        cores = CORES.get(cat, {"bg":"FFFFFF","font":"000000"})
        ws.insert_rows(1); ws.merge_cells(f"A1:{get_column_letter(len(COLUNAS))}1")
        c_titulo = ws["A1"]
        qtd = len(df_cat)
        c_titulo.value = f"{cat}  —  {qtd} AWB{'s' if qtd != 1 else ''}" if qtd > 0 else f"{cat}  —  Nenhuma AWB no momento"
        c_titulo.font = Font(bold=True, color=cores["font"], name="Arial", size=12)
        c_titulo.fill = make_fill(cores["bg"]); c_titulo.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28; ws.freeze_panes = "A3"

    # ── ABA MUDANÇAS ────────────────────────────────────────────────────────
    ws_mud = wb.create_sheet("🔄 MUDANÇAS"); ws_mud.sheet_properties.tabColor="7030A0"
    TRANSICOES_VALIDAS = {
        ("COMING TO BRAZIL","CUSTOMS INSPECTION"),("COMING TO BRAZIL","NATIONAL TRANSIT"),
        ("COMING TO BRAZIL","OUT FOR DELIVERY"),("COMING TO BRAZIL","DELIVERED"),
        ("CUSTOMS INSPECTION","NATIONAL TRANSIT"),("CUSTOMS INSPECTION","OUT FOR DELIVERY"),
        ("CUSTOMS INSPECTION","DELIVERED"),("NATIONAL TRANSIT","OUT FOR DELIVERY"),
        ("NATIONAL TRANSIT","DELIVERED"),("OUT FOR DELIVERY","DELIVERED"),
        ("LABEL CREATED","COMING TO BRAZIL"),("LABEL CREATED","CUSTOMS INSPECTION"),
        ("LABEL CREATED","NATIONAL TRANSIT"),("LABEL CREATED","OUT FOR DELIVERY"),
        ("LABEL CREATED","DELIVERED"),
    }
    mudancas = pd.DataFrame()
    try:
        path_hist_check = Path(arquivo_historico)
        if path_hist_check.exists():
            hist_df = pd.read_excel(path_hist_check)
            if all(c in hist_df.columns for c in ["DATA_CONSULTA","AWB","CATEGORIA"]):
                hist_df["AWB"] = hist_df["AWB"].astype(str).str.strip()
                hist_df["DATA_CONSULTA"] = pd.to_datetime(hist_df["DATA_CONSULTA"], dayfirst=True, errors="coerce")
                consulta_atual_str = df["DATA_CONSULTA"].iloc[0] if len(df) > 0 else ""
                consulta_atual_dt = pd.to_datetime(consulta_atual_str, dayfirst=True, errors="coerce")
                if pd.notna(consulta_atual_dt):
                    consulta_atual_date = consulta_atual_dt.date()
                    hist_ant = hist_df[hist_df["DATA_CONSULTA"].dt.date < consulta_atual_date]
                else:
                    hist_ant = hist_df
                if len(hist_ant) == 0:
                    pasta = Path(arquivo_historico).parent
                    for _cand in [pasta/"ultimo_status_anterior.xlsx"] + sorted(pasta.glob("snapshot_*.xlsx"),reverse=True) + [pasta/"ultimo_status_gerado.xlsx"]:
                        if not _cand.exists(): continue
                        try:
                            ult_df = pd.read_excel(_cand, sheet_name="📋 TODOS") if "snapshot" not in _cand.name else pd.read_excel(_cand)
                            col_data = "DATA CONSULTA" if "DATA CONSULTA" in ult_df.columns else "DATA_CONSULTA"
                            ult_df = ult_df.rename(columns={col_data:"DATA_CONSULTA"})
                            ult_df["AWB"] = ult_df["AWB"].astype(str).str.strip()
                            ult_df["DATA_CONSULTA"] = pd.to_datetime(ult_df["DATA_CONSULTA"], dayfirst=True, errors="coerce")
                            _cands = ult_df[ult_df["DATA_CONSULTA"] < consulta_atual_dt] if pd.notna(consulta_atual_dt) else ult_df
                            if len(_cands) > 0:
                                hist_ant = _cands; logger.info(f"Fallback: usando '{_cand.name}' ({len(hist_ant)} registros)"); break
                        except Exception as e_ult:
                            logger.warning(f"Fallback '{_cand.name}' falhou: {e_ult}")
                if len(hist_ant) > 0:
                    import re as _re3
                    def _norm_cat(c):
                        c = str(c); c = _re3.sub(r"^[^A-Z]*","",c); c = _re3.sub(r"\s*\(\d+d\)\s*$","",c); c = _re3.sub(r"\s*MEMPHIS.*$","",c); return c.strip()
                    hist_ant = hist_ant.copy(); hist_ant["CATEGORIA"] = hist_ant["CATEGORIA"].apply(_norm_cat)
                    hist_ant["AWB"] = hist_ant["AWB"].astype(str).str.strip()
                    ultimo_ant = hist_ant.sort_values("DATA_CONSULTA").groupby("AWB",as_index=False).last()[["AWB","CATEGORIA"]].rename(columns={"CATEGORIA":"CAT_ANTERIOR"})
                    import re as _re4
                    def _nc(c):
                        c=str(c); c=_re4.sub(r"^⚠\s*","",c); c=_re4.sub(r"\s*\(\d+d\)\s*$","",c); c=_re4.sub(r"\s*MEMPHIS.*$","",c); return c.strip()
                    df_comp = df.copy(); df_comp["AWB"] = df_comp["AWB"].astype(str).str.strip()
                    df_comp = df_comp.merge(ultimo_ant, on="AWB", how="left")
                    df_comp["CATEGORIA_NORM"] = df_comp["CATEGORIA"].apply(_nc)
                    mudancas = df_comp[df_comp.apply(lambda row: pd.notna(row.get("CAT_ANTERIOR")) and bool(row.get("CAT_ANTERIOR")) and (str(row.get("CAT_ANTERIOR")), str(row.get("CATEGORIA_NORM"))) in TRANSICOES_VALIDAS, axis=1)].copy().reset_index(drop=True)
    except Exception as e:
        logger.warning(f"Erro ao comparar histórico para mudanças: {e}")

    n_cols = len(COLUNAS) + 1
    ws_mud.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c_t = ws_mud["A1"]
    c_t.value = f"AWBs QUE MUDARAM DE STATUS  —  {len(mudancas)} alteração(ões)" if len(mudancas) > 0 else "AWBs QUE MUDARAM DE STATUS  —  Nenhuma mudança detectada"
    c_t.font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    c_t.fill = make_fill("7030A0") if len(mudancas) > 0 else make_fill("595959")
    c_t.alignment = Alignment(horizontal="center", vertical="center"); ws_mud.row_dimensions[1].height = 28
    headers_mud = HEADERS + ["STATUS ANTERIOR"]; larguras_mud = LARGURAS + [22]
    for col_idx, (h, larg) in enumerate(zip(headers_mud, larguras_mud), 1):
        cell = ws_mud.cell(row=2, column=col_idx, value=h); header_style(cell)
        ws_mud.column_dimensions[get_column_letter(col_idx)].width = larg
    ws_mud.freeze_panes = "A3"; ws_mud.row_dimensions[2].height = 28
    if len(mudancas) == 0:
        ws_mud.merge_cells(f"A3:{get_column_letter(n_cols)}3")
        c_vazio = ws_mud["A3"]; c_vazio.value = "Nenhuma mudança de status desde a última consulta."
        c_vazio.font = Font(italic=True, color="595959", name="Arial", size=10)
        c_vazio.alignment = Alignment(horizontal="center", vertical="center"); ws_mud.row_dimensions[3].height = 28
    else:
        for row_idx, (_, row) in enumerate(mudancas.iterrows(), 3):
            cat = row["CATEGORIA"]; cat_ant = row.get("CAT_ANTERIOR","")
            cores = CORES.get(cat,{"bg":"FFFFFF","font":"000000"}); cores_ant = CORES.get(cat_ant,{"bg":"EEEEEE","font":"333333"})
            dias_m = row.get("_dias"); data_mem = str(row.get("DATA_CHEGADA_MEMPHIS",""))
            dias_mem_m = None
            if data_mem and data_mem != "nan":
                try:
                    from datetime import date as _d2
                    dias_mem_m = dias_uteis_br(datetime.strptime(data_mem[:10],"%Y-%m-%d").date(), _d2.today())
                except: pass
            em_c = (cat=="CUSTOMS INSPECTION" and dias_m and dias_m>=DIAS_ATRASO_CUSTOMS)
            em_m = (cat=="COMING TO BRAZIL" and dias_mem_m is not None and dias_mem_m>=DIAS_ATRASO_MEMPHIS)
            at = em_c or em_m
            for col_idx, col_name in enumerate(COLUNAS, 1):
                val = row.get(col_name,"")
                cell = ws_mud.cell(row=row_idx, column=col_idx, value=str(val) if pd.notna(val) else "")
                if at: data_style(cell, bg="C00000", fg="FFFFFF", bold=True, wrap=(col_idx in [5,6]))
                else: data_style(cell, bg=cores["bg"], fg=cores["font"], wrap=(col_idx in [5,6]))
            cell_ant = ws_mud.cell(row=row_idx, column=len(COLUNAS)+1, value=cat_ant)
            data_style(cell_ant, bg=cores_ant["bg"], fg=cores_ant["font"], bold=True)
            cell_cat = ws_mud.cell(row=row_idx, column=3); cell_cat.value = f"{cat_ant} → {cat}"
            if at: cell_cat.font=Font(bold=True,color="FFFFFF",name="Arial",size=9); cell_cat.fill=make_fill("C00000")
            else: cell_cat.font=Font(bold=True,color=cores["font"],name="Arial",size=9); cell_cat.fill=make_fill(cores["bg"])
            ws_mud.row_dimensions[row_idx].height = 28

    df_save = df.drop(columns=["_dias","_dias_memphis"], errors="ignore").copy()
    import re as _re2
    def _limpar_cat(c):
        c=str(c); c=_re2.sub(r"^⚠\s*","",c); c=_re2.sub(r"\s*\(\d+d\)\s*$","",c); c=_re2.sub(r"\s*MEMPHIS.*$","",c); return c.strip()
    df_save["CATEGORIA"] = df_save["CATEGORIA"].apply(_limpar_cat)

    import shutil as _shutil
    path_ult_pre = Path(arquivo_ultimo); path_anterior = path_ult_pre.parent/"ultimo_status_anterior.xlsx"
    if path_ult_pre.exists():
        try: _shutil.copy2(str(path_ult_pre), str(path_anterior)); logger.info(f"Anterior salvo → '{path_anterior}'")
        except Exception as _e_s: logger.warning(f"Backup anterior falhou: {_e_s}")

    # ── ABA MENSAL ──────────────────────────────────────────────────────────
    try:
        from datetime import datetime as _dtm
        hoje = _dtm.now(); mes_ant = (hoje.month-2)%12+1; ano_ant = hoje.year if hoje.month>1 else hoje.year-1
        nome_mes = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"][mes_ant-1]
        df_rel = df[df["CATEGORIA"].apply(lambda c:"DELIVERED" in str(c).upper())].copy()
        if "DATA_ENTREGA" in df_rel.columns:
            df_rel["_de"] = pd.to_datetime(df_rel["DATA_ENTREGA"],errors="coerce")
            df_rel = df_rel[(df_rel["_de"].dt.month==mes_ant)&(df_rel["_de"].dt.year==ano_ant)]
        ws_rel = wb.create_sheet(f"📅 MENSAL {nome_mes.upper()[:3]}"); ws_rel.sheet_properties.tabColor="1F4E79"
        REGIOES = ["América do Norte","Europa","Oriente Médio","Ásia","Outros"]
        COR_REG = {"América do Norte":"BDD7EE","Europa":"C6EFCE","Oriente Médio":"FFE699","Ásia":"F4CCFF","Outros":"D9D9D9"}
        ws_rel.merge_cells("A1:H1"); ct=ws_rel["A1"]; ct.value=f"RELATÓRIO MENSAL — {nome_mes.upper()} {ano_ant}"
        ct.font=Font(bold=True,color="FFFFFF",name="Arial",size=13); ct.fill=make_fill("1F4E79")
        ct.alignment=Alignment(horizontal="center",vertical="center"); ws_rel.row_dimensions[1].height=32
        row_idx=3; total_geral=len(df_rel)
        ws_rel.merge_cells(f"A{row_idx}:H{row_idx}"); c=ws_rel[f"A{row_idx}"]
        c.value=f"Total geral Expresso: {total_geral} envios"; c.font=Font(bold=True,name="Arial",size=11)
        c.alignment=Alignment(vertical="center"); ws_rel.row_dimensions[row_idx].height=24; row_idx+=1
        totais_reg = df_rel["REGIAO"].value_counts().to_dict() if "REGIAO" in df_rel.columns else {}
        partes = [f"{r}: {totais_reg.get(r,0)}" for r in REGIOES if totais_reg.get(r,0)>0]
        ws_rel.merge_cells(f"A{row_idx}:H{row_idx}"); c2=ws_rel[f"A{row_idx}"]
        c2.value="  •  "+"  •  ".join(partes) if partes else "Sem dados de região"
        c2.font=Font(name="Arial",size=10,color="595959"); c2.alignment=Alignment(vertical="center")
        ws_rel.row_dimensions[row_idx].height=20; row_idx+=2
        for regiao in REGIOES:
            df_r = df_rel[df_rel["REGIAO"].astype(str).str.strip()==regiao] if "REGIAO" in df_rel.columns else pd.DataFrame()
            if len(df_r)==0: continue
            cor=COR_REG.get(regiao,"EEEEEE")
            ws_rel.merge_cells(f"A{row_idx}:H{row_idx}"); ch=ws_rel[f"A{row_idx}"]
            ch.value=regiao.upper(); ch.font=Font(bold=True,color="FFFFFF",name="Arial",size=11)
            ch.fill=make_fill("2E4057"); ch.alignment=Alignment(horizontal="center",vertical="center")
            ws_rel.row_dimensions[row_idx].height=24; row_idx+=1
            lt_vals=df_r["LEAD_TIME_DIAS"].dropna().astype(float) if "LEAD_TIME_DIAS" in df_r.columns else pd.Series()
            media_lt=round(lt_vals.mean()) if len(lt_vals)>0 else "-"
            min_lt=int(lt_vals.min()) if len(lt_vals)>0 else "-"
            max_lt=int(lt_vals.max()) if len(lt_vals)>0 else "-"
            prod_top=df_r["PRODUTO"].value_counts().index[0] if "PRODUTO" in df_r.columns and len(df_r)>0 else "-"
            for col_i,(label,valor) in enumerate([("Total envios",str(len(df_r))),("Lead time médio",f"{media_lt} dias"),("Lead time mín.",f"{min_lt} dias"),("Lead time máx.",f"{max_lt} dias"),("Produto top",str(prod_top))],1):
                cl=ws_rel.cell(row=row_idx,column=col_i*2-1,value=label)
                cl.font=Font(bold=True,name="Arial",size=9,color="595959"); cl.alignment=Alignment(horizontal="center",vertical="center")
                cv=ws_rel.cell(row=row_idx,column=col_i*2,value=valor)
                cv.font=Font(bold=True,name="Arial",size=11); cv.fill=make_fill(cor); cv.alignment=Alignment(horizontal="center",vertical="center")
            ws_rel.row_dimensions[row_idx].height=28; row_idx+=2
            for ci,h in enumerate(["AWB","PEDIDO","PRODUTO","PAÍS","DATA ENVIO","DATA ENTREGA","LEAD TIME","STATUS"],1):
                c=ws_rel.cell(row=row_idx,column=ci,value=h); header_style(c)
                ws_rel.column_dimensions[get_column_letter(ci)].width=[18,10,22,8,14,14,12,18][ci-1]
            ws_rel.row_dimensions[row_idx].height=22; row_idx+=1
            for _,rr in df_r.iterrows():
                lt_val=rr.get("LEAD_TIME_DIAS",""); lt_str=f"{int(lt_val)}d" if lt_val and str(lt_val) not in ("","nan","None") else "-"
                vals=[str(rr.get("AWB","")),str(rr.get("PEDIDO","")),str(rr.get("PRODUTO","")),str(rr.get("PAIS_ORIGEM","")),str(rr.get("DATA_CRIACAO",""))[:10],str(rr.get("DATA_ENTREGA",""))[:10],lt_str,str(rr.get("STATUS_FEDEX",""))[:30]]
                for ci,v in enumerate(vals,1):
                    c=ws_rel.cell(row=row_idx,column=ci,value=v); data_style(c,bg=cor,fg="000000",wrap=False)
                ws_rel.row_dimensions[row_idx].height=20; row_idx+=1
            row_idx+=1
    except Exception as _e_rel:
        logger.warning(f"Aba mensal: {_e_rel}")

    wb.save(arquivo_ultimo)
    logger.info(f"Relatório salvo → '{arquivo_ultimo}'")
    _fazer_backup(arquivo_historico)
    path_hist = Path(arquivo_historico)
    if path_hist.exists():
        hist = pd.read_excel(path_hist)
        final = pd.concat([hist, df_save], ignore_index=True)
    else:
        final = df_save
    final = final.drop_duplicates(subset=["AWB","DATA_CONSULTA"], keep="last")
    try:
        if "DATA_CONSULTA" in final.columns:
            final["_dc_dt"] = pd.to_datetime(final["DATA_CONSULTA"], dayfirst=True, errors="coerce")
            cutoff = pd.Timestamp.now() - pd.DateOffset(months=3)
            final = final[final["_dc_dt"].isna()|(final["_dc_dt"]>=cutoff)]
            final = final.drop(columns=["_dc_dt"])
    except Exception as _eh:
        logger.warning(f"Não foi possível truncar histórico: {_eh}")
    final.to_excel(arquivo_historico, index=False)
    logger.info(f"Histórico → '{arquivo_historico}' ({len(final)} registros)")


def gerar_resumo(resultados) -> str:
    total = len(resultados); sucessos = sum(1 for r in resultados if r["SUCESSO"])
    ordem = ["LABEL CREATED","COMING TO BRAZIL","CUSTOMS INSPECTION","NATIONAL TRANSIT","OUT FOR DELIVERY","DELIVERED","AWB NAO ENCONTRADO","ERRO"]
    contagem = {}
    for r in resultados: contagem[r["CATEGORIA"]] = contagem.get(r["CATEGORIA"],0)+1
    linhas = ["","="*60,f"  RESUMO DO RASTREIO — {datetime.now().strftime('%d/%m/%Y %H:%M')}","="*60,
              f"  Total consultado : {total}",f"  Com sucesso      : {sucessos}",f"  Com erro         : {total-sucessos}","","  Por categoria (ordem de progressão):"]
    for cat in ordem:
        if cat in contagem: linhas.append(f"    {cat:<30} {contagem[cat]:>4}x")
    linhas.append("="*60)
    return "\n".join(linhas)


# ==========================================================
# HTML
# ==========================================================

def gerar_html_relatorio(arquivo_excel: str, arquivo_html: str, arquivo_historico: str = "") -> None:
    import json, re as _re_html
    logger = logging.getLogger("fedex_tracker")

    def _norm(c):
        c=str(c); c=_re_html.sub(r"^[^A-Z]*","",c); c=_re_html.sub(r"\s*\(\d+d\)\s*$","",c); c=_re_html.sub(r"\s*MEMPHIS.*$","",c); return c.strip()

    try: sheets = pd.read_excel(arquivo_excel, sheet_name=None)
    except Exception as e: logger.warning(f"HTML: não foi possível abrir {arquivo_excel}: {e}"); return

    df = sheets.get("📋 TODOS", pd.DataFrame())
    if df.empty: return
    df.columns = df.columns.str.strip()
    df = df.rename(columns={"STATUS FEDEX":"STATUS_FEDEX","ÚLTIMO EVENTO":"ULTIMO_EVENTO","DATA CONSULTA":"DATA_CONSULTA","CHEGADA MEMPHIS":"DATA_CHEGADA_MEMPHIS","REGIÃO":"REGIAO","TIMELINE":"TIMELINE_JSON","EVENTOS":"EVENTOS_JSON","DEST":"ENDERECO_ENTREGA"})
    df["CATEGORIA_NORM"] = df["CATEGORIA"].apply(_norm)
    df = df.fillna("")

    def _recalc_dias(s):
        try:
            partes = str(s).split("—"); data_str = partes[-1].strip().split(" ")
            data_part = [p for p in data_str if re.match(r"\d{4}-\d{2}-\d{2}",p)]
            if data_part:
                from datetime import date
                return dias_uteis_br(datetime.strptime(data_part[0],"%Y-%m-%d").date(), date.today())
        except: pass
        return None

    def _recalc_dias_memphis(s):
        try:
            s=str(s)
            if s and s not in ("nan",""):
                from datetime import date
                return dias_uteis_br(datetime.strptime(s[:10],"%Y-%m-%d").date(), date.today())
        except: pass
        return None

    df["_dias"] = df["ULTIMO_EVENTO"].apply(_recalc_dias)
    df["_dias_memphis"] = df.get("DATA_CHEGADA_MEMPHIS", pd.Series(dtype=str)).apply(_recalc_dias_memphis)

    df_mud = sheets.get("🔄 MUDANÇAS", pd.DataFrame())
    data_consulta = df["DATA_CONSULTA"].iloc[0] if len(df)>0 else ""
    total = len(df)
    ORDEM = ["COMING TO BRAZIL","CUSTOMS INSPECTION","NATIONAL TRANSIT","OUT FOR DELIVERY","DELIVERED","LABEL CREATED"]
    ICONS = {"COMING TO BRAZIL":"✈","CUSTOMS INSPECTION":"🔍","NATIONAL TRANSIT":"🚚","OUT FOR DELIVERY":"📦","DELIVERED":"✅","LABEL CREATED":"🏷"}
    cat_classes = {"COMING TO BRAZIL":"coming","CUSTOMS INSPECTION":"customs","NATIONAL TRANSIT":"national","OUT FOR DELIVERY":"ofd","DELIVERED":"delivered","LABEL CREATED":"label"}
    cats = df["CATEGORIA_NORM"].value_counts().to_dict()

    mudancas_rows = []
    if not df_mud.empty:
        df_mud.columns = df_mud.columns.str.strip()
        col_data = "DATA_CONSULTA" if "DATA_CONSULTA" in df_mud.columns else None
        for _, r in df_mud.iterrows():
            vals = [str(v) if pd.notna(v) else "" for v in r.values]
            if vals[0] and vals[0] not in ("nan","None") and _re_html.match(r"\d{9,}",vals[0]):
                ts = str(r[col_data]) if col_data and pd.notna(r.get(col_data,"")) else data_consulta
                mudancas_rows.append(vals[:3]+[ts])

    _hist_rows = {}; _eta_map = {}
    try:
        _path_h = Path(arquivo_historico if arquivo_historico else "historico_status.xlsx")
        if _path_h.exists():
            _hdf = pd.read_excel(_path_h); _hdf.columns = _hdf.columns.str.strip()
            if all(c in _hdf.columns for c in ["AWB","CATEGORIA","DATA_CONSULTA"]):
                _hdf["AWB"] = _hdf["AWB"].astype(str).str.strip()
                _hdf["_dc"] = pd.to_datetime(_hdf["DATA_CONSULTA"],dayfirst=True,errors="coerce")
                _hdf = _hdf.sort_values("_dc")
                for _awb, _grp in _hdf.groupby("AWB"):
                    _timeline = []; _prev = None
                    for _, _row in _grp.iterrows():
                        import re as _re2
                        _cat = _re2.sub(r"^⚠\s*","",str(_row.get("CATEGORIA",""))).strip()
                        _cat = _re2.sub(r"\s*\(\d+d\)\s*$","",_cat).strip()
                        _cat = _re2.sub(r"\s*MEMPHIS.*$","",_cat).strip()
                        _dt = str(_row["_dc"])[:10] if pd.notna(_row["_dc"]) else ""
                        if _cat != _prev and _cat and _cat not in ("nan","ERRO"):
                            _timeline.append({"cat":_cat,"data":_dt}); _prev=_cat
                    _hist_rows[_awb] = _timeline
                _delivered = _hdf[_hdf["CATEGORIA"].astype(str).str.upper().str.contains("DELIVERED",na=False)].copy()
                if "LEAD_TIME_DIAS" in _delivered.columns and "REGIAO" in _delivered.columns:
                    for _reg, _grp2 in _delivered.groupby("REGIAO"):
                        _lt = pd.to_numeric(_grp2["LEAD_TIME_DIAS"],errors="coerce").dropna()
                        if len(_lt)>=2: _eta_map[str(_reg)] = round(float(_lt.mean()))
    except: pass

    rows_json = []
    for _, r in df.iterrows():
        _awb = str(r["AWB"]); _dias = r.get("_dias")
        _dias_val = int(_dias) if _dias is not None and not pd.isna(_dias) else None
        _regiao = str(r.get("REGIAO","")) if "REGIAO" in df.columns else ""
        _eta_dias = _eta_map.get(_regiao); _eta_str = ""
        if _eta_dias and _dias_val is not None and r["CATEGORIA_NORM"] not in ("DELIVERED","ERRO","AWB NAO ENCONTRADO"):
            _eta_str = f"~{max(0,_eta_dias-_dias_val)}d úteis"
        _tl_raw = str(r.get("TIMELINE_JSON","")) if "TIMELINE_JSON" in df.columns else ""
        _timeline = []
        if _tl_raw and _tl_raw not in ("nan",""):
            try: _timeline = json.loads(_tl_raw)
            except: _timeline = _hist_rows.get(_awb,[])
        else: _timeline = _hist_rows.get(_awb,[])
        _ev_raw = str(r.get("EVENTOS_JSON","")) if "EVENTOS_JSON" in df.columns else ""
        _eventos = []
        if _ev_raw and _ev_raw not in ("nan",""):
            try: _eventos = json.loads(_ev_raw)
            except: pass
        rows_json.append({
            "awb":_awb,"pedido":str(r["PEDIDO"]),"produto":str(r.get("PRODUTO","")) if "PRODUTO" in df.columns else "",
            "cat":r["CATEGORIA_NORM"],"status":str(r.get("STATUS_FEDEX","")),"evento":str(r.get("ULTIMO_EVENTO","")),
            "data":str(r.get("DATA_CONSULTA","")),"alerta":"⚠" in str(r["CATEGORIA"]),
            "dias":_dias_val,"eta":_eta_str,"timeline":_timeline,"eventos":_eventos,
            "endereco":str(r.get("ENDERECO_ENTREGA","")) if "ENDERECO_ENTREGA" in df.columns else "",
            "fp":{},"risco":{},"previsao":{},"fator_cal":1.0,"desc_cal":"",
        })

    kpi_html = ""
    for cat in ORDEM:
        n=cats.get(cat,0); cls=cat_classes.get(cat,"label"); icon=ICONS.get(cat,"•")
        short=cat.replace("COMING TO BRAZIL","COMING").replace("CUSTOMS INSPECTION","CUSTOMS").replace("NATIONAL TRANSIT","NATIONAL").replace("OUT FOR DELIVERY","OUT FOR DEL.").replace("LABEL CREATED","LABEL")
        kpi_html += f'  <div class="kpi {cls}" onclick="filterCat(\'{cat}\')" data-cat="{cat}"><div class="kpi-icon">{icon}</div><div class="kpi-num">{n}</div><div class="kpi-label">{short}</div></div>\n'

    filter_tags = ""
    for cat in ORDEM:
        icon=ICONS.get(cat,"")
        short=cat.replace("COMING TO BRAZIL","COMING").replace("CUSTOMS INSPECTION","CUSTOMS").replace("NATIONAL TRANSIT","NATIONAL").replace("OUT FOR DELIVERY","OUT FOR DEL.").replace("LABEL CREATED","LABEL")
        filter_tags += f'  <button class="filter-tag" onclick="filterCat(\'{cat}\')" data-ftag="{cat}">{icon} {short}</button>\n'

    # ── Monta JSON mensal ───────────────────────────────────────────────────
    REGIOES_DISPLAY = ["América do Norte","Europa","Oriente Médio","Ásia","Outros"]
    def _build_mensal_json(arquivo_hist, df_atual):
        frames = []
        if arquivo_hist:
            try:
                df_h = pd.read_excel(arquivo_hist); df_h.columns = df_h.columns.str.strip(); frames.append(df_h)
            except: pass
        frames.append(df_atual)
        if not frames: return {"meses":[]}
        df_all = pd.concat(frames, ignore_index=True); df_all.columns = df_all.columns.str.strip()
        if "CATEGORIA" not in df_all.columns and "CATEGORIA_NORM" in df_all.columns:
            df_all["CATEGORIA"] = df_all["CATEGORIA_NORM"]
        mask_del = df_all.get("CATEGORIA",pd.Series(dtype=str)).astype(str).str.upper().str.contains("DELIVERED",na=False)
        df_del = df_all[mask_del].copy()
        if df_del.empty or "DATA_ENTREGA" not in df_del.columns: return {"meses":[]}
        df_del["_de"] = pd.to_datetime(df_del["DATA_ENTREGA"],errors="coerce")
        df_del = df_del[df_del["_de"].notna()]
        if df_del.empty: return {"meses":[]}
        if "AWB" in df_del.columns:
            df_del = df_del.sort_values("_de").drop_duplicates(subset=["AWB"],keep="last")
        df_del["_mes"] = df_del["_de"].dt.month; df_del["_ano"] = df_del["_de"].dt.year
        df_del["_chave"] = df_del["_ano"].astype(str)+"-"+df_del["_mes"].astype(str).str.zfill(2)
        nomes_mes = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
        chaves = sorted(df_del["_chave"].unique(),reverse=True); meses_out = []
        for chave in chaves:
            df_m = df_del[df_del["_chave"]==chave]
            ano_m=int(chave.split("-")[0]); mes_m=int(chave.split("-")[1])
            periodo=f"{nomes_mes[mes_m-1]} {ano_m}"; regioes_out=[]
            for reg in REGIOES_DISPLAY:
                df_r = df_m[df_m["REGIAO"].astype(str).str.strip()==reg] if "REGIAO" in df_m.columns else pd.DataFrame()
                if df_r.empty: continue
                lt = pd.to_numeric(df_r["LEAD_TIME_DIAS"],errors="coerce") if "LEAD_TIME_DIAS" in df_r.columns else pd.Series(dtype=float)
                lt_valid = lt.dropna()
                prod_top = str(df_r["PRODUTO"].value_counts().index[0]) if "PRODUTO" in df_r.columns and not df_r["PRODUTO"].isna().all() else "-"
                awbs_list = []
                for _, row in df_r.iterrows():
                    lt_val = lt.get(row.name)
                    awbs_list.append({"awb":str(row.get("AWB","")),"pedido":str(row.get("PEDIDO","")).replace(".0","") if pd.notna(row.get("PEDIDO","")) else "","produto":str(row.get("PRODUTO","")) if pd.notna(row.get("PRODUTO","")) else "","lt":int(lt_val) if pd.notna(lt_val) else None})
                awbs_com_lt = [a for a in awbs_list if a["lt"] is not None]
                awb_min = min(awbs_com_lt,key=lambda a:a["lt"]) if awbs_com_lt else None
                awb_max = max(awbs_com_lt,key=lambda a:a["lt"]) if awbs_com_lt else None
                regioes_out.append({"nome":reg,"total":len(df_r),"media_lt":round(float(lt_valid.mean())) if len(lt_valid)>0 else "-","min_lt":int(lt_valid.min()) if len(lt_valid)>0 else "-","max_lt":int(lt_valid.max()) if len(lt_valid)>0 else "-","awb_min":awb_min,"awb_max":awb_max,"produto_top":prod_top,"awbs":awbs_list})
            meses_out.append({"periodo":periodo,"total":len(df_m),"regioes":regioes_out})
        return {"meses":meses_out}

    mensal_json = _build_mensal_json(arquivo_historico, df)

    # ── Template HTML ───────────────────────────────────────────────────────
    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>FedEx Tracker — {data_consulta}</title>
<meta http-equiv="refresh" content="3600">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
:root{{--bg:#0d1117;--surface:#161b2a;--surface2:#1c2238;--surface3:#212840;--border:#21293f;--border2:#2d3755;--text:#dde1f0;--muted:#4b5577;--muted2:#6b7599;--accent:#4f7dff;--accent2:#3d6be8;--coming:#1a3a6c;--coming-fg:#93c5fd;--coming-bar:#4f7dff;--customs:#3d2e00;--customs-fg:#fcd34d;--customs-bar:#f59e0b;--national:#0f3320;--national-fg:#6ee7b7;--national-bar:#10b981;--ofd:#1e2f5c;--ofd-fg:#a5b4fc;--ofd-bar:#818cf8;--delivered:#052e16;--delivered-fg:#4ade80;--delivered-bar:#22c55e;--label:#161b2a;--label-fg:#64748b;--label-bar:#475569;--alert:#3d0f0f;--alert-fg:#fca5a5;font-size:14px}}
body{{background:var(--bg);color:var(--text);font-family:'DM Sans',sans-serif;min-height:100vh}}
.header{{background:var(--surface);border-bottom:1px solid var(--border);padding:16px 32px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:100}}
.logo{{font-family:'DM Mono',monospace;font-size:12px;color:var(--accent);letter-spacing:.1em;display:flex;align-items:center;gap:10px}}
.logo-dot{{width:7px;height:7px;background:var(--accent);border-radius:50%;animation:pulse 2.5s ease-in-out infinite}}
@keyframes pulse{{0%,100%{{opacity:1;transform:scale(1)}}50%{{opacity:.3;transform:scale(.7)}}}}
.header-right{{display:flex;align-items:center;gap:20px}}
.header-date{{font-family:'DM Mono',monospace;font-size:11px;color:var(--muted2)}}
.header-countdown{{font-family:'DM Mono',monospace;font-size:11px;color:var(--muted);background:var(--surface2);padding:4px 10px;border-radius:20px;border:1px solid var(--border)}}
.kpi-strip{{display:grid;grid-template-columns:repeat(6,1fr);gap:10px;padding:20px 32px 0}}
.kpi{{background:var(--surface);border:1px solid var(--border);border-left:3px solid transparent;border-radius:0 8px 8px 0;padding:14px 16px;cursor:pointer;transition:background .15s,border-color .15s,transform .15s;position:relative}}
.kpi:hover{{background:var(--surface2);transform:translateY(-1px)}}.kpi.active{{background:var(--surface2);border-color:var(--accent)!important}}
.kpi.coming{{border-left-color:var(--coming-bar)}}.kpi.customs{{border-left-color:var(--customs-bar)}}.kpi.national{{border-left-color:var(--national-bar)}}.kpi.ofd{{border-left-color:var(--ofd-bar)}}.kpi.delivered{{border-left-color:var(--delivered-bar)}}.kpi.label{{border-left-color:var(--label-bar)}}
.kpi-icon{{font-size:16px;margin-bottom:6px}}.kpi-num{{font-family:'DM Mono',monospace;font-size:26px;font-weight:500;line-height:1}}.kpi-label{{font-size:9px;color:var(--muted2);text-transform:uppercase;letter-spacing:.12em;margin-top:5px}}
.mid-strip{{display:grid;grid-template-columns:1fr 1fr;gap:10px;padding:10px 32px 0}}
.panel{{background:var(--surface);border:1px solid var(--border);border-radius:8px;padding:18px}}
.panel-title{{font-size:10px;text-transform:uppercase;letter-spacing:.14em;color:var(--muted2);margin-bottom:14px;display:flex;align-items:center;gap:8px}}
.panel-title span{{color:var(--text);font-size:12px;letter-spacing:0;text-transform:none}}
.chart-bars{{display:flex;flex-direction:column;gap:8px}}
.bar-row{{display:grid;grid-template-columns:110px 1fr 34px;align-items:center;gap:8px}}
.bar-name{{font-size:10px;color:var(--muted2);text-align:right}}.bar-track{{background:var(--surface2);border-radius:3px;height:14px;overflow:hidden}}.bar-fill{{height:100%;border-radius:3px;transition:width .9s cubic-bezier(.4,0,.2,1)}}.bar-pct{{font-family:'DM Mono',monospace;font-size:10px;color:var(--muted2);text-align:right}}
.mud-filter{{padding:3px 9px;border-radius:12px;border:1px solid var(--border);background:transparent;color:var(--muted2);font-size:9px;cursor:pointer;transition:.15s;font-family:'DM Sans',sans-serif;letter-spacing:.06em}}
.mud-filter.active{{background:var(--accent);border-color:var(--accent);color:#fff}}
.mud-list{{display:flex;flex-direction:column;gap:6px;max-height:210px;overflow-y:auto}}.mud-list::-webkit-scrollbar{{width:3px}}.mud-list::-webkit-scrollbar-thumb{{background:var(--border2);border-radius:3px}}
.mud-row{{background:var(--surface2);border-left:2px solid var(--border2);border-radius:0 6px 6px 0;padding:8px 12px;display:flex;align-items:center;gap:10px;font-size:12px;transition:border-color .15s}}
.mud-row:hover{{border-left-color:var(--accent)}}
.mud-awb{{font-family:'DM Mono',monospace;font-size:10px;color:var(--muted2);min-width:105px}}.mud-pedido{{font-family:'DM Mono',monospace;font-size:10px;color:var(--accent);min-width:48px}}
.mud-trans{{display:flex;align-items:center;gap:5px;flex:1}}.mud-cat{{padding:2px 7px;border-radius:3px;font-size:9px;font-weight:600;text-transform:uppercase;letter-spacing:.06em}}.mud-arrow{{color:var(--muted);font-size:10px}}.mud-empty{{color:var(--muted);font-size:12px;text-align:center;padding:20px}}
.search-bar{{padding:14px 32px 0;display:flex;gap:8px;align-items:center;flex-wrap:wrap}}
.search-input{{flex:1;min-width:200px;background:var(--surface);border:1px solid var(--border);border-radius:7px;padding:9px 14px;color:var(--text);font-family:'DM Sans',sans-serif;font-size:12px;outline:none;transition:border-color .15s}}.search-input:focus{{border-color:var(--accent)}}.search-input::placeholder{{color:var(--muted)}}
.filter-tag{{padding:5px 12px;border-radius:20px;border:1px solid var(--border);background:transparent;color:var(--muted2);font-size:10px;cursor:pointer;transition:.15s;font-family:'DM Sans',sans-serif;text-transform:uppercase;letter-spacing:.08em}}
.filter-tag.active,.filter-tag:hover{{background:var(--accent);border-color:var(--accent);color:#fff}}
.table-wrap{{padding:12px 32px 32px;overflow-x:auto}}
table{{width:100%;border-collapse:separate;border-spacing:0 3px}}
thead th{{font-size:9px;text-transform:uppercase;letter-spacing:.14em;color:var(--muted);padding:6px 14px;text-align:left}}
tbody tr{{background:var(--surface);border-left:3px solid transparent;transition:background .12s,border-color .12s;cursor:pointer}}
tbody tr:hover{{background:var(--surface2)}}
tbody tr.alerta{{background:var(--alert)!important;border-left-color:#ef4444!important}}
tbody tr.alerta:hover{{background:#4a1010!important}}
td{{padding:10px 14px;vertical-align:middle;font-size:11px}}
td:first-child{{border-radius:0}}.td-last{{border-radius:0 6px 6px 0}}
.cat-badge{{display:inline-flex;align-items:center;gap:4px;padding:3px 9px;border-radius:4px;font-size:9px;font-weight:600;text-transform:uppercase;letter-spacing:.07em;white-space:nowrap}}
.pedido-badge{{font-family:'DM Mono',monospace;font-size:10px;color:var(--accent)}}
.evento-cell{{max-width:280px;color:var(--muted2);font-size:10px;line-height:1.5}}
.tab-bar{{display:flex;border-bottom:1px solid var(--border);padding:0 32px;margin-top:14px;background:var(--surface)}}
.tab-btn{{padding:11px 18px;font-size:10px;color:var(--muted2);cursor:pointer;border-bottom:2px solid transparent;margin-bottom:-1px;transition:.15s;text-transform:uppercase;letter-spacing:.1em;background:none;border-top:none;border-left:none;border-right:none;font-family:'DM Sans',sans-serif}}
.tab-btn.active{{color:var(--text);border-bottom-color:var(--accent)}}
.tab-pane{{display:none}}.tab-pane.active{{display:block}}
.mensal-wrap{{padding:20px 32px}}
.mensal-regiao{{background:var(--surface);border:1px solid var(--border);border-radius:8px;margin-bottom:16px;overflow:visible}}
.mensal-reg-header{{padding:12px 18px;font-size:11px;font-weight:600;letter-spacing:.08em;text-transform:uppercase;color:#fff;border-left:3px solid transparent;border-radius:0}}
.mensal-kpis{{display:grid;grid-template-columns:repeat(5,1fr);border-top:1px solid var(--border)}}
.mensal-kpi{{background:var(--surface2);padding:14px;text-align:center;border-right:1px solid var(--border)}}.mensal-kpi:last-child{{border-right:none}}
.mensal-kpi-label{{font-size:9px;color:var(--muted2);text-transform:uppercase;letter-spacing:.1em;margin-bottom:5px}}
.mensal-kpi-val{{font-family:'DM Mono',monospace;font-size:17px;font-weight:500}}
.mensal-kpi-val.destaque{{color:var(--accent);font-size:13px}}
.mensal-sem-dados{{color:var(--muted);font-size:13px;text-align:center;padding:40px}}
.mensal-kpi-clickable{{cursor:pointer;position:relative}}.mensal-kpi-clickable:hover{{background:var(--surface3)!important}}
.mensal-pop{{display:none;position:fixed;z-index:9999;background:var(--surface2);border:1px solid var(--accent);border-radius:8px;padding:12px 14px;min-width:230px;box-shadow:0 8px 24px #0009;font-size:12px;pointer-events:none}}
.mensal-pop.open{{display:block}}
.mensal-pop-awb{{font-family:'DM Mono',monospace;font-size:13px;font-weight:600;color:var(--text);margin-bottom:8px;padding-bottom:6px;border-bottom:1px solid var(--border2)}}
.mensal-pop-row{{display:flex;justify-content:space-between;gap:12px;padding:3px 0;color:var(--muted2)}}.mensal-pop-row span:last-child{{color:var(--text);text-align:right}}
.sparkline-wrap{{padding:14px 18px 10px;border-top:1px solid var(--border)}}.sparkline-title{{font-size:9px;color:var(--muted2);text-transform:uppercase;letter-spacing:.1em;margin-bottom:8px}}
thead th.sortable{{cursor:pointer;user-select:none}}thead th.sortable:hover{{color:var(--text)}}thead th.sort-asc::after{{content:" ↑"}}thead th.sort-desc::after{{content:" ↓"}}
.dias-cell{{font-family:'DM Mono',monospace;font-size:10px;text-align:center}}
.heat-0{{color:var(--muted)}}.heat-1{{color:#6ee7b7}}.heat-2{{color:#fcd34d}}.heat-3{{color:#fb923c}}.heat-4{{color:#f87171;font-weight:600}}
.eta-cell{{font-family:'DM Mono',monospace;font-size:9px;color:var(--muted2);text-align:center}}
.modal-overlay{{display:none;position:fixed;inset:0;background:#000a;z-index:1000;align-items:center;justify-content:center;padding:24px}}
.modal-overlay.open{{display:flex}}
.modal{{background:var(--surface);border:1px solid var(--border2);border-radius:12px;width:100%;max-width:720px;max-height:88vh;display:flex;flex-direction:column;overflow:hidden}}
.modal-header{{padding:18px 22px;border-bottom:1px solid var(--border);display:flex;align-items:flex-start;justify-content:space-between;gap:16px;flex-shrink:0}}
.modal-awb{{font-family:'DM Mono',monospace;font-size:17px;font-weight:600;color:var(--text)}}.modal-sub{{font-size:11px;color:var(--muted2);margin-top:3px}}
.modal-close{{background:none;border:1px solid var(--border);color:var(--muted2);cursor:pointer;font-size:16px;padding:4px 10px;border-radius:6px;line-height:1;transition:.15s;font-family:inherit}}.modal-close:hover{{color:var(--text);border-color:var(--muted2)}}
.modal-meta{{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));border-bottom:1px solid var(--border);flex-shrink:0}}
.modal-meta-item{{padding:12px 20px;border-right:1px solid var(--border)}}.modal-meta-item:last-child{{border-right:none}}
.modal-meta-label{{font-size:9px;text-transform:uppercase;letter-spacing:.1em;color:var(--muted2);margin-bottom:4px}}.modal-meta-val{{font-size:12px;color:var(--text);font-weight:500}}
.modal-body{{overflow-y:auto;flex:1}}.modal-body::-webkit-scrollbar{{width:4px}}.modal-body::-webkit-scrollbar-thumb{{background:var(--border2);border-radius:4px}}
.modal-day-group{{border-bottom:1px solid var(--border)}}.modal-day-group:last-child{{border-bottom:none}}
.modal-day-header{{padding:9px 22px;font-size:10px;font-weight:600;color:var(--muted2);text-transform:uppercase;letter-spacing:.1em;background:var(--surface2);position:sticky;top:0}}
.modal-event-row{{display:grid;grid-template-columns:72px 1fr auto;align-items:start;gap:12px;padding:10px 22px;border-top:1px solid var(--border);transition:background .1s;cursor:default}}
.modal-event-row:first-child{{border-top:none}}.modal-event-row:hover{{background:var(--surface2)}}
.modal-event-time{{font-family:'DM Mono',monospace;font-size:10px;color:var(--muted2);padding-top:1px}}
.modal-event-desc{{font-size:11px;color:var(--text);line-height:1.5}}.modal-event-subdesc{{font-size:10px;color:var(--muted2);margin-top:2px}}
.modal-event-local{{font-size:10px;color:var(--muted2);text-align:right;white-space:nowrap;padding-top:1px}}
.modal-intel{{padding:12px 20px;border-bottom:1px solid var(--border);display:flex;gap:8px;flex-wrap:wrap;align-items:center;flex-shrink:0;background:var(--surface2)}}
.intel-badge{{display:inline-flex;align-items:center;gap:5px;padding:4px 10px;border-radius:20px;font-size:10px;font-weight:600;border:1px solid;cursor:default}}
.intel-normal{{background:#0f3320;color:#6ee7b7;border-color:#185a30}}.intel-atencao{{background:#3d2e00;color:#fcd34d;border-color:#7a5c00}}
.intel-critico{{background:#3d0f0f;color:#fca5a5;border-color:#7a1f1f}}.intel-rapido{{background:#0d1a36;color:#93c5fd;border-color:#1a3a6c}}
.intel-risco{{background:#3d1a00;color:#fb923c;border-color:#7a3400}}.intel-previsao{{background:#1a0d36;color:#a78bfa;border-color:#3d1a7a}}
.intel-cal{{background:#0d2a0d;color:#86efac;border-color:#1a5a1a}}
tbody tr:hover td{{background:var(--surface3)!important}}
</style></head><body>
<div class="modal-overlay" id="modal-overlay" onclick="closeModal(event)">
  <div class="modal" id="modal">
    <div class="modal-header" id="modal-header"></div>
    <div class="modal-meta" id="modal-meta"></div>
    <div class="modal-intel" id="modal-intel" style="display:none"></div>
    <div class="modal-body" id="modal-body"></div>
  </div>
</div>
<div class="header">
  <div class="logo"><div class="logo-dot"></div>FEDEX TRACKER</div>
  <div class="header-right">
    <div class="header-date">Atualizado em {data_consulta}</div>
    <div class="header-countdown" id="countdown">↻ próxima em —</div>
  </div>
</div>
<div class="tab-bar">
  <button class="tab-btn active" onclick="switchTab('painel')">Painel</button>
  <button class="tab-btn" onclick="switchTab('mensal')">Relatório Mensal</button>
  <button class="tab-btn" onclick="switchTab('periodo')">Relatório por Período</button>
</div>
<div id="tab-painel" class="tab-pane active">
<div class="kpi-strip" id="kpi-strip">
{kpi_html}</div>
<div class="mid-strip">
  <div class="panel">
    <div class="panel-title">Distribuição <span id="total-badge"></span></div>
    <div class="chart-bars" id="chart-bars"></div>
  </div>
  <div class="panel">
    <div class="panel-title">🔄 Mudanças detectadas <span id="mud-count"></span></div>
    <div style="display:flex;gap:5px;margin-bottom:10px;flex-wrap:wrap;align-items:center">
      <button class="mud-filter mud-filter-top active" onclick="setMudFiltro(24,this)">24h</button>
      <button class="mud-filter mud-filter-top" onclick="setMudFiltro(12,this)">12h</button>
      <button class="mud-filter mud-filter-top" onclick="setMudFiltro(6,this)">6h</button>
      <button class="mud-filter mud-filter-top" onclick="setMudFiltro(1,this)">1h</button>
    </div>
    <div class="mud-list" id="mud-list"></div>
  </div>
</div>
<div style="padding:0 32px 0;margin-top:6px">
  <div class="panel" style="max-height:none">
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:12px;flex-wrap:wrap;gap:8px">
      <div style="font-size:9px;text-transform:uppercase;letter-spacing:.12em;color:var(--muted2)">📋 Histórico completo de mudanças <span id="mud-full-count" style="color:var(--text)"></span></div>
      <div style="display:flex;gap:6px;align-items:center">
        <button class="mud-filter mud-filter-full active" onclick="setMudFiltroFull(24,this)">24h</button>
        <button class="mud-filter mud-filter-full" onclick="setMudFiltroFull(12,this)">12h</button>
        <button class="mud-filter mud-filter-full" onclick="setMudFiltroFull(6,this)">6h</button>
        <button class="mud-filter mud-filter-full" onclick="setMudFiltroFull(1,this)">1h</button>
        <button onclick="exportarMudancas()" style="padding:4px 12px;background:var(--accent);border:none;border-radius:6px;color:#fff;font-size:10px;cursor:pointer;font-family:inherit;margin-left:4px">⬇ Exportar</button>
      </div>
    </div>
    <div style="overflow-x:auto">
      <table style="width:100%;border-collapse:collapse;font-size:11px" id="mud-full-table">
        <thead><tr style="border-bottom:1px solid var(--border)">
          <th style="padding:7px 10px;text-align:left;font-size:9px;text-transform:uppercase;letter-spacing:.1em;color:var(--muted2);white-space:nowrap">Horário</th>
          <th style="padding:7px 10px;text-align:left;font-size:9px;text-transform:uppercase;letter-spacing:.1em;color:var(--muted2)">AWB</th>
          <th style="padding:7px 10px;text-align:left;font-size:9px;text-transform:uppercase;letter-spacing:.1em;color:var(--muted2)">Pedido</th>
          <th style="padding:7px 10px;text-align:left;font-size:9px;text-transform:uppercase;letter-spacing:.1em;color:var(--muted2)">De</th>
          <th style="padding:7px 10px;text-align:left;font-size:9px;text-transform:uppercase;letter-spacing:.1em;color:var(--muted2)"></th>
          <th style="padding:7px 10px;text-align:left;font-size:9px;text-transform:uppercase;letter-spacing:.1em;color:var(--muted2)">Para</th>
        </tr></thead>
        <tbody id="mud-full-body"></tbody>
      </table>
      <div id="mud-full-empty" style="display:none;text-align:center;padding:24px;color:var(--muted2);font-size:12px">Nenhuma mudança no período selecionado</div>
    </div>
  </div>
</div>
<div class="search-bar">
  <input class="search-input" type="text" id="search" placeholder="Buscar AWB, pedido ou produto..." oninput="renderTable()">
  <button class="filter-tag active" onclick="filterCat('ALL')" data-ftag="ALL">Todos</button>
{filter_tags}</div>
<div class="table-wrap">
<table id="main-table"><thead><tr>
  <th class="sortable" data-col="awb">AWB</th>
  <th class="sortable" data-col="pedido">PEDIDO</th>
  <th class="sortable" data-col="cat">CATEGORIA</th>
  <th class="sortable" data-col="status">STATUS FEDEX</th>
  <th>ÚLTIMO EVENTO</th>
  <th class="sortable" data-col="dias">DIAS NO STATUS</th>
  <th class="sortable" data-col="eta">ETA</th>
  <th class="sortable" data-col="data">DATA</th>
</tr></thead>
<tbody id="tbody"></tbody></table>
<div style="padding:12px 0;font-size:11px;color:var(--muted);font-family:'DM Mono',monospace" id="row-count"></div>
</div>
</div>
<div id="tab-mensal" class="tab-pane">
  <div class="mensal-wrap" id="mensal-content"></div>
</div>
<div id="tab-periodo" class="tab-pane">
  <div style="padding:24px 32px">
    <div style="background:var(--surface);border:1px solid var(--border);border-radius:10px;padding:20px;margin-bottom:20px;max-width:600px">
      <div style="font-size:10px;text-transform:uppercase;letter-spacing:.12em;color:var(--muted2);margin-bottom:14px">Selecione o período</div>
      <div style="display:grid;grid-template-columns:1fr 1fr auto;gap:10px;align-items:end">
        <div><div style="font-size:10px;color:var(--muted2);margin-bottom:5px">Data início</div><input type="date" id="periodo-inicio" style="background:var(--surface2);border:1px solid var(--border);border-radius:6px;padding:8px 12px;color:var(--text);font-family:'DM Mono',monospace;font-size:12px;width:100%;outline:none"></div>
        <div><div style="font-size:10px;color:var(--muted2);margin-bottom:5px">Data fim</div><input type="date" id="periodo-fim" style="background:var(--surface2);border:1px solid var(--border);border-radius:6px;padding:8px 12px;color:var(--text);font-family:'DM Mono',monospace;font-size:12px;width:100%;outline:none"></div>
        <button onclick="gerarRelatorioPeriodo()" style="background:var(--accent);border:none;border-radius:6px;padding:9px 18px;color:#fff;font-size:12px;cursor:pointer;font-family:'DM Sans',sans-serif;white-space:nowrap">Gerar relatório</button>
      </div>
      <div style="display:flex;gap:8px;margin-top:10px;flex-wrap:wrap">
        <button onclick="atalhosPeriodo(7)"  style="padding:4px 10px;background:var(--surface2);border:1px solid var(--border);border-radius:4px;color:var(--muted2);font-size:10px;cursor:pointer;font-family:inherit">Últimos 7 dias</button>
        <button onclick="atalhosPeriodo(15)" style="padding:4px 10px;background:var(--surface2);border:1px solid var(--border);border-radius:4px;color:var(--muted2);font-size:10px;cursor:pointer;font-family:inherit">Últimos 15 dias</button>
        <button onclick="atalhosPeriodo(30)" style="padding:4px 10px;background:var(--surface2);border:1px solid var(--border);border-radius:4px;color:var(--muted2);font-size:10px;cursor:pointer;font-family:inherit">Últimos 30 dias</button>
        <button onclick="atalhosPeriodo(90)" style="padding:4px 10px;background:var(--surface2);border:1px solid var(--border);border-radius:4px;color:var(--muted2);font-size:10px;cursor:pointer;font-family:inherit">Últimos 90 dias</button>
      </div>
    </div>
    <div id="periodo-status" style="font-size:12px;color:var(--muted2)"></div>
  </div>
</div>
<script>
const ROWS=__ROWS_JSON__;
const MUDANCAS=__MUD_JSON__;
const CAT_COLORS={{"COMING TO BRAZIL":{{bg:"var(--coming)",fg:"var(--coming-fg)"}},"CUSTOMS INSPECTION":{{bg:"var(--customs)",fg:"var(--customs-fg)"}},"NATIONAL TRANSIT":{{bg:"var(--national)",fg:"var(--national-fg)"}},"OUT FOR DELIVERY":{{bg:"var(--ofd)",fg:"var(--ofd-fg)"}},"DELIVERED":{{bg:"var(--delivered)",fg:"var(--delivered-fg)"}},"LABEL CREATED":{{bg:"var(--label)",fg:"var(--label-fg)"}}}};
const ORDEM=["COMING TO BRAZIL","CUSTOMS INSPECTION","NATIONAL TRANSIT","OUT FOR DELIVERY","DELIVERED","LABEL CREATED"];
const ICONS={{"COMING TO BRAZIL":"✈","CUSTOMS INSPECTION":"🔍","NATIONAL TRANSIT":"🚚","OUT FOR DELIVERY":"📦","DELIVERED":"✅","LABEL CREATED":"🏷"}};
const CAT_BAR_COLOR={{"COMING TO BRAZIL":"#4f7dff","CUSTOMS INSPECTION":"#f59e0b","NATIONAL TRANSIT":"#10b981","OUT FOR DELIVERY":"#818cf8","DELIVERED":"#22c55e","LABEL CREATED":"#475569"}};
let sortCol=null,sortDir=1,currentCat="ALL",mudFiltroHoras=24,mudFiltroFullHoras=24,mensalIdx=0;

function heatClass(d,cat){{if(d===null||d===undefined||cat==="DELIVERED")return'heat-0';if(d<=2)return'heat-1';if(d<=5)return'heat-2';if(d<=10)return'heat-3';return'heat-4';}}
function sortBy(col){{if(sortCol===col)sortDir*=-1;else{{sortCol=col;sortDir=1;}}document.querySelectorAll('thead th').forEach(th=>th.classList.remove('sort-asc','sort-desc'));document.querySelectorAll('thead th[data-col="'+col+'"]').forEach(th=>th.classList.add(sortDir===1?'sort-asc':'sort-desc'));renderTable();}}
document.querySelectorAll('thead th.sortable').forEach(th=>th.addEventListener('click',()=>sortBy(th.dataset.col)));
function _fmtData(d){{if(!d)return"";try{{const[y,m,dy]=d.split("-");const meses=["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"];const ds=["Dom","Seg","Ter","Qua","Qui","Sex","Sáb"];const dt=new Date(+y,+m-1,+dy);return ds[dt.getDay()]+", "+meses[+m-1]+" "+dy+"/"+y.slice(2);}}catch(e){{return d;}}}}
function catBadge(cat){{const c=CAT_COLORS[cat]||{{bg:"#333",fg:"#aaa"}};return`<span class="cat-badge" style="background:${{c.bg}};color:${{c.fg}}">${{ICONS[cat]||""}} ${{cat}}</span>`;}}
function openModal(awb){{
  const r=ROWS.find(x=>x.awb===awb);if(!r)return;
  const c=CAT_COLORS[r.cat]||{{bg:"#333",fg:"#aaa"}};
  const diasColor=r.dias>10?"#f87171":r.dias>5?"#fb923c":r.dias>2?"#fcd34d":"#6ee7b7";
  document.getElementById("modal-header").innerHTML=`<div><div class="modal-awb">${{r.awb}}</div><div class="modal-sub">${{r.produto||""}}${{r.pedido?" · Pedido "+r.pedido:""}}</div></div><div style="display:flex;align-items:center;gap:10px"><span class="cat-badge" style="background:${{c.bg}};color:${{c.fg}};font-size:11px;padding:5px 12px">${{ICONS[r.cat]||""}} ${{r.cat}}</span><button class="modal-close" onclick="closeModal()">✕ Fechar</button></div>`;
  document.getElementById("modal-meta").innerHTML=`<div class="modal-meta-item"><div class="modal-meta-label">Status atual</div><div class="modal-meta-val">${{r.status}}</div></div><div class="modal-meta-item"><div class="modal-meta-label">Destino</div><div class="modal-meta-val">${{r.endereco||"—"}}</div></div><div class="modal-meta-item"><div class="modal-meta-label">Dias no status</div><div class="modal-meta-val" style="color:${{diasColor}}">${{r.dias!=null?r.dias+"d úteis":"—"}}</div></div><div class="modal-meta-item"><div class="modal-meta-label">ETA estimado</div><div class="modal-meta-val" style="color:var(--accent)">${{r.eta||"—"}}</div></div><div class="modal-meta-item"><div class="modal-meta-label">Última atualização</div><div class="modal-meta-val">${{(r.evento.split("—").pop()||"").trim().split(" ").slice(0,2).join(" ")||"—"}}</div></div>`;
  const byDate={{}};(r.eventos||[]).forEach(ev=>{{if(!byDate[ev.data])byDate[ev.data]=[];byDate[ev.data].push(ev);}});
  const dates=Object.keys(byDate).sort((a,b)=>b.localeCompare(a));let evHtml="";
  dates.forEach(d=>{{evHtml+=`<div class="modal-day-group"><div class="modal-day-header">${{_fmtData(d)}}</div>`;byDate[d].slice().reverse().forEach(ev=>{{evHtml+=`<div class="modal-event-row"><div class="modal-event-time">${{ev.hora||""}}</div><div><div class="modal-event-desc">${{ev.desc}}</div>${{ev.subdesc?`<div class="modal-event-subdesc">${{ev.subdesc}}</div>`:""}}</div><div class="modal-event-local">${{ev.local||""}}</div></div>`;}});evHtml+="</div>";}});
  document.getElementById("modal-body").innerHTML=evHtml||'<div style="padding:32px;text-align:center;color:var(--muted2);font-size:12px">Eventos não disponíveis — aguarde o próximo ciclo de consulta</div>';
  document.getElementById("modal-overlay").classList.add("open");document.body.style.overflow="hidden";
  const intel=document.getElementById("modal-intel");const badges=[];
  if(r.fp&&r.fp.status){{const cls={{"normal":"intel-normal","atencao":"intel-atencao","critico":"intel-critico","rapido":"intel-rapido"}}[r.fp.status]||"intel-normal";const ic={{"normal":"✓","atencao":"⚠","critico":"🔴","rapido":"↑"}}[r.fp.status]||"";const hist=r.fp.pct_historico!=null?` · Top ${{r.fp.pct_historico}}% hist.`:"";badges.push(`<span class="intel-badge ${{cls}}" title="${{r.fp.mensagem}}">${{ic}} Produto: ${{r.fp.mensagem.split("—")[0].trim()}}${{hist}}</span>`);}}
  if(r.risco&&r.risco.probabilidade!=null){{const cls=r.risco.probabilidade>=60?"intel-critico":r.risco.probabilidade>=40?"intel-risco":"intel-atencao";badges.push(`<span class="intel-badge ${{cls}}" title="${{r.risco.mensagem||""}}">⚡ Risco atraso: ${{r.risco.probabilidade}}% (${{r.risco.casos_similares}} casos similares)</span>`);}}
  if(r.previsao&&r.previsao.dias_restantes!=null){{badges.push(`<span class="intel-badge intel-previsao" title="Confiança: ${{r.previsao.confianca||0}}%">🔮 Previsão: ~${{r.previsao.dias_restantes}}d úteis${{r.previsao.ajuste_calendario?" "+r.previsao.ajuste_calendario:""}}</span>`);}}
  if(r.desc_cal){{badges.push(`<span class="intel-badge intel-cal">📅 ${{r.desc_cal}}</span>`);}}
  if(badges.length>0){{intel.innerHTML=badges.join("");intel.style.display="flex";}}else{{intel.style.display="none";}}
}}
function closeModal(e){{if(e&&e.target!==document.getElementById("modal-overlay")&&!e.target.classList.contains("modal-close"))return;document.getElementById("modal-overlay").classList.remove("open");document.body.style.overflow="";}}
document.addEventListener("keydown",e=>{{if(e.key==="Escape")closeModal({{target:document.getElementById("modal-overlay")}});}});
function renderTable(){{
  const q=document.getElementById("search").value.trim().toLowerCase();
  let rows=ROWS.filter(r=>{{const cm=currentCat==="ALL"||r.cat===currentCat;const qm=!q||r.awb.includes(q)||r.pedido.toLowerCase().includes(q)||(r.produto||"").toLowerCase().includes(q);return cm&&qm;}});
  if(sortCol){{rows=[...rows].sort((a,b)=>{{let va=a[sortCol]??'',vb=b[sortCol]??'';if(sortCol==='dias'){{va=va??9999;vb=vb??9999;return(va-vb)*sortDir;}}return String(va).localeCompare(String(vb))*sortDir;}});}}
  document.getElementById("tbody").innerHTML=rows.map(r=>{{const hc=heatClass(r.dias,r.cat);const bar=CAT_BAR_COLOR[r.cat]||'transparent';return`<tr class="${{r.alerta?'alerta':''}}" style="border-left:3px solid ${{bar}}" onclick="openModal('${{r.awb}}')"><td style="font-family:'DM Mono',monospace;font-size:10px">${{r.awb}}</td><td><span class="pedido-badge">${{r.pedido}}</span></td><td>${{catBadge(r.cat)}}</td><td style="font-size:10px;color:var(--muted2)">${{r.status}}</td><td class="evento-cell">${{r.evento}}</td><td class="dias-cell ${{hc}}">${{r.dias!=null?r.dias+'d':'—'}}</td><td class="eta-cell">${{r.eta||'—'}}</td><td class="td-last" style="font-family:'DM Mono',monospace;font-size:10px;color:var(--muted)">${{r.data}}</td></tr>`;}}).join("");
  document.getElementById("row-count").textContent=`${{rows.length}} de ${{ROWS.length}} remessas`;
}}
function renderChart(){{
  const total=ROWS.length;const cats={{}};ROWS.forEach(r=>cats[r.cat]=(cats[r.cat]||0)+1);const max=Math.max(...Object.values(cats),1);
  document.getElementById("total-badge").textContent=total+" remessas";
  document.getElementById("chart-bars").innerHTML=ORDEM.map(cat=>{{const n=cats[cat]||0;const pct=Math.round(n/total*100);const w=Math.round(n/max*100);const c=CAT_COLORS[cat]||{{fg:"#aaa"}};const short=cat.replace("COMING TO BRAZIL","COMING").replace("CUSTOMS INSPECTION","CUSTOMS").replace("NATIONAL TRANSIT","NATIONAL").replace("OUT FOR DELIVERY","OUT FOR DEL.").replace("LABEL CREATED","LABEL");return`<div class="bar-row"><div class="bar-name">${{short}}</div><div class="bar-track"><div class="bar-fill" style="width:${{w}}%;background:${{c.fg}};opacity:.7"></div></div><div class="bar-pct">${{pct}}%</div></div>`;}}).join("");
}}
function _filtrarMudancas(horas){{const agora=new Date();return MUDANCAS.filter(r=>{{if(!r[3])return true;try{{const[data,hora]=r[3].split(" ");if(!data||!hora)return true;const[d,m,y]=data.split("/");const[h,mi]=hora.split(":");const dt=new Date(+y,+m-1,+d,+h,+mi);return(agora-dt)/(1000*60*60)<=horas;}}catch(e){{return true;}}}});}}
function setMudFiltro(horas,btn){{mudFiltroHoras=horas;document.querySelectorAll('.mud-filter-top').forEach(b=>b.classList.remove('active'));if(btn)btn.classList.add('active');renderMudancas();}}
function setMudFiltroFull(horas,btn){{mudFiltroFullHoras=horas;document.querySelectorAll('.mud-filter-full').forEach(b=>b.classList.remove('active'));if(btn)btn.classList.add('active');renderMudancasFull();}}
function renderMudancas(){{
  const el=document.getElementById("mud-list");const filtradas=_filtrarMudancas(mudFiltroHoras);
  document.getElementById("mud-count").textContent=filtradas.length?`${{filtradas.length}} nas últimas ${{mudFiltroHoras}}h`:"";
  if(!filtradas.length){{el.innerHTML=`<div class="mud-empty">Nenhuma mudança nas últimas ${{mudFiltroHoras}}h</div>`;return;}}
  el.innerHTML=filtradas.map(r=>{{const trans=(r[2]||"").split("→").map(s=>s.trim());const ant=trans[0]||"";const atu=trans[1]||"";const ca=CAT_COLORS[ant]||{{bg:"#333",fg:"#aaa"}};const cb=CAT_COLORS[atu]||{{bg:"#333",fg:"#aaa"}};const ts=r[3]?`<div style="font-family:'DM Mono',monospace;font-size:9px;color:var(--muted);margin-top:3px">${{r[3].split(" ")[1]||""}}</div>`:"";return`<div class="mud-row" onclick="openModal('${{r[0]}}')" style="cursor:pointer"><div><div class="mud-awb">${{r[0]}}</div>${{ts}}</div><div class="mud-pedido">${{r[1]||""}}</div><div class="mud-trans"><span class="mud-cat" style="background:${{ca.bg}};color:${{ca.fg}}">${{ant}}</span><span class="mud-arrow">→</span><span class="mud-cat" style="background:${{cb.bg}};color:${{cb.fg}}">${{atu}}</span></div></div>`;}}).join("");
}}
function renderMudancasFull(){{
  const filtradas=_filtrarMudancas(mudFiltroFullHoras);const tbody=document.getElementById("mud-full-body");const empty=document.getElementById("mud-full-empty");
  document.getElementById("mud-full-count").textContent=filtradas.length?`(${{filtradas.length}} registros)`:"";
  if(!filtradas.length){{tbody.innerHTML="";empty.style.display="block";return;}}empty.style.display="none";
  tbody.innerHTML=filtradas.map(r=>{{const trans=(r[2]||"").split("→").map(s=>s.trim());const ant=trans[0]||"";const atu=trans[1]||"";const ca=CAT_COLORS[ant]||{{bg:"#333",fg:"#aaa"}};const cb=CAT_COLORS[atu]||{{bg:"#333",fg:"#aaa"}};const hora=r[3]?r[3].split(" ").pop():"—";const data=r[3]?r[3].split(" ")[0]:"—";return`<tr style="border-bottom:0.5px solid var(--border);cursor:pointer" onclick="openModal('${{r[0]}}')"><td style="padding:8px 10px;font-family:'DM Mono',monospace;font-size:10px;color:var(--muted2);white-space:nowrap">${{data}}<br><span style="color:var(--accent)">${{hora}}</span></td><td style="padding:8px 10px;font-family:'DM Mono',monospace;font-size:10px">${{r[0]}}</td><td style="padding:8px 10px;font-size:10px;color:var(--accent)">${{r[1]||"—"}}</td><td style="padding:8px 10px"><span class="mud-cat" style="background:${{ca.bg}};color:${{ca.fg}};font-size:9px;padding:2px 7px;border-radius:3px">${{ant}}</span></td><td style="padding:8px 4px;color:var(--muted);font-size:12px">→</td><td style="padding:8px 10px"><span class="mud-cat" style="background:${{cb.bg}};color:${{cb.fg}};font-size:9px;padding:2px 7px;border-radius:3px">${{atu}}</span></td></tr>`;}}).join("");
}}
function exportarMudancas(){{
  const filtradas=_filtrarMudancas(mudFiltroFullHoras);
  if(!filtradas.length){{alert("Nenhuma mudança para exportar no período selecionado.");return;}}
  const linhas=[["Horario","AWB","Pedido","Status Anterior","Status Novo"]];
  filtradas.forEach(r=>{{const trans=(r[2]||"").split("→").map(s=>s.trim());linhas.push([r[3]||"",r[0],r[1]||"",trans[0]||"",trans[1]||""]);  }});
  const csv=linhas.map(l=>l.map(v=>'"'+String(v).replace(/"/g,'""')+'"').join(",")).join("\\n");
  const blob=new Blob(["\uFEFF"+csv],{{type:"text/csv;charset=utf-8"}});const url=URL.createObjectURL(blob);const a=document.createElement("a");a.href=url;a.download=`mudancas_${{mudFiltroFullHoras}}h_${{new Date().toISOString().slice(0,10)}}.csv`;a.click();URL.revokeObjectURL(url);
}}
function filterCat(cat){{currentCat=cat;document.querySelectorAll("[data-ftag]").forEach(b=>b.classList.toggle("active",b.dataset.ftag===cat));renderTable();}}
function switchTab(tab){{document.querySelectorAll('.tab-pane').forEach(p=>p.classList.remove('active'));document.querySelectorAll('.tab-btn').forEach(b=>b.classList.remove('active'));document.getElementById('tab-'+tab).classList.add('active');event.target.classList.add('active');if(tab==='periodo')initPeriodo();}}
function initPeriodo(){{const hoje=new Date();const fim=hoje.toISOString().split('T')[0];const ini=new Date(hoje);ini.setDate(ini.getDate()-7);if(!document.getElementById('periodo-inicio').value){{document.getElementById('periodo-inicio').value=ini.toISOString().split('T')[0];document.getElementById('periodo-fim').value=fim;}}}}
function atalhosPeriodo(dias){{const hoje=new Date();const ini=new Date(hoje);ini.setDate(ini.getDate()-dias);document.getElementById('periodo-inicio').value=ini.toISOString().split('T')[0];document.getElementById('periodo-fim').value=hoje.toISOString().split('T')[0];}}
function gerarRelatorioPeriodo(){{
  const ini=document.getElementById('periodo-inicio').value;const fim=document.getElementById('periodo-fim').value;const st=document.getElementById('periodo-status');
  if(!ini||!fim){{st.textContent='Selecione as datas de início e fim.';return;}}
  if(ini>fim){{st.textContent='A data início deve ser anterior à data fim.';return;}}
  if(window.location.protocol==='file:'){{st.innerHTML='<span style="color:#f87171">Para usar esta função, acesse via servidor: <strong>http://localhost:8888</strong></span>';return;}}
  st.innerHTML='<span style="color:var(--muted2)">Gerando relatório...</span>';
  fetch('/gerar-relatorio?inicio='+ini+'&fim='+fim).then(r=>r.json()).then(d=>{{if(d.ok){{st.innerHTML='Relatório gerado! <a href="'+d.arquivo+'" target="_blank" style="color:var(--accent);font-weight:500">Abrir →</a>';}}else{{st.innerHTML='<span style="color:#f87171">Erro: '+d.erro+'</span>';}}  }}).catch(()=>{{st.innerHTML='<span style="color:#f87171">Falha na conexão com o servidor.</span>';}});
}}
(function(){{const meta=document.querySelector('meta[http-equiv="refresh"]');if(!meta)return;let secs=parseInt(meta.content)||3600;const el=document.getElementById('countdown');function fmt(s){{const m=Math.floor(s/60),r=s%60;return'↻ próxima em '+m+'min '+(r<10?'0':'')+r+'s';}}if(el)el.textContent=fmt(secs);const iv=setInterval(function(){{secs--;if(secs<=0){{clearInterval(iv);if(el)el.textContent='↻ atualizando...';return;}}if(el)el.textContent=fmt(secs);}},1000);}})();
const MENSAL=__MENSAL_JSON__;
const CAT_COR_REG={{"América do Norte":"#BDD7EE","Europa":"#C6EFCE","Oriente Médio":"#FFE699","Ásia":"#F4CCFF","Outros":"#D9D9D9"}};
function _kpiPopover(label,awbObj){{if(!awbObj)return`<div class="mensal-kpi"><div class="mensal-kpi-label">${{label}}</div><div class="mensal-kpi-val">—</div></div>`;const dias=awbObj.lt!==null?awbObj.lt+'d':'—';const tip=`${{awbObj.awb}}${{awbObj.pedido?' · '+awbObj.pedido:''}}${{awbObj.produto?'\\n'+awbObj.produto:''}}`;return`<div class="mensal-kpi mensal-kpi-clickable" onclick="togglePop(this)" title="${{tip}}"><div class="mensal-kpi-label">${{label}} <span style="font-size:9px;opacity:.6">ⓘ</span></div><div class="mensal-kpi-val">${{dias}}</div><div class="mensal-pop"><div class="mensal-pop-awb">${{awbObj.awb}}</div>${{awbObj.pedido?`<div class="mensal-pop-row"><span>Pedido</span><span>${{awbObj.pedido}}</span></div>`:''}}<div class="mensal-pop-row"><span>Lead time</span><span style="font-weight:600">${{dias}} úteis</span></div></div></div>`;}}
function togglePop(el){{const pop=el.querySelector('.mensal-pop');if(!pop)return;const isOpen=pop.classList.contains('open');document.querySelectorAll('.mensal-pop.open').forEach(p=>p.classList.remove('open'));if(!isOpen){{pop.classList.add('open');const rect=el.getBoundingClientRect();const popW=230;let left=rect.left+rect.width/2-popW/2;left=Math.max(8,Math.min(left,window.innerWidth-popW-8));pop.style.top=(rect.bottom+8)+'px';pop.style.left=left+'px';pop.style.width=popW+'px';}}}}
document.addEventListener('click',function(e){{if(!e.target.closest('.mensal-kpi-clickable')){{document.querySelectorAll('.mensal-pop.open').forEach(p=>p.classList.remove('open'));}}  }});
function renderMensal(){{
  const el=document.getElementById('mensal-content');
  if(!MENSAL||!MENSAL.meses||MENSAL.meses.length===0){{el.innerHTML='<div class="mensal-sem-dados">Nenhum dado de entrega disponível ainda.<br>As remessas entregues aparecerão aqui automaticamente.</div>';return;}}
  const meses=MENSAL.meses;const m=meses[mensalIdx];
  const navH=`<div style="display:flex;align-items:center;gap:12px;margin-bottom:20px;flex-wrap:wrap"><button onclick="mudaMes(-1)" style="background:var(--surface);border:1px solid var(--border);color:var(--text);padding:6px 14px;border-radius:6px;cursor:pointer;font-size:13px" ${{mensalIdx===meses.length-1?'disabled':''}}>&larr; Anterior</button><div style="display:flex;gap:6px;flex-wrap:wrap">${{meses.map((mm,i)=>`<button onclick="irMes(${{i}})" style="background:${{i===mensalIdx?'var(--accent)':'var(--surface)'}};border:1px solid ${{i===mensalIdx?'var(--accent)':'var(--border)'}};color:${{i===mensalIdx?'#fff':'var(--text)'}};padding:4px 10px;border-radius:6px;cursor:pointer;font-size:12px">${{mm.periodo}}</button>`).join('')}}</div><button onclick="mudaMes(1)" style="background:var(--surface);border:1px solid var(--border);color:var(--text);padding:6px 14px;border-radius:6px;cursor:pointer;font-size:13px" ${{mensalIdx===0?'disabled':''}}>Próximo &rarr;</button></div>`;
  let h=navH;h+=`<div style="margin-bottom:20px"><span style="font-size:13px;color:var(--muted)">Período: </span><span style="font-size:15px;font-weight:600">${{m.periodo}}</span><span style="margin-left:20px;font-size:13px;color:var(--muted)">Total entregue: </span><span style="font-size:15px;font-weight:600;color:var(--accent)">${{m.total}} envios</span></div>`;
  if(!m.regioes||m.regioes.length===0){{h+='<div class="mensal-sem-dados">Nenhuma remessa entregue neste mês.</div>';el.innerHTML=h;return;}}
  if(meses.length>1){{const maxVol=Math.max(...meses.map(mm=>mm.total),1);const barW=Math.max(14,Math.floor(180/meses.length));let spark='<div class="sparkline-wrap"><div class="sparkline-title">Volume histórico de entregas</div><div style="display:flex;align-items:flex-end;gap:3px;height:40px">';[...meses].reverse().forEach((mm,i)=>{{const h2=Math.max(4,Math.round(mm.total/maxVol*36));const isActive=meses.length-1-i===mensalIdx;spark+=`<div title="${{mm.periodo}}: ${{mm.total}} envios" onclick="irMes(${{meses.length-1-i}})" style="width:${{barW}}px;height:${{h2}}px;background:${{isActive?'var(--accent)':'var(--border2)'}};border-radius:2px 2px 0 0;cursor:pointer;transition:.15s;flex-shrink:0"></div>`;}});spark+='</div></div>';h+=spark;}}
  const REG_COLORS={{"América do Norte":"#4f7dff","Europa":"#10b981","Oriente Médio":"#f59e0b","Ásia":"#a78bfa","Outros":"#475569"}};
  const REG_BG={{"América do Norte":"#0d1a36","Europa":"#062010","Oriente Médio":"#1c1200","Ásia":"#150d2e","Outros":"#111827"}};
  m.regioes.forEach(r=>{{const cor=REG_COLORS[r.nome]||'#4f7dff';const bg=REG_BG[r.nome]||'#111827';h+=`<div class="mensal-regiao"><div class="mensal-reg-header" style="background:${{bg}};border-left-color:${{cor}}">${{r.nome}} — ${{r.total}} envio${{r.total>1?'s':''}}</div><div class="mensal-kpis"><div class="mensal-kpi"><div class="mensal-kpi-label">Total</div><div class="mensal-kpi-val">${{r.total}}</div></div><div class="mensal-kpi"><div class="mensal-kpi-label">Lead time médio</div><div class="mensal-kpi-val">${{r.media_lt!==null&&r.media_lt!=="-"?r.media_lt+"d":"—"}}</div></div>${{_kpiPopover('Mínimo',r.awb_min)}}${{_kpiPopover('Máximo',r.awb_max)}}<div class="mensal-kpi"><div class="mensal-kpi-label">Produto top</div><div class="mensal-kpi-val destaque">${{r.produto_top}}</div></div></div>${{r.awbs&&r.awbs.length>0?`<details style="margin-top:8px"><summary style="cursor:pointer;font-size:12px;color:var(--muted);user-select:none">Ver remessas (${{r.awbs.length}})</summary><div style="margin-top:8px;display:flex;flex-wrap:wrap;gap:6px">${{r.awbs.map(a=>`<span style="font-family:'DM Mono',monospace;font-size:11px;background:var(--bg);padding:2px 8px;border-radius:4px;border:1px solid var(--border)" title="${{a.produto}}">${{a.awb}}${{a.pedido?' · '+a.pedido:''}}${{a.lt!=null?' ('+a.lt+'d)':''}}</span>`).join('')}}</div></details>`:''}}</div>`;}});
  el.innerHTML=h;
}}
function mudaMes(dir){{mensalIdx=Math.max(0,Math.min(MENSAL.meses.length-1,mensalIdx-dir));renderMensal();}}
function irMes(i){{mensalIdx=i;renderMensal();}}
renderChart();renderMudancas();renderMudancasFull();renderTable();renderMensal();
</script></body></html>"""

    html = html.replace("__ROWS_JSON__", json.dumps(rows_json, ensure_ascii=False))
    html = html.replace("__MUD_JSON__", json.dumps(mudancas_rows, ensure_ascii=False))
    html = html.replace("__MENSAL_JSON__", json.dumps(mensal_json, ensure_ascii=False))

    with open(arquivo_html, "w", encoding="utf-8") as f:
        f.write(html)
    logger.info(f"HTML salvo → '{arquivo_html}'")


# ==========================================================
# RELATÓRIO POR PERÍODO
# ==========================================================

def gerar_relatorio_periodo(arquivo_historico: str, data_inicio: str, data_fim: str, arquivo_saida: str) -> bool:
    logger = logging.getLogger("fedex_tracker")
    try:
        from datetime import date as _d
        d1 = datetime.strptime(data_inicio, "%Y-%m-%d").date()
        d2 = datetime.strptime(data_fim,    "%Y-%m-%d").date()
        path = Path(arquivo_historico)
        if not path.exists(): return False
        df = pd.read_excel(path); df.columns = df.columns.str.strip()
        df["_dc"] = pd.to_datetime(df.get("DATA_CONSULTA", pd.Series()), dayfirst=True, errors="coerce")
        df = df[df["_dc"].dt.date.between(d1, d2)].copy()
        if df.empty: return False
        def _nc(c):
            c=re.sub(r"^⚠\s*","",str(c)); c=re.sub(r"\s*\(\d+d\)\s*$","",c); c=re.sub(r"\s*MEMPHIS.*$","",c); return c.strip()
        df["CAT_NORM"] = df["CATEGORIA"].apply(_nc)
        movimentacoes = []
        if "AWB" in df.columns:
            for awb, grp in df.sort_values("_dc").groupby("AWB"):
                cats=grp["CAT_NORM"].tolist(); datas=grp["_dc"].tolist()
                pedido=str(grp["PEDIDO"].iloc[0]).replace(".0","") if "PEDIDO" in grp.columns else ""
                produto=str(grp["PRODUTO"].iloc[0]) if "PRODUTO" in grp.columns else ""
                prev=None
                for cat,dt in zip(cats,datas):
                    if cat!=prev and cat and cat not in ("nan","ERRO"):
                        movimentacoes.append({"awb":str(awb),"pedido":pedido,"produto":produto,"cat":cat,"cat_ant":prev or "—","data":dt.strftime("%d/%m/%Y") if pd.notna(dt) else "","hora":dt.strftime("%H:%M") if pd.notna(dt) else ""})
                        prev=cat
        total_awbs=df["AWB"].nunique() if "AWB" in df.columns else 0
        total_entregas=len(df[df["CAT_NORM"]=="DELIVERED"]["AWB"].unique()) if "AWB" in df.columns else 0
        total_mud=len(movimentacoes); lt_media="—"
        if "LEAD_TIME_DIAS" in df.columns:
            lts=pd.to_numeric(df["LEAD_TIME_DIAS"],errors="coerce").dropna()
            if len(lts)>0: lt_media=f"{round(float(lts.mean()),1)}d úteis"
        df["_data"]=df["_dc"].dt.date; resumo_diario=df.groupby(["_data","CAT_NORM"]).size().reset_index(name="n")
        CAT_CORES_HTML = {"COMING TO BRAZIL":("#1a3a6c","#93c5fd"),"CUSTOMS INSPECTION":("#3d2e00","#fcd34d"),"NATIONAL TRANSIT":("#0f3320","#6ee7b7"),"OUT FOR DELIVERY":("#1e2f5c","#a5b4fc"),"DELIVERED":("#052e16","#4ade80"),"LABEL CREATED":("#1a1a2e","#94a3b8")}
        def badge_html(cat,ant=False):
            bg,fg=CAT_CORES_HTML.get(str(cat).strip(),("#333","#aaa")); opacity="opacity:.6;" if ant else ""
            return f'<span style="background:{bg};color:{fg};{opacity}padding:2px 8px;border-radius:4px;font-size:10px;font-weight:600">{cat}</span>'
        rows_mov="".join(f"<tr><td style='padding:8px 12px;font-family:monospace;font-size:11px'>{m['awb']}</td><td style='padding:8px 12px;font-size:11px;color:#4f7dff'>{m['pedido']}</td><td style='padding:8px 12px;font-size:11px;color:#888'>{m['produto'][:30] if m['produto'] else '—'}</td><td style='padding:8px 12px'>{badge_html(m['cat_ant'],ant=True)}</td><td style='padding:8px 12px;color:#888'>→</td><td style='padding:8px 12px'>{badge_html(m['cat'])}</td><td style='padding:8px 12px;font-family:monospace;font-size:10px;color:#888'>{m['data']} {m['hora']}</td></tr>" for m in movimentacoes) or "<tr><td colspan='7' style='padding:16px;text-align:center;color:#888'>Nenhuma movimentação no período</td></tr>"
        rows_diario=""
        for d in sorted(resumo_diario["_data"].unique(),reverse=True):
            dia_df=resumo_diario[resumo_diario["_data"]==d]; cats_str=" ".join(badge_html(r["CAT_NORM"])+f' <span style="color:#888;font-size:10px">×{r["n"]}</span>' for _,r in dia_df.iterrows()); total_dia=dia_df["n"].sum()
            rows_diario+=f"<tr><td style='padding:8px 12px;font-family:monospace;font-size:11px'>{d.strftime('%d/%m/%Y')}</td><td style='padding:8px 12px;font-size:11px;text-align:center'>{total_dia}</td><td style='padding:8px 12px'>{cats_str}</td></tr>"
        html = f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
<title>Relatório de Movimentações — {d1.strftime('%d/%m/%Y')} a {d2.strftime('%d/%m/%Y')}</title>
<style>@media print{{.no-print{{display:none}}}}*{{box-sizing:border-box;margin:0;padding:0;font-family:Arial,sans-serif}}body{{background:#f5f5f5;color:#222}}.page{{background:#fff;max-width:1000px;margin:20px auto;padding:36px;border-radius:8px}}.header{{border-bottom:3px solid #4f7dff;padding-bottom:16px;margin-bottom:24px;display:flex;justify-content:space-between;align-items:flex-end}}.title{{font-size:20px;font-weight:700;color:#1a1a2e}}.sub{{font-size:12px;color:#888;margin-top:3px}}.kpis{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:28px}}.kpi{{background:#f8f9fc;border-radius:8px;padding:14px;border-left:4px solid #4f7dff}}.kpi-label{{font-size:9px;text-transform:uppercase;letter-spacing:.1em;color:#888;margin-bottom:5px}}.kpi-val{{font-size:24px;font-weight:700;color:#1a1a2e}}.section{{margin-bottom:28px}}.section-title{{font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:.08em;color:#4f7dff;border-bottom:1px solid #eee;padding-bottom:6px;margin-bottom:14px}}table{{width:100%;border-collapse:collapse;font-size:12px}}thead tr{{background:#f0f2ff}}th{{padding:8px 12px;text-align:left;font-size:9px;text-transform:uppercase;letter-spacing:.08em;color:#4f7dff}}tbody tr:hover{{background:#f8f9fc}}tbody tr td{{border-bottom:1px solid #f0f0f0}}.print-btn{{padding:9px 20px;background:#4f7dff;color:#fff;border:none;border-radius:6px;font-size:12px;cursor:pointer;margin-bottom:16px}}</style>
</head><body>
<div class="no-print" style="text-align:center;padding:12px;background:#4f7dff"><button class="print-btn" onclick="window.print()">🖨️ Imprimir / Salvar como PDF</button></div>
<div class="page">
  <div class="header"><div><div class="title">📦 Relatório de Movimentações</div><div class="sub">Período: {d1.strftime('%d/%m/%Y')} a {d2.strftime('%d/%m/%Y')} · Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}</div></div></div>
  <div class="kpis"><div class="kpi"><div class="kpi-label">AWBs no período</div><div class="kpi-val">{total_awbs}</div></div><div class="kpi" style="border-left-color:#22c55e"><div class="kpi-label">Entregas</div><div class="kpi-val" style="color:#16a34a">{total_entregas}</div></div><div class="kpi" style="border-left-color:#818cf8"><div class="kpi-label">Movimentações</div><div class="kpi-val">{total_mud}</div></div><div class="kpi" style="border-left-color:#f59e0b"><div class="kpi-label">Lead time médio</div><div class="kpi-val" style="font-size:16px">{lt_media}</div></div></div>
  <div class="section"><div class="section-title">Resumo diário</div><table><thead><tr><th>Data</th><th>Total</th><th>Distribuição</th></tr></thead><tbody>{rows_diario}</tbody></table></div>
  <div class="section"><div class="section-title">Todas as movimentações ({total_mud})</div><table><thead><tr><th>AWB</th><th>Pedido</th><th>Produto</th><th>Status anterior</th><th></th><th>Status novo</th><th>Data/Hora</th></tr></thead><tbody>{rows_mov}</tbody></table></div>
</div></body></html>"""
        with open(arquivo_saida,"w",encoding="utf-8") as f: f.write(html)
        logger.info(f"Relatório do período salvo → '{arquivo_saida}'"); return True
    except Exception as e:
        logger.error(f"Erro ao gerar relatório por período: {e}"); return False


# ==========================================================
# MAIN
# ==========================================================

# ==========================================================
# MÓDULO DE CONFIGURAÇÃO — interface web para setup inicial
# ==========================================================

CONFIG_FILE = "config.json"

def carregar_config() -> dict:
    """Lê config.json — retorna dict com credenciais e preferências."""
    path = Path(CONFIG_FILE)
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def salvar_config(dados: dict) -> None:
    """Salva config.json com as configurações do cliente."""
    Path(CONFIG_FILE).write_text(json.dumps(dados, indent=2, ensure_ascii=False), encoding="utf-8")

def config_completa() -> bool:
    """Retorna True se as credenciais estão configuradas."""
    c = carregar_config()
    return bool(c.get("client_id") and c.get("client_secret"))

HTML_CONFIG = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>FedEx Tracker — Configuração</title>
<link href="https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600&display=swap" rel="stylesheet">
<style>
*{box-sizing:border-box;margin:0;padding:0}
:root{--bg:#0d1117;--surface:#161b2a;--surface2:#1c2238;--border:#21293f;--text:#dde1f0;--muted:#6b7599;--accent:#4f7dff}
body{background:var(--bg);color:var(--text);font-family:'DM Sans',sans-serif;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:24px}
.card{background:var(--surface);border:1px solid var(--border);border-radius:16px;padding:40px;width:100%;max-width:520px}
.logo{font-family:'DM Mono',monospace;font-size:12px;color:var(--accent);letter-spacing:.1em;margin-bottom:32px;display:flex;align-items:center;gap:8px}
.logo-dot{width:7px;height:7px;background:var(--accent);border-radius:50%}
h1{font-size:20px;font-weight:600;margin-bottom:6px}
.sub{font-size:13px;color:var(--muted);margin-bottom:32px;line-height:1.6}
.section{margin-bottom:28px}
.section-title{font-size:10px;text-transform:uppercase;letter-spacing:.12em;color:var(--muted);margin-bottom:14px;padding-bottom:8px;border-bottom:1px solid var(--border)}
.field{margin-bottom:16px}
label{display:block;font-size:11px;color:var(--muted);margin-bottom:6px;text-transform:uppercase;letter-spacing:.08em}
input,select{width:100%;background:var(--surface2);border:1px solid var(--border);border-radius:8px;padding:10px 14px;color:var(--text);font-family:'DM Sans',sans-serif;font-size:13px;outline:none;transition:border-color .15s}
input:focus,select:focus{border-color:var(--accent)}
input::placeholder{color:var(--muted)}
.hint{font-size:11px;color:var(--muted);margin-top:5px;line-height:1.5}
.hint a{color:var(--accent);text-decoration:none}
.btn{width:100%;background:var(--accent);border:none;border-radius:8px;padding:13px;color:#fff;font-family:'DM Sans',sans-serif;font-size:14px;font-weight:500;cursor:pointer;transition:background .15s;margin-top:8px}
.btn:hover{background:#3d6be8}
.btn:disabled{background:#2a3255;cursor:not-allowed;color:var(--muted)}
.alert{padding:12px 16px;border-radius:8px;font-size:12px;margin-bottom:20px;display:none}
.alert.error{background:#3d0f0f;border:1px solid #7a1f1f;color:#fca5a5;display:block}
.alert.success{background:#0f3320;border:1px solid #185a30;color:#6ee7b7;display:block}
.status-ok{display:flex;align-items:center;gap:8px;background:#0f3320;border:1px solid #185a30;border-radius:8px;padding:12px 16px;margin-bottom:20px}
.status-dot{width:8px;height:8px;background:#4ade80;border-radius:50%;flex-shrink:0}
.status-ok span{font-size:12px;color:#6ee7b7}
</style>
</head>
<body>
<div class="card">
  <div class="logo"><div class="logo-dot"></div>FEDEX TRACKER</div>
  <h1>Configuração inicial</h1>
  <p class="sub">Insira suas credenciais da API FedEx para começar a rastrear suas remessas.</p>

  __STATUS__

  <div id="alert" class="alert"></div>

  <form onsubmit="salvar(event)">
    <div class="section">
      <div class="section-title">Credenciais FedEx API</div>
      <div class="field">
        <label>Client ID</label>
        <input type="text" id="client_id" placeholder="l7xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx" value="__CLIENT_ID__" required autocomplete="off">
      </div>
      <div class="field">
        <label>Client Secret</label>
        <input type="password" id="client_secret" placeholder="••••••••••••••••••••••••••••••••" value="__CLIENT_SECRET__" required autocomplete="off">
        <div class="hint">Obtenha em <a href="https://developer.fedex.com" target="_blank">developer.fedex.com</a> → My Projects → API Key</div>
      </div>
    </div>

    <div class="section">
      <div class="section-title">Preferências</div>
      <div class="field">
        <label>Intervalo de atualização automática</label>
        <select id="intervalo">
          <option value="3600" __SEL_1H__>A cada 1 hora</option>
          <option value="7200" __SEL_2H__>A cada 2 horas</option>
          <option value="14400" __SEL_4H__>A cada 4 horas</option>
          <option value="28800" __SEL_8H__>A cada 8 horas</option>
        </select>
      </div>
      <div class="field">
        <label>Porta do servidor web</label>
        <input type="number" id="porta" value="__PORTA__" min="1024" max="65535">
        <div class="hint">Padrão: 8888. Acesso em http://localhost:8888</div>
      </div>
    </div>

    <button type="submit" class="btn" id="btn-salvar">Salvar e iniciar</button>
  </form>
</div>
<script>
async function salvar(e) {
  e.preventDefault();
  const btn = document.getElementById('btn-salvar');
  const alert = document.getElementById('alert');
  btn.disabled = true; btn.textContent = 'Verificando credenciais...';
  alert.className = 'alert'; alert.textContent = '';
  try {
    const resp = await fetch('/salvar-config', {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({
        client_id: document.getElementById('client_id').value.trim(),
        client_secret: document.getElementById('client_secret').value.trim(),
        intervalo: parseInt(document.getElementById('intervalo').value),
        porta: parseInt(document.getElementById('porta').value) || 8888,
      })
    });
    const d = await resp.json();
    if (d.ok) {
      alert.className = 'alert success'; alert.textContent = '✓ Configurado! Iniciando rastreamento...';
      setTimeout(() => window.location.href = '/', 2000);
    } else {
      alert.className = 'alert error'; alert.textContent = '✗ ' + (d.erro || 'Erro desconhecido');
      btn.disabled = false; btn.textContent = 'Salvar e iniciar';
    }
  } catch(err) {
    alert.className = 'alert error'; alert.textContent = '✗ Erro de conexão: ' + err.message;
    btn.disabled = false; btn.textContent = 'Salvar e iniciar';
  }
}
</script>
</body></html>"""


def executar():
    # Carrega credenciais do config.json se existir
    cfg_saved = carregar_config()
    if cfg_saved.get("client_id"):
        import os
        os.environ["FEDEX_CLIENT_ID"]     = cfg_saved["client_id"]
        os.environ["FEDEX_CLIENT_SECRET"] = cfg_saved.get("client_secret","")

    config = Config()
    logger = setup_logger(config.arquivo_log)
    logger.info("=" * 60)
    logger.info("  FedEx Tracker — API Oficial")
    logger.info("=" * 60)

    # Se não configurado, aguarda — o usuário vai configurar pelo browser
    if not config_completa():
        logger.info("Credenciais não configuradas.")
        logger.info(f"Acesse http://localhost:{config.porta_servidor}/config para configurar.")
        return  # não executa até ter credenciais

    try: config.validar()
    except ValueError as e: logger.error(str(e)); sys.exit(1)
    try: lista_awbs = carregar_awbs(config.arquivo_awbs)
    except (FileNotFoundError, ValueError) as e: logger.error(str(e)); sys.exit(1)
    if not lista_awbs: logger.error("Nenhum AWB válido encontrado."); sys.exit(1)
    try: token_mgr = TokenManager(config); token = token_mgr.get_token()
    except Exception as e: logger.error(f"Falha na autenticação: {e}"); sys.exit(1)
    logger.info(f"Iniciando rastreio de {len(lista_awbs)} AWBs...")
    resultados = []
    with ThreadPoolExecutor(max_workers=config.max_workers) as executor:
        futures = {executor.submit(consultar_awb, orig, limpo, token, config, ped, prod): limpo for orig, limpo, ped, prod in lista_awbs}
        with tqdm(total=len(lista_awbs), desc="Rastreando", unit="AWB", ncols=72) as pbar:
            for future in as_completed(futures):
                r = future.result(); resultados.append(r)
                pbar.set_postfix_str(f"{r['AWB']} → {r['CATEGORIA'][:18]}"); pbar.update(1)
    try: salvar_resultados(resultados, config.arquivo_historico, config.arquivo_ultimo_status)
    except Exception as e: logger.error(f"Erro ao salvar: {e}")
    try:
        arquivo_html = config.arquivo_ultimo_status.replace(".xlsx", ".html")
        gerar_html_relatorio(config.arquivo_ultimo_status, arquivo_html, config.arquivo_historico)
    except Exception as e:
        import traceback; logger.error(f"Erro ao gerar HTML: {e}"); logger.error(traceback.format_exc())
    try:
        from datetime import date as _d
        if _d.today().weekday() == 0:
            arquivo_exec = config.arquivo_ultimo_status.replace(".xlsx", "_relatorio_executivo.html")
            if gerar_relatorio_executivo(config.arquivo_historico, arquivo_exec):
                logger.info(f"Relatório executivo disponível em: {arquivo_exec}")
    except Exception as e: logger.error(f"Erro ao gerar relatório executivo: {e}")
    logger.info(gerar_resumo(resultados))


def get_ip_local() -> str:
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM); s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]; s.close(); return ip
    except: return "localhost"


def iniciar_servidor(pasta: str, porta: int) -> None:
    import urllib.parse

    class Handler(http.server.SimpleHTTPRequestHandler):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, directory=pasta, **kwargs)
        def do_GET(self):
            # Redireciona para config se credenciais não estiverem configuradas
            if not config_completa() and self.path not in ("/config", "/salvar-config"):
                self.send_response(302); self.send_header("Location", "/config"); self.end_headers(); return

            if self.path == "/":
                self.send_response(302); self.send_header("Location", "/ultimo_status_gerado.html"); self.end_headers()
            elif self.path == "/config":
                # Página de configuração
                cfg = carregar_config()
                sel = {str(cfg.get("intervalo",3600)): "selected"}
                html = HTML_CONFIG.replace("__CLIENT_ID__", cfg.get("client_id",""))
                html = html.replace("__CLIENT_SECRET__", cfg.get("client_secret",""))
                html = html.replace("__PORTA__", str(cfg.get("porta",8888)))
                html = html.replace("__SEL_1H__",  sel.get("3600",""))
                html = html.replace("__SEL_2H__",  sel.get("7200",""))
                html = html.replace("__SEL_4H__",  sel.get("14400",""))
                html = html.replace("__SEL_8H__",  sel.get("28800",""))
                if config_completa():
                    html = html.replace("__STATUS__", '<div class="status-ok"><div class="status-dot"></div><span>Credenciais configuradas e ativas</span></div>')
                else:
                    html = html.replace("__STATUS__", "")
                body = html.encode("utf-8")
                self.send_response(200); self.send_header("Content-Type","text/html;charset=utf-8"); self.send_header("Content-Length",str(len(body))); self.end_headers(); self.wfile.write(body)
            elif self.path.startswith("/gerar-relatorio"):
                try:
                    params = urllib.parse.parse_qs(urllib.parse.urlparse(self.path).query)
                    inicio = params.get("inicio",[""])[0]; fim = params.get("fim",[""])[0]
                    if not inicio or not fim: raise ValueError("Datas inválidas")
                    nome_arquivo = f"relatorio_periodo_{inicio}_{fim}.html"
                    ok = gerar_relatorio_periodo(str(Path(pasta)/"historico_status.xlsx"), inicio, fim, str(Path(pasta)/nome_arquivo))
                    resp = json.dumps({"ok":ok,"arquivo":"/"+nome_arquivo}).encode()
                    self.send_response(200); self.send_header("Content-Type","application/json"); self.send_header("Content-Length",str(len(resp))); self.end_headers(); self.wfile.write(resp)
                except Exception as e:
                    resp = json.dumps({"ok":False,"erro":str(e)}).encode()
                    self.send_response(500); self.send_header("Content-Type","application/json"); self.send_header("Content-Length",str(len(resp))); self.end_headers(); self.wfile.write(resp)
            else: super().do_GET()
        def do_POST(self):
            if self.path == "/salvar-config":
                try:
                    length = int(self.headers.get("Content-Length", 0))
                    body = json.loads(self.rfile.read(length).decode("utf-8"))
                    client_id     = body.get("client_id","").strip()
                    client_secret = body.get("client_secret","").strip()
                    intervalo     = int(body.get("intervalo", 3600))
                    porta         = int(body.get("porta", 8888))

                    if not client_id or not client_secret:
                        raise ValueError("Client ID e Client Secret são obrigatórios")

                    # Valida credenciais na API FedEx antes de salvar
                    r = requests.post(
                        "https://apis.fedex.com/oauth/token",
                        headers={"Content-Type":"application/x-www-form-urlencoded"},
                        data={"grant_type":"client_credentials","client_id":client_id,"client_secret":client_secret},
                        timeout=15
                    )
                    if r.status_code != 200:
                        raise ValueError(f"Credenciais inválidas — verifique Client ID e Secret (HTTP {r.status_code})")

                    salvar_config({"client_id":client_id,"client_secret":client_secret,"intervalo":intervalo,"porta":porta})

                    # Atualiza config em memória sem reiniciar
                    import os; os.environ["FEDEX_CLIENT_ID"] = client_id; os.environ["FEDEX_CLIENT_SECRET"] = client_secret

                    resp = json.dumps({"ok":True}).encode()
                    self.send_response(200); self.send_header("Content-Type","application/json"); self.send_header("Content-Length",str(len(resp))); self.end_headers(); self.wfile.write(resp)
                except Exception as e:
                    resp = json.dumps({"ok":False,"erro":str(e)}).encode()
                    self.send_response(400); self.send_header("Content-Type","application/json"); self.send_header("Content-Length",str(len(resp))); self.end_headers(); self.wfile.write(resp)
            else:
                self.send_response(404); self.end_headers()

        def log_message(self, format, *args): pass

    def _run():
        with socketserver.TCPServer(("", porta), Handler) as httpd:
            httpd.allow_reuse_address = True; httpd.serve_forever()

    threading.Thread(target=_run, daemon=True).start()


if __name__ == "__main__":
    config = Config()
    pasta  = str(Path(config.arquivo_ultimo_status).parent.resolve())
    ip     = get_ip_local()
    iniciar_servidor(pasta, config.porta_servidor)

    def _print_banner():
        print(); print("=" * 55); print("  FEDEX TRACKER — SERVIDOR ATIVO"); print("=" * 55)
        print(f"  Seu computador : http://localhost:{config.porta_servidor}")
        print(f"  Rede local     : http://{ip}:{config.porta_servidor}")
        print("=" * 55); print("  Ctrl+C para parar."); print()

    _print_banner()

    # Se não configurado, mostra mensagem e aguarda
    if not config_completa():
        print("  ⚠  Credenciais não configuradas.")
        print(f"  Abra http://localhost:{config.porta_servidor}/config no navegador.")
        print("  O rastreamento iniciará automaticamente após a configuração.")
        print()

    while True:
        # Recarrega config a cada ciclo — pega mudanças feitas pela tela de config
        cfg_saved = carregar_config()
        if cfg_saved.get("client_id"):
            import os
            os.environ["FEDEX_CLIENT_ID"]     = cfg_saved["client_id"]
            os.environ["FEDEX_CLIENT_SECRET"] = cfg_saved.get("client_secret","")
            # Atualiza intervalo se mudou
            intervalo = cfg_saved.get("intervalo", 3600)
        else:
            intervalo = 3600

        try: executar()
        except Exception as e: logging.getLogger("fedex_tracker").error(f"Erro no ciclo: {e}")

        _print_banner()
        print(f"  Proxima consulta em {intervalo//60} min... ({datetime.now().strftime('%H:%M')})")
        time.sleep(intervalo)
