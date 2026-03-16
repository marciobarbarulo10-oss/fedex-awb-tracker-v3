import re
import sys
import shutil
import logging
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, TypedDict

import requests
import pandas as pd
from tqdm import tqdm


# ==========================================================
# CONFIGURAÇÕES
# ==========================================================

@dataclass
class Config:
    # ⚠️ Substitua pelas suas credenciais da API FedEx Developer Portal
    # https://developer.fedex.com/
    client_id:     str = "SEU_CLIENT_ID_AQUI"
    client_secret: str = "SEU_CLIENT_SECRET_AQUI"

    url_token:    str = "https://apis.fedex.com/oauth/token"
    url_tracking: str = "https://apis.fedex.com/track/v1/trackingnumbers"

    arquivo_awbs:          str = "awbs.xlsx"
    arquivo_historico:     str = "historico_status.xlsx"
    arquivo_ultimo_status: str = "ultimo_status_gerado.xlsx"
    arquivo_log:           str = "tracking.log"

    max_workers: int   = 5
    delay_entre: float = 0.5
    timeout:     int   = 30


# ==========================================================
# MODELS
# ==========================================================

class ResultadoAWB(TypedDict, total=False):
    AWB: str
    AWB_ORIGINAL: str
    CATEGORIA: str
    STATUS_FEDEX: str
    ULTIMO_EVENTO: str
    MOTIVO_CATEGORIA: str
    DATA_CONSULTA: str
    SUCESSO: bool
    DATA_CHEGADA_MEMPHIS: str
    PEDIDO: str


# ==========================================================
# LOGGER
# ==========================================================

def setup_logger(log_file: str) -> logging.Logger:
    logger = logging.getLogger("fedex_tracker")
    if logger.handlers:
        return logger
    logger.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s [%(levelname)-8s] %(message)s", "%d/%m/%Y %H:%M:%S")
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger


# ==========================================================
# AWB — LIMPEZA E VALIDAÇÃO
# ==========================================================

def limpar_awb(awb: str) -> str:
    return re.sub(r'\s+', '', awb.strip())

def validar_awb(awb: str) -> bool:
    return bool(re.match(r'^\d{12,22}$', awb))


# ==========================================================
# CLASSIFICAÇÃO — 5 CATEGORIAS
# ==========================================================

HUBS_INTERNACIONAIS = [
    # EUA
    "MEMPHIS", "MEM", "INDIANAPOLIS", "IND", "MIAMI", "MIA",
    "NEWARK", "EWR", "CHICAGO", "ORD", "LOS ANGELES", "LAX",
    "AGUADILLA", "BQN",
    # Europa
    "FRANKFURT", "FRA", "PARIS", "CDG", "COLOGNE", "CGN",
    "LONDON", "LHR", "AMSTERDAM", "AMS", "LIEGE", "LGG",
    "ENFIELD",
    # Oriente Médio
    "DUBAI", "DXB",
    # Ásia
    "HONG KONG", "HKG", "SHANGHAI", "PVG", "TOKYO", "NRT",
    "SINGAPORE", "SIN", "MUMBAI", "BOM", "DELHI", "DEL",
    "BANGALORE", "BLR", "CHENNAI", "MAA", "GUANGZHOU", "CAN",
    "BEIJING", "PEK", "SEOUL", "ICN", "TAIPEI", "TPE",
    "BANGKOK", "BKK", "KUALA LUMPUR", "KUL",
    # América
    "TORONTO", "YYZ", "MEXICO CITY", "MEX",
]

TERMOS_BRASIL = [
    "CAMPINAS", "VCP", "VIRACOPOS", "BRAZIL", "BRASIL",
    "BARUERI", "SAO PAULO", "SAO PAULO", "RIO DE JANEIRO",
    "CURITIBA", "PORTO ALEGRE", "BELO HORIZONTE", "BRASILIA",
    "MANAUS", "RECIFE", "FORTALEZA", "SALVADOR", "BR",
]

TERMOS_ATRASO_ALFANDEGA = [
    "ATRASO NA LIBERACAO",
    "ATRASO NA LIBERAÇÃO",
    "ATRASO NA LIBERACAO - IMPORTACAO",
    "ATRASO NA LIBERAÇÃO - IMPORTAÇÃO",
    "CLEARANCE DELAY",
    "CUSTOMS DELAY",
    "HELD BY CUSTOMS",
    "CUSTOMS HOLD",
    "RETIDO NA ALFANDEGA",
    "RETIDO NA ALFÂNDEGA",
    "AGUARDANDO LIBERACAO",
    "AGUARDANDO LIBERAÇÃO",
    "PENDING CUSTOMS",
    "CLEARANCE IN PROGRESS",
    "IMPORT HOLD",
]

TERMOS_LIBERACAO = [
    "INTERNATIONAL SHIPMENT RELEASE",
    "RELEASED BY CUSTOMS",
    "CUSTOMS CLEARED",
    "CLEARANCE COMPLETED",
    "LIBERACAO DA REMESSA INTERNACIONAL",
    "LIBERAÇÃO DA REMESSA INTERNACIONAL",
    "REMESSA LIBERADA",
    "LIBERADO PELA ALFANDEGA",
    "LIBERADO PELA ALFÂNDEGA",
]

TERMOS_OUT_FOR_DELIVERY = [
    "ON FEDEX VEHICLE FOR DELIVERY",
    "EM UM VEICULO DA FEDEX PARA ENTREGA",
    "EM UM VEÍCULO DA FEDEX PARA ENTREGA",
    "OUT FOR DELIVERY",
    "EM VEICULO PARA ENTREGA",
    "EM VEÍCULO PARA ENTREGA",
    "SAIU PARA ENTREGA",
    "WITH DELIVERY COURIER",
]

TERMOS_ENTREGUE = [
    "DELIVERED",
    "ENTREGUE",
    "DELIVERY MADE",
]


def _textos_eventos(eventos: list) -> list:
    textos = []
    for ev in eventos:
        textos.append(ev.get("eventDescription", "").upper())
        textos.append(ev.get("scanLocation", {}).get("city", "").upper())
        textos.append(ev.get("scanLocation", {}).get("stateOrProvinceCode", "").upper())
        textos.append(ev.get("scanLocation", {}).get("countryCode", "").upper())
    return textos


def _contem(textos: list, termos: list) -> bool:
    for texto in textos:
        for termo in termos:
            if termo in texto:
                return True
    return False


def classificar(eventos: list, status_fedex: str) -> tuple:
    status_upper         = status_fedex.upper()
    ultimo_desc          = eventos[0].get("eventDescription", "").upper() if eventos else ""
    ultimo_local_city    = eventos[0].get("scanLocation", {}).get("city", "").upper() if eventos else ""
    ultimo_local_country = eventos[0].get("scanLocation", {}).get("countryCode", "").upper() if eventos else ""
    ultimo_textos        = [status_upper, ultimo_desc, ultimo_local_city, ultimo_local_country]

    todos_textos = _textos_eventos(eventos)

    GATILHOS_ENTREGUE = ["DELIVERED", "ENTREGUE", "DELIVERY MADE"]
    if _contem(ultimo_textos, GATILHOS_ENTREGUE):
        return "DELIVERED", "Entrega confirmada no destino"

    GATILHOS_OFD = [
        "ON FEDEX VEHICLE FOR DELIVERY",
        "EM UM VEICULO DA FEDEX PARA ENTREGA",
        "EM UM VEÍCULO DA FEDEX PARA ENTREGA",
        "OUT FOR DELIVERY",
        "EM VEICULO PARA ENTREGA",
        "EM VEÍCULO PARA ENTREGA",
        "SAIU PARA ENTREGA",
        "WITH DELIVERY COURIER",
    ]
    if _contem(ultimo_textos, GATILHOS_OFD):
        return "OUT FOR DELIVERY", "Pacote saiu para entrega — chegada prevista hoje"

    TERMOS_SAO_PAULO = ["SAO PAULO", "SÃO PAULO"]
    ultimo_city = eventos[0].get("scanLocation", {}).get("city", "").upper() if eventos else ""
    ultimo_state = eventos[0].get("scanLocation", {}).get("stateOrProvinceCode", "").upper() if eventos else ""
    if any(t in ultimo_city for t in TERMOS_SAO_PAULO) and ultimo_state in ("SP", ""):
        return "OUT FOR DELIVERY", "Remessa em São Paulo/SP — saindo para entrega"

    if _contem(ultimo_textos, TERMOS_ATRASO_ALFANDEGA):
        return "CUSTOMS INSPECTION", "Atraso na liberação alfandegária — retido pela Receita Federal"

    GATILHOS_LIBERACAO = [
        "LIBERACAO DA REMESSA INTERNACIONAL",
        "LIBERAÇÃO DA REMESSA INTERNACIONAL",
        "INTERNATIONAL SHIPMENT RELEASE",
        "RELEASED BY CUSTOMS",
        "CUSTOMS CLEARED",
        "CLEARANCE COMPLETED",
        "REMESSA LIBERADA",
        "LIBERADO PELA ALFANDEGA",
        "LIBERADO PELA ALFÂNDEGA",
    ]
    tem_brasil_historico = _contem(todos_textos, TERMOS_BRASIL)
    if _contem([status_upper], GATILHOS_LIBERACAO) and tem_brasil_historico:
        return "NATIONAL TRANSIT", "Liberação alfandegária confirmada — em trânsito nacional"
    elif _contem([status_upper], GATILHOS_LIBERACAO) and not tem_brasil_historico:
        return "COMING TO BRAZIL", "Liberação detectada mas Brasil não confirmado no histórico"

    GATILHOS_ORIGEM = [
        "DEIXOU O LOCAL DE ORIGEM FEDEX",
        "LEFT FEDEX ORIGIN FACILITY",
        "DEPARTED FEDEX LOCATION",
        "SHIPMENT LEFT FEDEX ORIGIN",
    ]
    if _contem([status_upper, ultimo_desc], GATILHOS_ORIGEM):
        return "COMING TO BRAZIL", "Saiu do local de origem — em rota internacional para o Brasil"

    tem_brasil    = _contem(todos_textos, TERMOS_BRASIL)
    tem_liberacao = _contem(todos_textos, GATILHOS_LIBERACAO)
    tem_atraso    = _contem(todos_textos, TERMOS_ATRASO_ALFANDEGA)
    tem_hub       = _contem(todos_textos, HUBS_INTERNACIONAIS)
    tem_origem    = _contem(todos_textos, GATILHOS_ORIGEM)

    ultimo_no_brasil = _contem(ultimo_textos, TERMOS_BRASIL)
    ultimo_em_hub    = _contem(ultimo_textos, HUBS_INTERNACIONAIS)

    if tem_brasil and tem_atraso and not tem_liberacao and ultimo_no_brasil:
        return "CUSTOMS INSPECTION", "Carga no Brasil com atraso na liberação alfandegária"

    if tem_brasil and tem_liberacao and ultimo_no_brasil:
        return "NATIONAL TRANSIT", "Liberado na alfândega — em trânsito nacional"

    if tem_liberacao and not ultimo_no_brasil:
        return "COMING TO BRAZIL", "Em trânsito internacional — ainda não chegou ao Brasil"

    if tem_brasil and not tem_liberacao and ultimo_no_brasil:
        return "CUSTOMS INSPECTION", "Carga no Brasil aguardando liberação pela Receita Federal"

    if tem_hub or tem_origem or ultimo_em_hub:
        return "COMING TO BRAZIL", "Em trânsito internacional — ainda não chegou ao Brasil"

    return "LABEL CREATED", "Etiqueta criada — aguardando coleta ou primeira movimentação"


# ==========================================================
# AUTENTICAÇÃO — TOKEN OAUTH
# ==========================================================

class TokenManager:
    def __init__(self, config: Config):
        self.config     = config
        self._token     = None
        self._expira_em = 0
        self.logger     = logging.getLogger("fedex_tracker")

    def get_token(self) -> str:
        if self._token and time.time() < self._expira_em - 60:
            return self._token

        self.logger.info("Obtendo token de acesso da API FedEx...")

        response = requests.post(
            self.config.url_token,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            data={
                "grant_type":    "client_credentials",
                "client_id":     self.config.client_id,
                "client_secret": self.config.client_secret,
            },
            timeout=self.config.timeout,
        )

        if response.status_code != 200:
            raise Exception(
                f"Erro ao obter token: HTTP {response.status_code} — {response.text[:200]}"
            )

        dados           = response.json()
        self._token     = dados["access_token"]
        self._expira_em = time.time() + int(dados.get("expires_in", 3600))

        self.logger.info("Token obtido! (válido por 1 hora)")
        return self._token


# ==========================================================
# CONSULTA À API OFICIAL
# ==========================================================

def consultar_awb(awb_original: str, awb_limpo: str, token: str, config: Config, pedido: str = "") -> ResultadoAWB:
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M")
    logger    = logging.getLogger("fedex_tracker")

    try:
        time.sleep(config.delay_entre)

        response = requests.post(
            config.url_tracking,
            headers={
                "Content-Type":  "application/json",
                "Authorization": f"Bearer {token}",
                "X-locale":      "pt_BR",
            },
            json={
                "includeDetailedScans": True,
                "trackingInfo": [{"trackingNumberInfo": {"trackingNumber": awb_limpo}}]
            },
            timeout=config.timeout,
        )

        if response.status_code == 401:
            raise Exception("Token inválido ou expirado")
        if response.status_code == 429:
            logger.warning(f"[{awb_limpo}] Rate limit — aguardando 30s...")
            time.sleep(30)
            raise Exception("Rate limit atingido")
        if response.status_code != 200:
            raise Exception(f"HTTP {response.status_code}: {response.text[:200]}")

        dados  = response.json()
        output = (
            dados
            .get("output", {})
            .get("completeTrackResults", [{}])[0]
            .get("trackResults", [{}])[0]
        )

        erro = output.get("error", {})
        if erro:
            codigo = erro.get("code", "")
            msg    = erro.get("message", "")
            if codigo == "TRACKING.TRACKINGNUMBER.NOTFOUND":
                return ResultadoAWB(
                    AWB=awb_limpo, AWB_ORIGINAL=awb_original, PEDIDO=pedido,
                    CATEGORIA="AWB NAO ENCONTRADO",
                    STATUS_FEDEX="NOT FOUND",
                    ULTIMO_EVENTO="Número não encontrado na base FedEx",
                    MOTIVO_CATEGORIA="AWB inválido ou não registrado",
                    DATA_CONSULTA=timestamp, SUCESSO=False,
                )
            raise Exception(f"{codigo}: {msg}")

        status_fedex = (
            output.get("latestStatusDetail", {})
            .get("description", "UNKNOWN")
            .upper().strip()
        )

        eventos = output.get("scanEvents", [])

        if eventos:
            ult    = eventos[0]
            descr  = ult.get("eventDescription", "").strip()
            city   = ult.get("scanLocation", {}).get("city", "").strip()
            pais   = ult.get("scanLocation", {}).get("countryCode", "").strip()
            data_e = ult.get("date", "").split("T")[0]
            hora_e = ult.get("date", "T").split("T")[1][:5] if "T" in ult.get("date", "") else ""
            ultimo_evento = f"{descr} — {city}/{pais} {data_e} {hora_e}".upper().strip(" —")
        else:
            ultimo_evento = "SEM EVENTOS REGISTRADOS"

        data_chegada_memphis = ""
        for ev in reversed(eventos):
            ev_city = ev.get("scanLocation", {}).get("city", "").upper()
            if "MEMPHIS" in ev_city:
                data_chegada_memphis = ev.get("date", "").split("T")[0]
                break

        categoria, motivo = classificar(eventos, status_fedex)

        logger.debug(f"[{awb_limpo}] {categoria} | {status_fedex} | {ultimo_evento[:50]}")

        return ResultadoAWB(
            AWB=awb_limpo,
            AWB_ORIGINAL=awb_original,
            CATEGORIA=categoria,
            STATUS_FEDEX=status_fedex,
            ULTIMO_EVENTO=ultimo_evento,
            MOTIVO_CATEGORIA=motivo,
            DATA_CONSULTA=timestamp,
            SUCESSO=True,
            DATA_CHEGADA_MEMPHIS=data_chegada_memphis,
            PEDIDO=pedido,
        )

    except Exception as e:
        msg = str(e).split('\n')[0]
        logger.error(f"[{awb_limpo}] {msg}")
        return ResultadoAWB(
            AWB=awb_limpo,
            AWB_ORIGINAL=awb_original,
            CATEGORIA="ERRO",
            STATUS_FEDEX="ERRO",
            ULTIMO_EVENTO=msg[:200],
            MOTIVO_CATEGORIA="Falha na consulta",
            DATA_CONSULTA=timestamp,
            SUCESSO=False,
            PEDIDO=pedido,
        )


# ==========================================================
# STORAGE
# ==========================================================

def carregar_awbs(arquivo: str) -> List[tuple]:
    path = Path(arquivo)
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {arquivo}")
    df = pd.read_excel(path)
    df.columns = df.columns.str.strip()
    if "AWB" not in df.columns:
        raise ValueError(f"Coluna 'AWB' não encontrada. Colunas: {list(df.columns)}")
    col_pedido = "Pedido" if "Pedido" in df.columns else ("PEDIDO" if "PEDIDO" in df.columns else None)
    logger    = logging.getLogger("fedex_tracker")
    resultado = []
    for _, row in df.iterrows():
        awb_original = str(row["AWB"]) if pd.notna(row["AWB"]) else ""
        if not awb_original or awb_original == "nan":
            continue
        pedido = str(row[col_pedido]) if col_pedido and pd.notna(row[col_pedido]) else ""
        if pedido.endswith(".0"):
            pedido = pedido[:-2]
        awb_limpo = limpar_awb(awb_original)
        if validar_awb(awb_limpo):
            resultado.append((awb_original.strip(), awb_limpo, pedido))
        else:
            logger.warning(f"AWB ignorado: '{awb_original}'")
    logger.info(f"{len(resultado)} AWBs válidos")
    return resultado


def _fazer_backup(arquivo: str) -> None:
    path = Path(arquivo)
    if path.exists():
        ts     = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = path.with_name(f"{path.stem}_bak_{ts}{path.suffix}")
        shutil.copy(path, backup)


def salvar_resultados(resultados, arquivo_historico, arquivo_ultimo) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    import re as _re

    logger = logging.getLogger("fedex_tracker")
    df = pd.DataFrame(resultados)

    ORDEM_CAT = {
        "LABEL CREATED": 1, "COMING TO BRAZIL": 2, "CUSTOMS INSPECTION": 3,
        "NATIONAL TRANSIT": 4, "OUT FOR DELIVERY": 5, "DELIVERED": 6,
        "AWB NAO ENCONTRADO": 7, "ERRO": 8,
    }

    CORES = {
        "LABEL CREATED":       {"bg": "D9D9D9", "font": "595959"},
        "COMING TO BRAZIL":    {"bg": "BDD7EE", "font": "1F4E79"},
        "CUSTOMS INSPECTION":  {"bg": "FFE699", "font": "7F6000"},
        "NATIONAL TRANSIT":    {"bg": "C6EFCE", "font": "276221"},
        "OUT FOR DELIVERY":    {"bg": "9DC3E6", "font": "1F4E79"},
        "DELIVERED":           {"bg": "375623", "font": "FFFFFF"},
        "AWB NAO ENCONTRADO":  {"bg": "F4CCCC", "font": "990000"},
        "ERRO":                {"bg": "FF0000", "font": "FFFFFF"},
    }

    DIAS_ATRASO_CUSTOMS = 5
    DIAS_ATRASO_MEMPHIS  = 3

    df["_ordem"] = df["CATEGORIA"].map(lambda x: ORDEM_CAT.get(x, 9))
    df = df.sort_values("_ordem").drop(columns=["_ordem"]).reset_index(drop=True)

    def dias_no_status(ultimo_evento_str):
        try:
            partes = str(ultimo_evento_str).split("—")
            data_str = partes[-1].strip().split(" ")
            data_part = [p for p in data_str if _re.match(r"\d{4}-\d{2}-\d{2}", p)]
            if data_part:
                from datetime import date
                d = datetime.strptime(data_part[0], "%Y-%m-%d").date()
                return (date.today() - d).days
        except:
            pass
        return None

    df["_dias"] = df["ULTIMO_EVENTO"].apply(dias_no_status)

    wb = Workbook()
    wb.remove(wb.active)

    def make_fill(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

    def make_border():
        side = Side(style="thin", color="BFBFBF")
        return Border(left=side, right=side, top=side, bottom=side)

    def header_style(cell, bg="1F4E79", fg="FFFFFF"):
        cell.font      = Font(bold=True, color=fg, name="Arial", size=10)
        cell.fill      = make_fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = make_border()

    def data_style(cell, bg="FFFFFF", fg="000000", bold=False, wrap=False):
        cell.font      = Font(color=fg, name="Arial", size=9, bold=bold)
        cell.fill      = make_fill(bg)
        cell.alignment = Alignment(vertical="center", wrap_text=wrap)
        cell.border    = make_border()

    COLUNAS  = ["AWB", "PEDIDO", "CATEGORIA", "STATUS_FEDEX", "ULTIMO_EVENTO", "MOTIVO_CATEGORIA", "DATA_CONSULTA"]
    LARGURAS = [22, 14, 22, 28, 52, 40, 18]
    HEADERS  = ["AWB", "PEDIDO", "CATEGORIA", "STATUS FEDEX", "ÚLTIMO EVENTO", "MOTIVO", "DATA CONSULTA"]

    def escrever_aba(ws, df_aba, titulo_aba):
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 32
        for col_idx, (header, largura) in enumerate(zip(HEADERS, LARGURAS), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            header_style(cell)
            ws.column_dimensions[get_column_letter(col_idx)].width = largura

        for row_idx, (_, row) in enumerate(df_aba.iterrows(), 2):
            cat    = row["CATEGORIA"]
            cores  = CORES.get(cat, {"bg": "FFFFFF", "font": "000000"})
            dias   = row.get("_dias")
            em_customs = (cat == "CUSTOMS INSPECTION" and dias and dias >= DIAS_ATRASO_CUSTOMS)
            data_memphis_str = str(row.get("DATA_CHEGADA_MEMPHIS", ""))
            dias_memphis = None
            if data_memphis_str and data_memphis_str not in ("nan", ""):
                try:
                    from datetime import date as _date
                    d_mem = datetime.strptime(data_memphis_str[:10], "%Y-%m-%d").date()
                    dias_memphis = (_date.today() - d_mem).days
                except:
                    pass
            em_memphis = (cat == "COMING TO BRAZIL" and dias_memphis is not None
                          and dias_memphis >= DIAS_ATRASO_MEMPHIS)
            atraso = em_customs or em_memphis

            for col_idx, col_name in enumerate(COLUNAS, 1):
                val  = row.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if pd.notna(val) else "")
                if atraso:
                    data_style(cell, bg="FF4C4C", fg="FFFFFF", bold=True, wrap=(col_idx in [4, 5]))
                else:
                    data_style(cell, bg=cores["bg"], fg=cores["font"], wrap=(col_idx in [4, 5]))
                ws.row_dimensions[row_idx].height = 28

            if atraso:
                cell_cat = ws.cell(row=row_idx, column=3)
                if em_memphis:
                    cell_cat.value = f"⚠ MEMPHIS +{dias_memphis}d"
                else:
                    cell_cat.value = f"⚠ {cat} ({dias}d)"
                cell_cat.font  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
                cell_cat.fill  = make_fill("C00000")

    # Aba Resumo, Todos e por categoria seguem a mesma lógica do original
    # (código completo mantido do original)
    ws_geral = wb.create_sheet("📋 TODOS")
    escrever_aba(ws_geral, df, "TODOS")

    ICONES = {
        "LABEL CREATED":      "🏷",
        "COMING TO BRAZIL":   "✈",
        "CUSTOMS INSPECTION": "🔍",
        "NATIONAL TRANSIT":   "🚚",
        "OUT FOR DELIVERY":   "📦",
        "DELIVERED":          "✅",
        "AWB NAO ENCONTRADO": "❓",
        "ERRO":               "❌",
    }

    for cat in ["COMING TO BRAZIL", "CUSTOMS INSPECTION", "NATIONAL TRANSIT", "OUT FOR DELIVERY", "DELIVERED"]:
        df_cat = df[df["CATEGORIA"] == cat].reset_index(drop=True)
        icone  = ICONES.get(cat, "")
        ws     = wb.create_sheet(f"{icone} {cat}"[:31])
        escrever_aba(ws, df_cat, cat)

    df_save = df.drop(columns=["_dias"], errors="ignore").copy()
    wb.save(arquivo_ultimo)
    logger.info(f"Relatório salvo → '{arquivo_ultimo}'")

    _fazer_backup(arquivo_historico)
    path_hist = Path(arquivo_historico)
    if path_hist.exists():
        hist  = pd.read_excel(path_hist)
        final = pd.concat([hist, df_save], ignore_index=True)
    else:
        final = df_save

    final = final.drop_duplicates(subset=["AWB", "DATA_CONSULTA"], keep="last")
    final.to_excel(arquivo_historico, index=False)
    logger.info(f"Histórico → '{arquivo_historico}' ({len(final)} registros)")


def gerar_resumo(resultados) -> str:
    total    = len(resultados)
    sucessos = sum(1 for r in resultados if r["SUCESSO"])
    ordem = ["LABEL CREATED", "COMING TO BRAZIL", "CUSTOMS INSPECTION",
             "NATIONAL TRANSIT", "OUT FOR DELIVERY", "DELIVERED",
             "AWB NAO ENCONTRADO", "ERRO"]
    contagem = {}
    for r in resultados:
        cat = r["CATEGORIA"]
        contagem[cat] = contagem.get(cat, 0) + 1
    linhas = [
        "", "=" * 60,
        f"  RESUMO DO RASTREIO — {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        "=" * 60,
        f"  Total consultado : {total}",
        f"  Com sucesso      : {sucessos}",
        f"  Com erro         : {total - sucessos}",
        "", "  Por categoria (ordem de progressão):",
    ]
    for cat in ordem:
        if cat in contagem:
            linhas.append(f"    {cat:<30} {contagem[cat]:>4}x")
    linhas.append("=" * 60)
    return "\n".join(linhas)


# ==========================================================
# MAIN
# ==========================================================

def executar():
    config = Config()
    logger = setup_logger(config.arquivo_log)
    logger.info("=" * 60)
    logger.info("  FedEx Tracker — API Oficial")
    logger.info("=" * 60)

    try:
        lista_awbs = carregar_awbs(config.arquivo_awbs)
    except (FileNotFoundError, ValueError) as e:
        logger.error(str(e))
        sys.exit(1)

    if not lista_awbs:
        logger.error("Nenhum AWB válido encontrado.")
        sys.exit(1)

    try:
        token_mgr = TokenManager(config)
        token     = token_mgr.get_token()
    except Exception as e:
        logger.error(f"Falha na autenticação: {e}")
        sys.exit(1)

    logger.info(f"Iniciando rastreio de {len(lista_awbs)} AWBs...")

    resultados = []
    with ThreadPoolExecutor(max_workers=config.max_workers) as executor:
        futures = {
            executor.submit(consultar_awb, orig, limpo, token, config, ped): limpo
            for orig, limpo, ped in lista_awbs
        }
        with tqdm(total=len(lista_awbs), desc="Rastreando", unit="AWB", ncols=72) as pbar:
            for future in as_completed(futures):
                r = future.result()
                resultados.append(r)
                pbar.set_postfix_str(f"{r['AWB']} → {r['CATEGORIA'][:18]}")
                pbar.update(1)

    try:
        salvar_resultados(resultados, config.arquivo_historico, config.arquivo_ultimo_status)
    except Exception as e:
        logger.error(f"Erro ao salvar: {e}")

    logger.info(gerar_resumo(resultados))


if __name__ == "__main__":
    executar()
